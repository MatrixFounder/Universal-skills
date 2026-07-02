#!/usr/bin/env python3
"""Deterministic script-grader for the xlsx recalc evals.

Pure function: no LLM, no network, no DB, no eval/exec/shell. Same
inputs → same grading.json, which is what makes pinning possible
(guide §7.2 / advanced-eval-patterns §1, §3).

Verdict parity (§7.1): the "was the workbook really recalculated" step
IMPORTS and CALLS the production gate `xlsx_recalc.verify_cached_values`
— the exact function the shipped CLI runs — instead of re-implementing
cache detection here. If production tightens or fixes the gate, this
grader follows automatically; it cannot drift.

Usage:
    # grade one arm of one case
    ./grade.py --case X-01 --run-dir <dir with the produced output file>

    # grade a whole workspace (all cases × arms) + aggregate
    ./grade.py --workspace <ws>   # expects <ws>/<case>/<arm>/outputs/

    # re-verify a pinned report (CI drift guard)
    ./grade.py --verify-pin <ws> reports/benchmark-v1.json

    # self-test the grader on synthetic outputs (RED/GREEN)
    ./grade.py --selftest

Exit codes: 0 = graded (or pin matches / selftest green);
2 = pin mismatch or selftest failure; 1 = usage/IO error.
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

HERE = Path(__file__).resolve().parent
SCRIPTS = HERE.parent / "scripts"  # skills/xlsx/scripts
sys.path.insert(0, str(SCRIPTS))

from openpyxl import load_workbook  # noqa: E402

# Production gate — IMPORTED, never copied (guide §7.1).
from xlsx_recalc import RecalcVerificationError, verify_cached_values  # noqa: E402

EVALS_FILE = HERE / "evals-v1.json"
ARMS = ("with_skill", "without_skill")


def _load_cases() -> dict:
    data = json.loads(EVALS_FILE.read_text(encoding="utf-8"))
    return {case["id"]: case for case in data["evals"]}


def _cell(wb_values, wb_formulas, ref: str):
    """Return (cached_value, formula_view_value, data_type) for 'Sheet!A1'."""
    sheet, coord = ref.split("!", 1)
    return (
        wb_values[sheet][coord].value,
        wb_formulas[sheet][coord].value,
        wb_values[sheet][coord].data_type,
    )


def grade_case(case: dict, run_dir: Path) -> dict:
    """Grade one produced output against one case spec. Pure."""
    checks: list[dict] = []

    def check(name: str, passed: bool, evidence: str) -> None:
        checks.append({"check": name, "passed": bool(passed), "evidence": evidence})

    out_path = run_dir / case["output_file"]
    if not out_path.is_file():
        check("output_exists", False, f"missing: {out_path.name}")
        return {"case": case["id"], "checks": checks, "passed": False}
    check("output_exists", True, out_path.name)

    machine = case.get("machine", {})

    # 1. Verdict parity: CALL the production gate.
    if machine.get("expect_gate_pass"):
        try:
            verify_cached_values(out_path)
            check("gate_verify_cached_values", True, "production gate passed")
        except RecalcVerificationError as exc:
            check("gate_verify_cached_values", False, str(exc)[:200])
        except Exception as exc:  # unreadable/corrupt output
            check("gate_verify_cached_values", False, f"{type(exc).__name__}: {exc}")
            return {"case": case["id"], "checks": checks, "passed": False}

    try:
        wb_v = load_workbook(str(out_path), data_only=True)
        wb_f = load_workbook(str(out_path))
    except Exception as exc:
        check("workbook_readable", False, f"{type(exc).__name__}: {exc}")
        return {"case": case["id"], "checks": checks, "passed": False}
    check("workbook_readable", True, "openpyxl opened both views")

    def safe_cell(ref: str):
        try:
            return _cell(wb_v, wb_f, ref)
        except Exception as exc:
            return ("<unreadable>", f"{type(exc).__name__}: {exc}", "?")

    # 2. Formulas preserved (recall: the contract is recalc, not flatten).
    for ref in machine.get("expect_formula_cells", []):
        _, formula, _ = safe_cell(ref)
        ok = isinstance(formula, str) and formula.startswith("=")
        check(f"formula_preserved:{ref}", ok, f"formula view = {formula!r}")

    # 3. Expected cached values (ground truth by construction).
    for ref, want in machine.get("expected_values", {}).items():
        got, _, _ = safe_cell(ref)
        ok = got == want or (
            isinstance(got, float) and isinstance(want, (int, float))
            and abs(got - want) < 1e-9
        )
        check(f"cached_value:{ref}", ok, f"want {want!r}, got {got!r}")

    # 4. Forbidden cached values (negative checks — guide §6.6).
    for ref, banned in machine.get("forbidden_values", {}).items():
        got, _, _ = safe_cell(ref)
        banned_norm = [None if b is None else b for b in banned]
        ok = got not in banned_norm
        check(f"forbidden_value:{ref}", ok, f"banned {banned!r}, got {got!r}")

    # 5. Seeded error cells: cached as Excel errors with the right token.
    for ref, want_err in machine.get("expected_error_cells", {}).items():
        got, _, dtype = safe_cell(ref)
        allowed = [want_err] if isinstance(want_err, str) else list(want_err)
        ok = dtype == "e" and got in allowed
        check(f"error_cell:{ref}", ok, f"want one of {allowed}, got {got!r} (type {dtype})")

    wb_v.close()
    wb_f.close()
    return {
        "case": case["id"],
        "checks": checks,
        "passed": all(c["passed"] for c in checks),
    }


def _aggregate(records: list[dict]) -> dict:
    by_arm: dict[str, dict] = {}
    for rec in records:
        arm = by_arm.setdefault(
            rec["arm"], {"cases": 0, "passed": 0, "checks": 0, "checks_passed": 0}
        )
        arm["cases"] += 1
        arm["passed"] += int(rec["passed"])
        arm["checks"] += len(rec["checks"])
        arm["checks_passed"] += sum(c["passed"] for c in rec["checks"])
    for arm in by_arm.values():
        arm["case_pass_rate"] = round(arm["passed"] / arm["cases"], 4) if arm["cases"] else None
        arm["check_pass_rate"] = (
            round(arm["checks_passed"] / arm["checks"], 4) if arm["checks"] else None
        )
    delta = None
    if "with_skill" in by_arm and "without_skill" in by_arm:
        a, b = by_arm["with_skill"], by_arm["without_skill"]
        if a["case_pass_rate"] is not None and b["case_pass_rate"] is not None:
            delta = round(a["case_pass_rate"] - b["case_pass_rate"], 4)
    return {"eval_set": "evals-v1.json", "skill": "xlsx", "arms": by_arm,
            "case_pass_rate_delta": delta}


def grade_workspace(ws: Path) -> dict:
    cases = _load_cases()
    records = []
    for case_id, case in sorted(cases.items()):
        for arm in ARMS:
            run_dir = ws / case_id / arm / "outputs"
            if not run_dir.is_dir():
                continue
            rec = grade_case(case, run_dir)
            rec["arm"] = arm
            records.append(rec)
            (ws / case_id / arm / "grading.json").write_text(
                json.dumps(rec, ensure_ascii=False, indent=2), encoding="utf-8"
            )
    return {"records": records, "aggregate": _aggregate(records)}


def verify_pin(ws: Path, pinned_path: Path) -> int:
    """Recompute the aggregate from the workspace and compare with the
    committed benchmark (pinning — guide §7.2). Volatile fields are not
    stored in the pin, so equality is exact."""
    got = grade_workspace(ws)["aggregate"]
    want = json.loads(pinned_path.read_text(encoding="utf-8"))["aggregate"]
    if got == want:
        print("PIN OK: aggregate matches", pinned_path)
        return 0
    print("PIN MISMATCH:\n want:", json.dumps(want, ensure_ascii=False),
          "\n got: ", json.dumps(got, ensure_ascii=False))
    return 2


def selftest() -> int:
    """RED/GREEN self-test on synthetic outputs: the grader must fail a
    wrong output and pass a correct one. Guards the grader itself."""
    import tempfile

    sys.path.insert(0, str(HERE / "fixtures"))
    from openpyxl import Workbook
    import zipfile as _zip

    case = _load_cases()["X-02"]
    failures = []
    with tempfile.TemporaryDirectory(prefix="grade-selftest-") as td:
        run = Path(td) / "outputs"
        run.mkdir()
        # RED: an openpyxl-only "fix" (formula rewritten, still uncached).
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws["A2"], ws["A3"], ws["A4"] = 10, 20, 30
        ws["A5"] = "=SUM(A2:A4)"
        wb.save(str(run / case["output_file"]))
        rec = grade_case(case, run)
        if rec["passed"]:
            failures.append("RED expected: uncached output must FAIL")
        # GREEN: same file with a correct cached value planted.
        src = run / case["output_file"]
        fixed = run / "_fixed.xlsx"
        zin = _zip.ZipFile(str(src))
        with _zip.ZipFile(str(fixed), "w", _zip.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == "xl/worksheets/sheet1.xml":
                    data = data.decode().replace(
                        "<f>SUM(A2:A4)</f><v></v>", "<f>SUM(A2:A4)</f><v>60</v>"
                    ).encode()
                zout.writestr(item, data)
        zin.close()
        fixed.replace(src)
        rec = grade_case(case, run)
        if not rec["passed"]:
            failures.append(
                "GREEN expected: cached-60 output must PASS, got "
                + json.dumps([c for c in rec["checks"] if not c["passed"]])
            )
    if failures:
        print("SELFTEST FAIL:", *failures, sep="\n  ")
        return 2
    print("SELFTEST OK (RED fails, GREEN passes)")
    return 0


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("--case")
    parser.add_argument("--run-dir", type=Path)
    parser.add_argument("--workspace", type=Path)
    parser.add_argument("--verify-pin", nargs=2, metavar=("WORKSPACE", "PINNED_JSON"))
    parser.add_argument("--selftest", action="store_true")
    args = parser.parse_args(argv)

    if args.selftest:
        return selftest()
    if args.verify_pin:
        return verify_pin(Path(args.verify_pin[0]), Path(args.verify_pin[1]))
    if args.workspace:
        result = grade_workspace(args.workspace)
        print(json.dumps(result, ensure_ascii=False, indent=2))
        return 0
    if args.case and args.run_dir:
        case = _load_cases().get(args.case)
        if case is None:
            print(f"unknown case: {args.case}", file=sys.stderr)
            return 1
        print(json.dumps(grade_case(case, args.run_dir), ensure_ascii=False, indent=2))
        return 0
    parser.print_help()
    return 1


if __name__ == "__main__":
    sys.exit(main())
