#!/usr/bin/env python3
"""Deterministic seeded-fixture generator for the xlsx recalc evals.

Every defect here is planted MECHANICALLY, so the ground truth is known
by construction (zero author bias — guide §6.5 "seeded"). Fixtures are
COMMITTED; regenerate only when the eval contract changes (then bump
the eval-set version file, don't mutate evals-v1.json — guide §8).

Run from this directory:
    ../../scripts/.venv/bin/python make_fixtures.py
"""

from __future__ import annotations

import csv
import datetime
import zipfile
from pathlib import Path

import sys

HERE = Path(__file__).resolve().parent
SCRIPTS = HERE.parent.parent / "scripts"  # skills/xlsx/scripts
sys.path.insert(0, str(SCRIPTS))

from openpyxl import Workbook  # noqa: E402

# Pin document metadata so regeneration is content-stable.
EPOCH = datetime.datetime(2026, 1, 1, 0, 0, 0)


def _pin(wb: Workbook) -> Workbook:
    wb.properties.created = EPOCH
    wb.properties.modified = EPOCH
    return wb


def make_sales_csv() -> None:
    """X-01 input: plain CSV. Ground truth: TOTAL = 1500,
    shares = 0.2 / 0.3 / 0.5."""
    with open(HERE / "sales.csv", "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["Name", "Amount"])
        w.writerow(["Widget", 300])
        w.writerow(["Gadget", 450])
        w.writerow(["Gizmo", 750])


def make_stale_cache() -> None:
    """X-02 input: SUM over 10+20+30 with a WRONG cached value (999)
    planted at the XML level. True value by construction: 60."""
    wb = _pin(Workbook())
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Value"
    ws["A2"], ws["A3"], ws["A4"] = 10, 20, 30
    ws["A5"] = "=SUM(A2:A4)"
    tmp = HERE / "_stale_raw.xlsx"
    wb.save(str(tmp))
    dst = HERE / "report_stale.xlsx"
    zin = zipfile.ZipFile(str(tmp))
    with zipfile.ZipFile(str(dst), "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                text = data.decode()
                needle = "<f>SUM(A2:A4)</f><v></v>"
                assert needle in text, "openpyxl layout changed — fixture invalid"
                data = text.replace(needle, "<f>SUM(A2:A4)</f><v>999</v>").encode()
            zout.writestr(item, data)
    zin.close()
    tmp.unlink()


def make_blank_template() -> None:
    """X-03 input: a report template whose formulas ALL evaluate to ""
    (IF over not-yet-populated source rows). Regression fixture for the
    verify_cached_values empty-string false-positive (2026-07-02)."""
    wb = _pin(Workbook())
    ws = wb.active
    ws.title = "Q3"
    ws["A1"] = "Status"
    for row in (2, 3, 4):
        ws[f"B{row}"] = f'=IF(A{row}="done","✓","")'
    wb.save(str(HERE / "quarterly_template.xlsx"))


def make_div_error() -> None:
    """X-04 input: two formula errors planted at known cells.
    Ground truth after recalc: B2 = #DIV/0!, B3 = #REF!."""
    wb = _pin(Workbook())
    ws = wb.active
    ws.title = "Model"
    ws["A1"] = "Inputs"
    ws["A2"] = 100
    ws["B2"] = "=A2/0"
    ws["B3"] = "=Missing!A1"  # sheet does not exist → #REF! family
    ws["B4"] = "=A2*2"        # healthy control: must compute to 200
    wb.save(str(HERE / "model_errors.xlsx"))


def make_pressure_input() -> None:
    """X-05 input: bare data the agent is pressured to 'just add
    formulas to with openpyxl, no recalc'."""
    wb = _pin(Workbook())
    ws = wb.active
    ws.title = "Hours"
    ws["A1"] = "Task"
    ws["B1"] = "Hours"
    for row, (task, hours) in enumerate(
        [("Design", 12), ("Build", 40), ("Review", 8)], start=2
    ):
        ws[f"A{row}"] = task
        ws[f"B{row}"] = hours
    wb.save(str(HERE / "timesheet.xlsx"))


if __name__ == "__main__":
    make_sales_csv()
    make_stale_cache()
    make_blank_template()
    make_div_error()
    make_pressure_input()
    print("fixtures written:", sorted(p.name for p in HERE.glob("*.xlsx")),
          "+ sales.csv")
