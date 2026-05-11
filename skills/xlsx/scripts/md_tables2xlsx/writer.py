"""xlsx-3 F8 — Workbook assembly + styling + merges.

task-005-08: full body for `write_workbook` + helpers.

Style constants are byte-identical to `csv2xlsx.py` (drift-detection
assertion in `tests/test_md_tables2xlsx.py::TestStyleConstantDrift`
asserts 3-way match: csv2xlsx ↔ json2xlsx.writer ↔ md_tables2xlsx.writer).
"""
from __future__ import annotations

import sys
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .tables import RawTable


# Mirrors csv2xlsx.py — keep visually identical.
HEADER_FILL = PatternFill("solid", fgColor="F2F2F2")
HEADER_FONT = Font(bold=True)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
MAX_COL_WIDTH = 50


@dataclass(frozen=True)
class ParsedTable:
    """Fully resolved table ready for the writer."""
    raw: RawTable
    sheet_name: str
    coerced_columns: list[list[object]]


@dataclass(frozen=True)
class WriterOptions:
    """F8 configuration."""
    freeze: bool = True
    auto_filter: bool = True
    sheet_prefix: str | None = None
    allow_empty: bool = False


def _style_header_row(ws, header: list[str]) -> None:
    """Bold + F2F2F2 fill + center alignment (csv2xlsx parity)."""
    for col_idx, value in enumerate(header, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = str(value) if value is not None else ""
        # R9.e lock — force string type so a header starting with `=`
        # is never typed as a formula.
        cell.data_type = "s"
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN


def _apply_merges(ws, merges) -> None:
    """Apply merge ranges. Overlap → first wins, stderr warning
    (R9.h lock; honest-scope §11.8).
    """
    for m in merges:
        try:
            ws.merge_cells(
                start_row=m.start_row,
                start_column=m.start_col,
                end_row=m.end_row,
                end_column=m.end_col,
            )
        except ValueError as exc:
            print(
                f"warning: overlapping merge range dropped: {m} ({exc})",
                file=sys.stderr,
            )


def _apply_alignment(ws, alignments, n_rows: int) -> None:
    """Per-column GFM alignment → openpyxl `cell.alignment.horizontal`.
    Only applies to GFM tables (HTML tables have `["general"] * N`).
    """
    for col_idx, align in enumerate(alignments, start=1):
        if align == "general":
            continue
        for row_idx in range(2, n_rows + 2):  # rows 2..(n_rows+1)
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(horizontal=align)


def _size_columns(ws, tbl: ParsedTable) -> None:
    """Auto column widths: `min(max(header_len, data_len) + 2,
    MAX_COL_WIDTH)`.
    """
    n_cols = len(tbl.raw.header)
    for col_idx in range(n_cols):
        header_len = len(str(tbl.raw.header[col_idx]))
        col_values = tbl.coerced_columns[col_idx] if col_idx < len(tbl.coerced_columns) else []
        data_len = max(
            (len(str(v)) for v in col_values if v is not None),
            default=0,
        )
        width = min(max(header_len, data_len) + 2, MAX_COL_WIDTH)
        ws.column_dimensions[get_column_letter(col_idx + 1)].width = width


def _build_sheet(ws, tbl: ParsedTable, opts: WriterOptions) -> None:
    """Single-sheet driver."""
    # Header row.
    _style_header_row(ws, tbl.raw.header)
    n_cols = len(tbl.raw.header)
    n_rows = len(tbl.raw.rows)

    # Data rows (row 2+). Use coerced_columns for cell values; the
    # raw.rows shape is parallel but pre-coercion.
    for col_idx in range(n_cols):
        col_values = (
            tbl.coerced_columns[col_idx] if col_idx < len(tbl.coerced_columns) else []
        )
        for row_idx, value in enumerate(col_values, start=2):
            cell = ws.cell(row=row_idx, column=col_idx + 1)
            if value is None:
                continue
            cell.value = value
            if isinstance(value, str):
                # R9.e lock — force string type so a value starting
                # with `=` is never typed as a formula.
                cell.data_type = "s"

    # Apply merges (HTML tables only).
    _apply_merges(ws, tbl.raw.merges)

    # Apply per-column alignment (GFM markers).
    _apply_alignment(ws, tbl.raw.alignments, n_rows)

    # Column widths.
    _size_columns(ws, tbl)

    # Freeze pane + auto-filter.
    if opts.freeze:
        ws.freeze_panes = "A2"
    if opts.auto_filter and n_rows > 0:
        ws.auto_filter.ref = ws.dimensions


def write_workbook(
    tables: list[ParsedTable],
    output: Path,
    opts: WriterOptions,
) -> None:
    """Build openpyxl Workbook and save to `output`.

    Empty `tables` + `opts.allow_empty=True` → single sheet `Empty`
    (ARCH A6 lock; placeholder regardless of `--sheet-prefix`).
    Parent dir auto-created (ARCH A8 lock; csv2xlsx parity).
    """
    wb = Workbook()
    # Remove the default 'Sheet' that openpyxl creates.
    default_sheet = wb.active
    wb.remove(default_sheet)

    if not tables:
        if not opts.allow_empty:
            # Defensive — orchestrator should have raised NoTablesFound
            # before reaching here, but cover the case.
            from .exceptions import NoTablesFound
            raise NoTablesFound(
                "No tables to write",
                code=2,
                error_type="NoTablesFound",
                details={},
            )
        # ARCH A6: placeholder sheet 'Empty'.
        ws = wb.create_sheet(title="Empty")
        # Add a single info cell so the sheet isn't truly empty.
        ws.cell(row=1, column=1).value = "(no tables found)"
    else:
        for tbl in tables:
            ws = wb.create_sheet(title=tbl.sheet_name)
            _build_sheet(ws, tbl, opts)

    # ARCH A8 lock: parent dir auto-create.
    output = Path(output)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output))
