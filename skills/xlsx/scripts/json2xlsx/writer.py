"""xlsx-2 workbook writer (F4).

Takes a `ParsedInput` (from F2 loaders) and produces a styled .xlsx
file on disk. Per-cell typing is routed through F3 (coerce). The
visual contract — bold header, light-grey fill, centre alignment,
freeze pane "A2", auto-filter, sized column widths — mirrors
csv2xlsx 1:1.

Style constants are COPIED (not imported) from csv2xlsx; the drift
assertion lives in `tests/test_json2xlsx.py::test_style_constants_drift_csv2xlsx`
(AQ-1 lock — accepts both 6-char and 8-char ARGB forms). Importing
csv2xlsx from production code would create a top-script cross-
dependency that the CLAUDE.md §"Независимость скиллов" principle
forbids.

Sheet-name Excel-rule validation owns the structural rules here
(R7.b) — F2 leaves names as-is.

Honest scope:
  §11.4 — Sheet-name auto-sanitization is OUT of scope v1. Invalid
  names hard-fail with `InvalidSheetName`. A `--sanitize-sheet-names`
  flag is a v2 candidate.
  §11.2 — Leading `=` in a JSON string value passes through to Excel
  as-is; the cell may render as a formula or as text per Excel's
  own heuristic. No `--escape-formulas` in v1 (joint v2 fix with
  csv2xlsx).
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore[import-untyped]

from .coerce import CellContext, CoerceOptions, coerce_cell
from .exceptions import InvalidSheetName
from .loaders import ParsedInput


# Mirrors csv2xlsx.py — keep visually identical.
# Drift detection: tests/test_json2xlsx.py::test_style_constants_drift_csv2xlsx
# (accepts both "F2F2F2" and "00F2F2F2" since openpyxl normalises 6-char
# fgColor to 8-char ARGB lazily on attribute access).
HEADER_FILL = PatternFill("solid", fgColor="F2F2F2")
HEADER_FONT = Font(bold=True)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
MAX_COL_WIDTH = 50


# Excel sheet-name rules — same set csv2xlsx omits but xlsx-2 enforces
# (csv2xlsx is single-sheet; multi-sheet input makes the rules load-
# bearing). Forbidden character set per ECMA-376 / Excel 2016+ docs.
_FORBIDDEN_SHEET_CHARS = frozenset("[]:*?/\\")
_RESERVED_SHEET_NAMES = frozenset({"history"})  # case-insensitive match
# VDD-multi Logic M1 + Security LOW-1 lock: Excel additionally rejects
# control chars (`\x00`-`\x1f`) and apostrophe (`'`) at first or last
# position. Without these guards a JSON sheet key like `"'foo'"` or
# `"benign‮exe.sh"` slips past xlsx-2 and either crashes openpyxl
# on title assignment or lands a Unicode RTL spoof in Excel's tab.
_CONTROL_CHAR_ORDS = frozenset(range(0x00, 0x20))


def write_workbook(
    parsed: ParsedInput,
    output: Path,
    *,
    freeze: bool = True,
    auto_filter: bool = True,
    sheet_override: str | None = None,
    coerce_opts: CoerceOptions,
) -> None:
    """Build the workbook in memory and save to `output`.

    Single-sheet shapes (`array_of_objects`, `jsonl`) write to
    `sheet_override` if provided, else the default "Sheet1" key in
    `parsed.sheets`.

    Multi-sheet shape (`multi_sheet_dict`) writes each top-level dict
    key as a sheet with the same name; `sheet_override` is IGNORED at
    this layer (CLI layer emits a stderr warning per R7.d).

    `parsed.sheets` is `dict[str, list[dict[str, Any]]]` (header
    invariant guaranteed by F2 `_validate_multi_sheet`).
    """
    wb = Workbook()
    default_ws = wb.active  # Workbook() creates one default sheet.

    first = True
    for raw_name, rows in parsed.sheets.items():
        if parsed.shape != "multi_sheet_dict" and sheet_override is not None:
            sheet_name = sheet_override
        else:
            sheet_name = raw_name
        _validate_sheet_name(sheet_name)

        if first:
            default_ws.title = sheet_name
            ws = default_ws
            first = False
        else:
            ws = wb.create_sheet(title=sheet_name)

        _build_sheet(ws, rows, coerce_opts, sheet_name)

        if freeze:
            ws.freeze_panes = "A2"
        if auto_filter and len(rows) > 0:
            ws.auto_filter.ref = ws.dimensions

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output))


def _build_sheet(
    ws: Worksheet,
    rows: list[dict[str, Any]],
    coerce_opts: CoerceOptions,
    sheet_name: str,
) -> None:
    headers = _union_headers(rows)

    # Header row — always strings. Force `data_type='s'` so a header
    # like "#REF!" is NOT classified by openpyxl as a cached formula
    # error (which xlsx_validate would then flag). Mirrors csv2xlsx.
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = str(header)
        cell.data_type = "s"
    _style_header_row(ws, len(headers))

    # Data rows.
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            if header not in row:
                # Missing key → empty cell (no value set).
                continue
            value = row[header]
            ctx = CellContext(
                sheet=sheet_name,
                row=row_idx - 1,  # 1-indexed data row (header excluded)
                column=get_column_letter(col_idx),
            )
            payload = coerce_cell(value, coerce_opts, ctx=ctx)
            if payload.value is None:
                # Explicit JSON null → empty cell (no value set).
                # R3.d locks the visual indistinguishability between
                # missing-key and null-value at the cell level.
                continue
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = payload.value
            if payload.number_format is not None:
                cell.number_format = payload.number_format
            if isinstance(payload.value, str):
                # Same defence as the header row: force 's' so e.g.
                # an agent-emitted value "#REF!" doesn't get reclassed
                # as data_type 'e'.
                cell.data_type = "s"

    _size_columns(ws, headers, rows)


def _union_headers(rows: list[dict[str, Any]]) -> list[str]:
    """First-seen wins (R5.b). Row 1's keys appear first; new keys
    from row 2 are appended in the order they appear; etc.

    Python's `dict` preserves insertion order (PEP 468 / 3.7+), so a
    plain dict-as-ordered-set does the job — no need for OrderedDict.
    """
    seen: dict[str, None] = {}
    for row in rows:
        for k in row.keys():
            if k not in seen:
                seen[k] = None
    return list(seen.keys())


def _style_header_row(ws: Worksheet, header_count: int) -> None:
    for col_idx in range(1, header_count + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN


def _size_columns(
    ws: Worksheet,
    headers: list[str],
    rows: list[dict[str, Any]],
) -> None:
    """Set column widths to min(max(header_len, max_value_len) + 2, MAX_COL_WIDTH).

    Signature matches ARCH §F4 exactly — no `coerce_opts` / `sheet_name`
    parameters. Column width is computed off the raw string length of
    the JSON value (not the coerced cell value); date-formatted Excel
    renders may differ slightly but the visual cap at 50 absorbs the
    difference.
    """
    for col_idx, header in enumerate(headers, start=1):
        header_len = len(str(header))
        max_val_len = 0
        for row in rows:
            v = row.get(header)
            if v is None:
                continue
            length = len(str(v))
            if length > max_val_len:
                max_val_len = length
        width = min(max(header_len, max_val_len) + 2, MAX_COL_WIDTH)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _validate_sheet_name(name: str) -> None:
    """Excel sheet-name rules — fail-loud per R7.b / honest-scope §11.4.

    Auto-sanitization (truncate/replace) is out of scope v1 because
    silently mutating user-supplied sheet keys is more dangerous than
    a clear error.

    Rule set (VDD-multi merged from Logic M1 + Security LOW-1):
      1. Non-empty.
      2. Length ≤ 31 chars.
      3. No `[]:*?/\\` characters (ECMA-376 forbidden set).
      4. No control characters (`\\x00`-`\\x1f` — break OOXML serialisation
         on some Excel builds; `\\u202E` RTL-override etc. are NOT
         covered by this gate, but are far less common and risk-rated
         lower than control chars).
      5. Apostrophe (`'`) not at first or last character — Excel
         hard-rejects on workbook load (the apostrophe is its
         single-quote sheet-name delimiter in formulas).
      6. Not case-insensitive `"history"` (Excel reserved).
    """
    if not name:
        raise InvalidSheetName(name=name, reason="empty")
    if len(name) > 31:
        raise InvalidSheetName(
            name=name,
            reason=f"length {len(name)} exceeds Excel limit of 31",
        )
    bad = sorted({c for c in name if c in _FORBIDDEN_SHEET_CHARS})
    if bad:
        raise InvalidSheetName(
            name=name,
            reason=f"contains forbidden character(s) {bad}",
        )
    ctrl_indices = [i for i, c in enumerate(name) if ord(c) in _CONTROL_CHAR_ORDS]
    if ctrl_indices:
        raise InvalidSheetName(
            name=name,
            reason=(
                f"contains control character(s) at offset(s) "
                f"{ctrl_indices} (Excel rejects \\x00-\\x1f)"
            ),
        )
    if name.startswith("'") or name.endswith("'"):
        raise InvalidSheetName(
            name=name,
            reason="apostrophe is not permitted at the start or end of a sheet name",
        )
    if name.lower() in _RESERVED_SHEET_NAMES:
        raise InvalidSheetName(
            name=name,
            reason="reserved sheet name (case-insensitive)",
        )


__all__ = [
    "HEADER_FILL",
    "HEADER_FONT",
    "HEADER_ALIGN",
    "MAX_COL_WIDTH",
    "write_workbook",
    "_validate_sheet_name",
]
