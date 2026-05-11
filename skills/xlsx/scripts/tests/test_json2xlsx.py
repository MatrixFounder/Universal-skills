"""Unit tests for `json2xlsx` (xlsx-2).

This module hosts one TestCase class per F-region per ARCHITECTURE
§2.1. In this task (004.02) every class carries:

  - A `test_smoke` method that imports the corresponding submodule.
    Smoke methods are NOT `@unittest.skip`'ed — they run and pass
    immediately because the 004.01 stubs are importable. This gives
    the suite an early positive signal before logic ships.

  - One representative `test_<feature>` method per major feature the
    F-region owns, decorated with `@unittest.skip("task-004-NN — not
    implemented")`. The owning task removes the skip + replaces the
    method body with real assertions.

Successive tasks (004.03–004.08) will:
  1. Remove the `@unittest.skip` decorator from the methods they own.
  2. Replace placeholder bodies with real assertions.
  3. Add additional methods as their feature surface grows.

The style-constant drift assertion (`test_style_constants_drift_csv2xlsx`)
is **NOT** skipped — it locks the AQ-1 contract live against csv2xlsx
and self-skips the writer-side mirror until 004.06 lands the constants.

The round-trip synthetic test (`test_synthetic_roundtrip`) is live
since 004.08 (writer + CLI + post-validate were all in place by then,
so the assertion against the golden JSON could be finalised). The
live round-trip test (`test_live_roundtrip`) is gated by
`@unittest.skipUnless(_xlsx2json_available(), …)` — AQ-5 lock —
and activates automatically once xlsx-8 lands.
"""
from __future__ import annotations

import importlib
import os
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

# Make `import json2xlsx`, `import csv2xlsx`, etc. resolve when the
# test runs from `skills/xlsx/scripts/tests/`.
_SCRIPTS_DIR = Path(__file__).resolve().parent.parent
if str(_SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS_DIR))


def _smoke_import(module_name: str) -> None:
    """Helper: import the named submodule. Raises ImportError on
    structural breakage; the caller's `test_smoke` then reports it.
    """
    importlib.import_module(module_name)


def _xlsx2json_available() -> bool:
    """AQ-5 lock — gates the live round-trip test. Returns True once
    xlsx-8 (`xlsx2json.py`) lands and is importable.
    """
    try:
        importlib.import_module("xlsx2json")
        return True
    except ImportError:
        return False


# ===========================================================================
# Loaders (F1 + F2, owned by 004.04)
# ===========================================================================
class TestLoaders(unittest.TestCase):
    """Tests for `json2xlsx.loaders` (F1 read_input + F2 detect_and_parse)."""

    def test_smoke(self) -> None:
        _smoke_import("json2xlsx.loaders")

    # ----- F1: read_input ---------------------------------------------------

    def test_read_input_file_utf8(self) -> None:
        from json2xlsx.loaders import read_input
        with tempfile.NamedTemporaryFile(mode="wb", suffix=".json", delete=False) as f:
            f.write(b'[{"a":1}]')
            path = f.name
        try:
            data, src = read_input(path)
            self.assertEqual(data, b'[{"a":1}]')
            self.assertEqual(src, path)
        finally:
            os.unlink(path)

    def test_read_input_stdin_dash(self) -> None:
        """`-` sentinel reads from `sys.stdin.buffer`. Use a subprocess
        for real pipe semantics (no monkeypatching)."""
        payload = b'[{"name":"Alice"}]'
        child = (
            "import sys, os; "
            "sys.path.insert(0, os.environ['XLSX2_SCRIPTS']); "
            "from json2xlsx.loaders import read_input; "
            "data, src = read_input('-'); "
            "sys.stdout.write(f'{len(data)}|{src}')"
        )
        proc = subprocess.run(
            [sys.executable, "-c", child],
            input=payload, capture_output=True, timeout=15,
            env={**os.environ, "XLSX2_SCRIPTS": str(_SCRIPTS_DIR)},
        )
        self.assertEqual(proc.returncode, 0, proc.stderr.decode("utf-8", "replace"))
        out = proc.stdout.decode()
        self.assertEqual(out, f"{len(payload)}|<stdin>")

    def test_read_input_file_not_found(self) -> None:
        from json2xlsx.loaders import read_input
        with self.assertRaises(FileNotFoundError):
            read_input("/nonexistent/path/to/file-that-does-not-exist.json")

    # ----- F2: detect_and_parse — shapes ------------------------------------

    def test_detect_array_of_objects(self) -> None:
        from json2xlsx.loaders import detect_and_parse
        parsed = detect_and_parse(
            b'[{"a": 1, "b": "x"}, {"a": 2, "b": "y"}]',
            source="t.json", is_jsonl_hint=False,
        )
        self.assertEqual(parsed.shape, "array_of_objects")
        self.assertEqual(list(parsed.sheets.keys()), ["Sheet1"])
        self.assertEqual(len(parsed.sheets["Sheet1"]), 2)
        self.assertEqual(parsed.sheets["Sheet1"][0], {"a": 1, "b": "x"})
        self.assertEqual(parsed.source_label, "t.json")

    def test_detect_multi_sheet_dict(self) -> None:
        from json2xlsx.loaders import detect_and_parse
        raw = b'{"A": [{"x": 1}], "B": [{"y": 2}, {"y": 3}]}'
        parsed = detect_and_parse(raw, source="m.json", is_jsonl_hint=False)
        self.assertEqual(parsed.shape, "multi_sheet_dict")
        self.assertEqual(list(parsed.sheets.keys()), ["A", "B"])
        self.assertEqual(parsed.sheets["A"], [{"x": 1}])
        self.assertEqual(parsed.sheets["B"], [{"y": 2}, {"y": 3}])

    def test_detect_jsonl_by_extension(self) -> None:
        from json2xlsx.loaders import detect_and_parse
        raw = b'{"a": 1}\n{"a": 2}\n'
        parsed = detect_and_parse(raw, source="t.jsonl", is_jsonl_hint=True)
        self.assertEqual(parsed.shape, "jsonl")
        self.assertEqual(list(parsed.sheets.keys()), ["Sheet1"])
        self.assertEqual(len(parsed.sheets["Sheet1"]), 2)

    # ----- F2: negative cases -----------------------------------------------

    def test_detect_unsupported_scalar(self) -> None:
        from json2xlsx.loaders import detect_and_parse
        from json2xlsx.exceptions import UnsupportedJsonShape
        with self.assertRaises(UnsupportedJsonShape) as cm:
            detect_and_parse(b'42', source="s.json", is_jsonl_hint=False)
        self.assertEqual(cm.exception.details["root_type"], "int")

    def test_detect_unsupported_list_of_lists(self) -> None:
        from json2xlsx.loaders import detect_and_parse
        from json2xlsx.exceptions import UnsupportedJsonShape
        with self.assertRaises(UnsupportedJsonShape) as cm:
            detect_and_parse(b'[[1,2],[3,4]]', source="l.json", is_jsonl_hint=False)
        self.assertEqual(cm.exception.details["root_type"], "list")
        self.assertEqual(cm.exception.details["first_element_type"], "list")

    def test_detect_empty_array_no_rows(self) -> None:
        from json2xlsx.loaders import detect_and_parse
        from json2xlsx.exceptions import NoRowsToWrite
        with self.assertRaises(NoRowsToWrite):
            detect_and_parse(b'[]', source="e.json", is_jsonl_hint=False)

    def test_jsonl_blank_line_tolerated(self) -> None:
        from json2xlsx.loaders import detect_and_parse
        raw = b'{"a": 1}\n\n   \n{"a": 2}\n'
        parsed = detect_and_parse(raw, source="t.jsonl", is_jsonl_hint=True)
        self.assertEqual(len(parsed.sheets["Sheet1"]), 2)

    def test_jsonl_malformed_line_reports_line_number(self) -> None:
        from json2xlsx.loaders import detect_and_parse
        from json2xlsx.exceptions import JsonDecodeError
        raw = b'{"a": 1}\n{not_json\n{"a": 3}\n'
        with self.assertRaises(JsonDecodeError) as cm:
            detect_and_parse(raw, source="t.jsonl", is_jsonl_hint=True)
        # Malformed line is line 2.
        self.assertEqual(cm.exception.details["line"], 2)

    # ----- Bonus coverage tests (F2 edge cases beyond the 11-tag inventory)
    # — pulled forward from honest-scope §11.5 + UC-2 alt scenarios.

    def test_jsonl_array_line_raises_unsupported_shape(self) -> None:
        """UC-3 A3 — JSONL line is an array, not an object."""
        from json2xlsx.loaders import detect_and_parse
        from json2xlsx.exceptions import UnsupportedJsonShape
        raw = b'{"a": 1}\n[1, 2, 3]\n'
        with self.assertRaises(UnsupportedJsonShape) as cm:
            detect_and_parse(raw, source="t.jsonl", is_jsonl_hint=True)
        self.assertIn("line 2", cm.exception.details["hint"])

    def test_multi_sheet_empty_sheet_attribution(self) -> None:
        """UC-2 A4 — multi-sheet input with an empty sheet → NoRowsToWrite
        with `details.empty_sheet` naming the offender."""
        from json2xlsx.loaders import detect_and_parse
        from json2xlsx.exceptions import NoRowsToWrite
        raw = b'{"A": [{"x": 1}], "B": []}'
        with self.assertRaises(NoRowsToWrite) as cm:
            detect_and_parse(raw, source="m.json", is_jsonl_hint=False)
        self.assertEqual(cm.exception.details.get("empty_sheet"), "B")

    def test_empty_input_raises(self) -> None:
        """R1.d — empty input string → EmptyInput envelope."""
        from json2xlsx.loaders import detect_and_parse
        from json2xlsx.exceptions import EmptyInput
        with self.assertRaises(EmptyInput) as cm:
            detect_and_parse(b'', source="empty.json", is_jsonl_hint=False)
        self.assertEqual(cm.exception.details["source"], "empty.json")
        with self.assertRaises(EmptyInput):
            detect_and_parse(b'   \n\t  ', source="ws.json", is_jsonl_hint=False)

    def test_invalid_json_top_level(self) -> None:
        """R1.e — JSON parse error reports line + column."""
        from json2xlsx.loaders import detect_and_parse
        from json2xlsx.exceptions import JsonDecodeError
        with self.assertRaises(JsonDecodeError) as cm:
            detect_and_parse(b'[{\n', source="t.json", is_jsonl_hint=False)
        self.assertGreaterEqual(cm.exception.details["line"], 1)
        self.assertGreaterEqual(cm.exception.details["column"], 1)

    # ----- VDD-multi Security M-1: invalid UTF-8 routes through envelope

    def test_invalid_utf8_raises_jsondecodeerror(self) -> None:
        """VDD-multi Security M-1: malformed UTF-8 bytes (e.g., latin-1
        / UTF-16-LE BOM) MUST surface through the typed `JsonDecodeError`
        envelope, NOT as an uncaught `UnicodeDecodeError` traceback
        (AQ-3 contract: every taxonomy error routes through report_error).
        """
        from json2xlsx.loaders import detect_and_parse
        from json2xlsx.exceptions import JsonDecodeError
        # Invalid lone-byte 0xff cannot appear in a UTF-8 stream.
        with self.assertRaises(JsonDecodeError) as cm:
            detect_and_parse(b"\xff\xfe garbage", source="bad.json", is_jsonl_hint=False)
        self.assertIn("UTF-8", cm.exception.details["msg"])
        self.assertEqual(cm.exception.details["line"], 1)

    def test_invalid_utf8_jsonl_line_reports_line_number(self) -> None:
        """VDD-multi Security M-1 on the JSONL path: malformed UTF-8
        on a specific line surfaces as JsonDecodeError with line N."""
        from json2xlsx.loaders import detect_and_parse
        from json2xlsx.exceptions import JsonDecodeError
        # line 1 OK, line 2 has invalid UTF-8.
        raw = b'{"a": 1}\n\xff\xfe\n{"a": 3}\n'
        with self.assertRaises(JsonDecodeError) as cm:
            detect_and_parse(raw, source="t.jsonl", is_jsonl_hint=True)
        self.assertEqual(cm.exception.details["line"], 2)

    # ----- VDD-multi Logic H2: NaN / Infinity rejection

    def test_json_nan_rejected_as_jsondecodeerror(self) -> None:
        """VDD-multi Logic H2: Python's json.loads silently accepts
        `NaN`/`Infinity`/`-Infinity` literals (non-strict mode). xlsx-2
        MUST reject them at decode time so they never reach openpyxl
        (where they would produce a workbook Excel renders as #NUM!
        or refuses to open).
        """
        from json2xlsx.loaders import detect_and_parse
        from json2xlsx.exceptions import JsonDecodeError
        for token in (b"NaN", b"Infinity", b"-Infinity"):
            payload = b'[{"v": ' + token + b'}]'
            with self.assertRaises(JsonDecodeError, msg=f"token={token!r}") as cm:
                detect_and_parse(payload, source="bad.json", is_jsonl_hint=False)
            self.assertIn("non-finite", cm.exception.details["msg"].lower())

    def test_jsonl_nan_rejected_with_line_number(self) -> None:
        """Same H2 guard on the JSONL path — reports the offending line."""
        from json2xlsx.loaders import detect_and_parse
        from json2xlsx.exceptions import JsonDecodeError
        raw = b'{"a": 1}\n{"a": NaN}\n'
        with self.assertRaises(JsonDecodeError) as cm:
            detect_and_parse(raw, source="t.jsonl", is_jsonl_hint=True)
        self.assertEqual(cm.exception.details["line"], 2)
        self.assertIn("non-finite", cm.exception.details["msg"].lower())


# ===========================================================================
# Coerce (F3, owned by 004.05)
# ===========================================================================
class TestCoerce(unittest.TestCase):
    """Tests for `json2xlsx.coerce` (F3 — per-cell type coercion + ISO-date
    heuristic + D7 --strict-dates rejection)."""

    CTX = None  # set in setUp

    def setUp(self) -> None:
        from json2xlsx.coerce import CellContext
        self.CTX = CellContext(sheet="Sheet1", row=1, column="A")

    def _opts(self, **kw):
        from json2xlsx.coerce import CoerceOptions
        return CoerceOptions(**kw)

    def test_smoke(self) -> None:
        _smoke_import("json2xlsx.coerce")

    # ----- type preservation -----------------------------------------------

    def test_coerce_int_to_int(self) -> None:
        from json2xlsx.coerce import coerce_cell
        p = coerce_cell(42, self._opts(), ctx=self.CTX)
        self.assertEqual(p.value, 42)
        self.assertIs(type(p.value), int)
        self.assertIsNone(p.number_format)

    def test_coerce_bool_to_bool_not_int(self) -> None:
        """ARCH §4.1 — bool MUST be classified BEFORE int."""
        from json2xlsx.coerce import coerce_cell
        p_true = coerce_cell(True, self._opts(), ctx=self.CTX)
        self.assertIs(type(p_true.value), bool)
        self.assertIs(p_true.value, True)
        p_false = coerce_cell(False, self._opts(), ctx=self.CTX)
        self.assertIs(type(p_false.value), bool)
        self.assertIs(p_false.value, False)

    def test_coerce_float_to_float(self) -> None:
        from json2xlsx.coerce import coerce_cell
        p = coerce_cell(3.14, self._opts(), ctx=self.CTX)
        self.assertEqual(p.value, 3.14)
        self.assertIs(type(p.value), float)

    def test_coerce_none_to_none(self) -> None:
        from json2xlsx.coerce import coerce_cell
        p = coerce_cell(None, self._opts(), ctx=self.CTX)
        self.assertIsNone(p.value)
        self.assertIsNone(p.number_format)

    # ----- ISO-date coercion -----------------------------------------------

    def test_coerce_iso_date_to_date(self) -> None:
        from datetime import date
        from json2xlsx.coerce import coerce_cell
        p = coerce_cell("2024-01-15", self._opts(), ctx=self.CTX)
        self.assertEqual(p.value, date(2024, 1, 15))
        self.assertEqual(p.number_format, "YYYY-MM-DD")

    def test_coerce_iso_datetime_to_datetime(self) -> None:
        from datetime import datetime
        from json2xlsx.coerce import coerce_cell
        p = coerce_cell("2024-01-15T09:00:00", self._opts(), ctx=self.CTX)
        self.assertEqual(p.value, datetime(2024, 1, 15, 9, 0, 0))
        self.assertEqual(p.number_format, "YYYY-MM-DD HH:MM:SS")

    def test_coerce_aware_dt_default_to_utc_naive(self) -> None:
        """R4.e — aware datetime → UTC naive (default, no --strict-dates)."""
        from datetime import datetime
        from json2xlsx.coerce import coerce_cell
        p = coerce_cell("2024-01-15T09:00:00+02:00", self._opts(), ctx=self.CTX)
        # 09:00 +02:00 → 07:00 UTC.
        self.assertEqual(p.value, datetime(2024, 1, 15, 7, 0, 0))
        self.assertIsNone(p.value.tzinfo)

    def test_coerce_aware_dt_strict_dates_raises(self) -> None:
        """D7 lock — aware → TimezoneNotSupported."""
        from json2xlsx.coerce import coerce_cell
        from json2xlsx.exceptions import TimezoneNotSupported
        with self.assertRaises(TimezoneNotSupported) as cm:
            coerce_cell("2024-01-15T09:00:00Z", self._opts(strict_dates=True),
                        ctx=self.CTX)
        self.assertEqual(cm.exception.code, 2)
        self.assertEqual(cm.exception.details["value"], "2024-01-15T09:00:00Z")
        self.assertEqual(cm.exception.details["sheet"], "Sheet1")
        self.assertEqual(cm.exception.details["row"], 1)
        self.assertEqual(cm.exception.details["column"], "A")

    def test_coerce_invalid_date_string_default_passthrough(self) -> None:
        """R4.f — invalid date candidate → silent string passthrough by default."""
        from json2xlsx.coerce import coerce_cell
        p = coerce_cell("2026-13-99", self._opts(), ctx=self.CTX)
        self.assertEqual(p.value, "2026-13-99")
        self.assertIsNone(p.number_format)

    def test_coerce_invalid_date_string_strict_raises(self) -> None:
        """R4.g — invalid date candidate under --strict-dates → InvalidDateString."""
        from json2xlsx.coerce import coerce_cell
        from json2xlsx.exceptions import InvalidDateString
        with self.assertRaises(InvalidDateString) as cm:
            coerce_cell("2026-13-99", self._opts(strict_dates=True), ctx=self.CTX)
        self.assertEqual(cm.exception.code, 2)
        self.assertEqual(cm.exception.details["value"], "2026-13-99")
        # And a "looks-like-date" plain-text variant.
        with self.assertRaises(InvalidDateString):
            coerce_cell("2024-Jan-15", self._opts(strict_dates=True), ctx=self.CTX)

    def test_coerce_no_date_coerce_flag(self) -> None:
        """`--no-date-coerce` → strings stay as text."""
        from json2xlsx.coerce import coerce_cell
        p = coerce_cell("2024-01-15", self._opts(date_coerce=False), ctx=self.CTX)
        self.assertEqual(p.value, "2024-01-15")
        self.assertIs(type(p.value), str)
        self.assertIsNone(p.number_format)

    def test_coerce_date_format_override(self) -> None:
        """`--date-format DD/MM/YYYY` overrides the default Excel format."""
        from datetime import date
        from json2xlsx.coerce import coerce_cell
        p = coerce_cell(
            "2024-01-15",
            self._opts(date_format_override="DD/MM/YYYY"),
            ctx=self.CTX,
        )
        self.assertEqual(p.value, date(2024, 1, 15))
        self.assertEqual(p.number_format, "DD/MM/YYYY")

    def test_coerce_non_date_string_passthrough(self) -> None:
        """Plain strings unaffected by date settings."""
        from json2xlsx.coerce import coerce_cell
        for flags in (self._opts(), self._opts(strict_dates=True),
                      self._opts(date_coerce=False)):
            p = coerce_cell("hello", flags, ctx=self.CTX)
            self.assertEqual(p.value, "hello")
            self.assertIsNone(p.number_format)

    def test_coerce_naive_dt_unchanged(self) -> None:
        """Naive datetime string → naive datetime (no tz conversion)."""
        from datetime import datetime
        from json2xlsx.coerce import coerce_cell
        p = coerce_cell("2024-01-15T09:00:00", self._opts(), ctx=self.CTX)
        self.assertEqual(p.value, datetime(2024, 1, 15, 9, 0, 0))
        self.assertIsNone(p.value.tzinfo)

    def test_coerce_datetime_space_separator(self) -> None:
        """Sarcasmotron-flagged gap (004.05): `[T ]` regex supports both
        `T` and space separators per ISO-8601; bench against both."""
        from datetime import datetime
        from json2xlsx.coerce import coerce_cell
        for s in ("2024-01-15T09:00:00", "2024-01-15 09:00:00"):
            p = coerce_cell(s, self._opts(), ctx=self.CTX)
            self.assertEqual(p.value, datetime(2024, 1, 15, 9, 0, 0))
            self.assertEqual(p.number_format, "YYYY-MM-DD HH:MM:SS")


# ===========================================================================
# Writer (F4, owned by 004.06)
# ===========================================================================
class TestWriter(unittest.TestCase):
    """Tests for `json2xlsx.writer` (F4 — workbook construction + styling)."""

    def test_smoke(self) -> None:
        _smoke_import("json2xlsx.writer")

    # ----- unit-level: header merging + sheet-name validation --------------

    def test_union_headers_first_seen_order(self) -> None:
        from json2xlsx.writer import _union_headers
        rows = [{"a": 1}, {"b": 2, "a": 3}, {"c": 4, "a": 5}]
        self.assertEqual(_union_headers(rows), ["a", "b", "c"])

    def test_validate_sheet_name_ok(self) -> None:
        from json2xlsx.writer import _validate_sheet_name
        # Must NOT raise.
        for ok in ("Employees", "Q1 2026", "Sheet 1", "Запись"):
            _validate_sheet_name(ok)

    def test_validate_sheet_name_empty(self) -> None:
        from json2xlsx.writer import _validate_sheet_name
        from json2xlsx.exceptions import InvalidSheetName
        with self.assertRaises(InvalidSheetName) as cm:
            _validate_sheet_name("")
        self.assertIn("empty", cm.exception.details["reason"])

    def test_validate_sheet_name_too_long(self) -> None:
        from json2xlsx.writer import _validate_sheet_name
        from json2xlsx.exceptions import InvalidSheetName
        with self.assertRaises(InvalidSheetName) as cm:
            _validate_sheet_name("x" * 32)
        self.assertIn("31", cm.exception.details["reason"])

    def test_validate_sheet_name_invalid_chars(self) -> None:
        from json2xlsx.writer import _validate_sheet_name
        from json2xlsx.exceptions import InvalidSheetName
        for bad in ("Q1/Q2", "A:B", "X*Y", "[name]", "back\\slash", "ques?n"):
            with self.assertRaises(InvalidSheetName, msg=bad):
                _validate_sheet_name(bad)

    def test_validate_sheet_name_reserved(self) -> None:
        from json2xlsx.writer import _validate_sheet_name
        from json2xlsx.exceptions import InvalidSheetName
        for variant in ("History", "history", "HISTORY", "HiStOrY"):
            with self.assertRaises(InvalidSheetName, msg=variant):
                _validate_sheet_name(variant)

    def test_validate_sheet_name_control_chars(self) -> None:
        """VDD-multi Logic M1 + Security LOW-1: Excel rejects control
        characters in sheet names. Plain `\\x00`, `\\n`, `\\t`, `\\r`
        all must trip the validator."""
        from json2xlsx.writer import _validate_sheet_name
        from json2xlsx.exceptions import InvalidSheetName
        for bad in ("\x00leading-nul", "tab\there", "new\nline", "ret\rurn", "bell\x07"):
            with self.assertRaises(InvalidSheetName, msg=repr(bad)) as cm:
                _validate_sheet_name(bad)
            self.assertIn("control character", cm.exception.details["reason"])

    def test_validate_sheet_name_apostrophe_edge(self) -> None:
        """VDD-multi Logic M1: Excel forbids `'` at first or last
        position (single-quote is its formula sheet-delimiter)."""
        from json2xlsx.writer import _validate_sheet_name
        from json2xlsx.exceptions import InvalidSheetName
        for bad in ("'leading", "trailing'", "'both'"):
            with self.assertRaises(InvalidSheetName, msg=bad) as cm:
                _validate_sheet_name(bad)
            self.assertIn("apostrophe", cm.exception.details["reason"])
        # Apostrophe in the middle is fine (Excel allows it).
        _validate_sheet_name("Bob's Sheet")

    # ----- integration-level: build then re-open via openpyxl --------------

    def _write(self, parsed_kwargs: dict, **write_kwargs):
        """Helper: build a ParsedInput, write to tempfile, return the
        openpyxl-reloaded Workbook + path.
        """
        from json2xlsx.coerce import CoerceOptions
        from json2xlsx.loaders import ParsedInput
        from json2xlsx.writer import write_workbook
        import openpyxl
        parsed = ParsedInput(**parsed_kwargs)
        coerce_opts = write_kwargs.pop("coerce_opts", CoerceOptions())
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = Path(f.name)
        write_workbook(parsed, path, coerce_opts=coerce_opts, **write_kwargs)
        return openpyxl.load_workbook(path), path

    def test_style_header_row_bold_grey_centre(self) -> None:
        wb, path = self._write({
            "shape": "array_of_objects",
            "sheets": {"Sheet1": [{"Name": "A", "Age": 1}]},
            "source_label": "t.json",
        })
        try:
            ws = wb["Sheet1"]
            self.assertTrue(ws["A1"].font.bold)
            self.assertIn(ws["A1"].fill.fgColor.rgb, ("F2F2F2", "00F2F2F2"))
            self.assertEqual(ws["A1"].alignment.horizontal, "center")
        finally:
            path.unlink(missing_ok=True)

    def test_size_columns_caps_at_max(self) -> None:
        long_val = "x" * 200  # well over MAX_COL_WIDTH=50
        wb, path = self._write({
            "shape": "array_of_objects",
            "sheets": {"Sheet1": [{"Long": long_val}]},
            "source_label": "t.json",
        })
        try:
            ws = wb["Sheet1"]
            self.assertLessEqual(ws.column_dimensions["A"].width, 50)
        finally:
            path.unlink(missing_ok=True)

    def test_write_workbook_freeze_pane_a2(self) -> None:
        wb, path = self._write({
            "shape": "array_of_objects",
            "sheets": {"Sheet1": [{"a": 1}]},
            "source_label": "t.json",
        })
        try:
            self.assertEqual(wb["Sheet1"].freeze_panes, "A2")
        finally:
            path.unlink(missing_ok=True)

    def test_write_workbook_auto_filter_set(self) -> None:
        wb, path = self._write({
            "shape": "array_of_objects",
            "sheets": {"Sheet1": [{"a": 1}, {"a": 2}]},
            "source_label": "t.json",
        })
        try:
            ref = wb["Sheet1"].auto_filter.ref
            self.assertIsNotNone(ref)
            self.assertTrue(ref.startswith("A1:"))
        finally:
            path.unlink(missing_ok=True)

    def test_write_workbook_multi_sheet_preserves_order(self) -> None:
        wb, path = self._write({
            "shape": "multi_sheet_dict",
            "sheets": {
                "Employees": [{"Name": "A"}],
                "Departments": [{"Dept": "E"}],
            },
            "source_label": "m.json",
        })
        try:
            self.assertEqual(wb.sheetnames, ["Employees", "Departments"])
        finally:
            path.unlink(missing_ok=True)

    def test_write_workbook_single_sheet_with_sheet_override(self) -> None:
        wb, path = self._write({
            "shape": "array_of_objects",
            "sheets": {"Sheet1": [{"a": 1}]},
            "source_label": "t.json",
        }, sheet_override="Custom")
        try:
            self.assertEqual(wb.sheetnames, ["Custom"])
        finally:
            path.unlink(missing_ok=True)

    def test_write_workbook_single_sheet_default_name(self) -> None:
        wb, path = self._write({
            "shape": "array_of_objects",
            "sheets": {"Sheet1": [{"a": 1}]},
            "source_label": "t.json",
        })
        try:
            self.assertEqual(wb.sheetnames, ["Sheet1"])
        finally:
            path.unlink(missing_ok=True)

    def test_write_workbook_multi_sheet_ignores_sheet_override(self) -> None:
        """Sarcasmotron-flagged gap (004.06): writer-level contract that
        multi-sheet input IGNORES sheet_override (CLI layer 004.07 owns
        the stderr warning). Lock the writer contract here so a future
        edit can't silently start applying the override on dict roots.
        """
        wb, path = self._write({
            "shape": "multi_sheet_dict",
            "sheets": {"A": [{"x": 1}], "B": [{"y": 2}]},
            "source_label": "m.json",
        }, sheet_override="ShouldBeIgnored")
        try:
            self.assertEqual(wb.sheetnames, ["A", "B"])
        finally:
            path.unlink(missing_ok=True)

    def test_write_workbook_missing_keys_empty_cells(self) -> None:
        wb, path = self._write({
            "shape": "array_of_objects",
            "sheets": {"Sheet1": [
                {"a": 1, "b": "x"},
                {"a": 2},  # missing "b"
                {"a": 3, "b": None},  # null b
            ]},
            "source_label": "t.json",
        })
        try:
            ws = wb["Sheet1"]
            self.assertEqual(ws["B2"].value, "x")
            self.assertIsNone(ws["B3"].value)  # missing key
            self.assertIsNone(ws["B4"].value)  # null value (R3.d)
        finally:
            path.unlink(missing_ok=True)

    def test_write_workbook_bool_cell_is_bool(self) -> None:
        wb, path = self._write({
            "shape": "array_of_objects",
            "sheets": {"Sheet1": [{"flag": True}, {"flag": False}]},
            "source_label": "t.json",
        })
        try:
            ws = wb["Sheet1"]
            self.assertIs(ws["A2"].value, True)
            self.assertIs(ws["A3"].value, False)
            # data_type 'b' = boolean in Excel.
            self.assertEqual(ws["A2"].data_type, "b")
            self.assertEqual(ws["A3"].data_type, "b")
        finally:
            path.unlink(missing_ok=True)

    def test_write_workbook_date_cell_number_format(self) -> None:
        """openpyxl reads date cells back as datetime (midnight) per its
        Excel-serial-number resolution — accept either type."""
        from datetime import date, datetime
        wb, path = self._write({
            "shape": "array_of_objects",
            "sheets": {"Sheet1": [{"Hired": "2024-01-15"}]},
            "source_label": "t.json",
        })
        try:
            cell = wb["Sheet1"]["A2"]
            self.assertIn(
                cell.value,
                (date(2024, 1, 15), datetime(2024, 1, 15, 0, 0)),
            )
            self.assertEqual(cell.number_format, "YYYY-MM-DD")
            # data_type 'd' = date in Excel (numeric serial + format).
            self.assertEqual(cell.data_type, "d")
        finally:
            path.unlink(missing_ok=True)

    # ----- AQ-1 drift assertion (was the only live test pre-004.06) --------

    def test_style_constants_drift_csv2xlsx(self) -> None:
        """AQ-1 lock — accepts both 6-char source literal and 8-char
        openpyxl-normalised ARGB form. The csv2xlsx side asserts live;
        the writer-side mirror self-skips until 004.06 lands the
        constants in `json2xlsx/writer.py`. Once that lands, the test
        asserts byte-equal style across the two modules.
        """
        import csv2xlsx  # noqa: WPS433 (intentional cross-script import in tests only)
        self.assertIn(
            csv2xlsx.HEADER_FILL.fgColor.rgb, ("F2F2F2", "00F2F2F2"),
            "csv2xlsx HEADER_FILL drift — task-001-baseline anchor broken",
        )
        self.assertTrue(csv2xlsx.HEADER_FONT.bold)
        self.assertEqual(csv2xlsx.HEADER_ALIGN.horizontal, "center")
        self.assertEqual(csv2xlsx.MAX_COL_WIDTH, 50)

        # Writer-side mirror — lands in 004.06.
        try:
            from json2xlsx.writer import (
                HEADER_FILL as W_FILL,
                HEADER_FONT as W_FONT,
                HEADER_ALIGN as W_ALIGN,
                MAX_COL_WIDTH as W_MAX_COL_WIDTH,
            )
        except ImportError:
            self.skipTest("writer-side style constants land in task-004-06")
        self.assertIn(W_FILL.fgColor.rgb, ("F2F2F2", "00F2F2F2"))
        self.assertEqual(W_FONT.bold, csv2xlsx.HEADER_FONT.bold)
        self.assertEqual(W_ALIGN.horizontal, csv2xlsx.HEADER_ALIGN.horizontal)
        self.assertEqual(W_MAX_COL_WIDTH, csv2xlsx.MAX_COL_WIDTH)


# ===========================================================================
# Exceptions (owned by 004.03)
# ===========================================================================
class TestExceptions(unittest.TestCase):
    """Tests for `json2xlsx.exceptions` — full `_AppError` attribute model
    and the 9 typed subclasses with cross-5 envelope payloads.
    """

    def test_apperror_is_plain_exception_not_dataclass(self) -> None:
        """ARCH §3.2 m1 lock — `_AppError` is a plain Exception subclass,
        NOT a frozen dataclass."""
        from json2xlsx.exceptions import _AppError
        # `_AppError` does NOT have `__dataclass_fields__` (frozen dataclasses do).
        self.assertFalse(hasattr(_AppError, "__dataclass_fields__"))
        self.assertTrue(issubclass(_AppError, Exception))

    def test_empty_input_envelope_payload(self) -> None:
        """TC-UNIT-EXC-01."""
        from json2xlsx.exceptions import EmptyInput
        exc = EmptyInput("file.json")
        self.assertEqual(exc.code, 2)
        self.assertEqual(exc.error_type, "EmptyInput")
        self.assertEqual(exc.details, {"source": "file.json"})
        self.assertTrue(exc.message)
        self.assertIn("file.json", exc.message)

    def test_self_overwrite_refused_code_6(self) -> None:
        """TC-UNIT-EXC-02 — exit code is the cross-7 H1 contract value."""
        from json2xlsx.exceptions import SelfOverwriteRefused
        exc = SelfOverwriteRefused(input_path="/tmp/a", output_path="/tmp/b")
        self.assertEqual(exc.code, 6)
        self.assertEqual(exc.error_type, "SelfOverwriteRefused")
        self.assertEqual(exc.details, {"input": "/tmp/a", "output": "/tmp/b"})
        # VDD-multi Logic L5 fix: human-readable message must include
        # BOTH paths (helps users debug which side typo'd).
        self.assertIn("/tmp/a", exc.message)
        self.assertIn("/tmp/b", exc.message)

    def test_post_validate_failed_truncates_8192(self) -> None:
        """TC-UNIT-EXC-03 — validator output capped at 8192 BYTES of UTF-8.

        Contract is byte-cap (NOT character-cap) because the envelope is
        emitted as a UTF-8-encoded JSON line and a Cyrillic / CJK payload
        would otherwise overflow up to 4x with a naive `s[:8192]` slice.
        """
        from json2xlsx.exceptions import PostValidateFailed
        # ASCII payload — bytes == chars, so the cap is a clean 8192.
        ascii_big = "x" * 20000
        exc_ascii = PostValidateFailed(validator_output=ascii_big)
        self.assertEqual(exc_ascii.code, 7)
        self.assertEqual(exc_ascii.error_type, "PostValidateFailed")
        out_ascii = exc_ascii.details["validator_output"]
        self.assertEqual(len(out_ascii.encode("utf-8")), 8192)
        # Multi-byte payload — char count would have overflowed in
        # naive slicing; ensure the byte cap holds.
        cyr_big = "ё" * 20000  # ё encodes to 2 UTF-8 bytes
        exc_cyr = PostValidateFailed(validator_output=cyr_big)
        out_cyr = exc_cyr.details["validator_output"]
        self.assertLessEqual(len(out_cyr.encode("utf-8")), 8192)
        # And the short-input path (no truncation needed).
        exc_tiny = PostValidateFailed(validator_output="ok")
        self.assertEqual(exc_tiny.details["validator_output"], "ok")

    def test_all_subclasses_carry_full_payload(self) -> None:
        """TC-UNIT-EXC-04 — every typed error has the four-attr model."""
        from json2xlsx.exceptions import (
            _AppError,
            EmptyInput, NoRowsToWrite, JsonDecodeError, UnsupportedJsonShape,
            InvalidSheetName, TimezoneNotSupported, InvalidDateString,
            SelfOverwriteRefused, PostValidateFailed,
        )
        samples = [
            EmptyInput("x"),
            NoRowsToWrite(),
            JsonDecodeError(line=1, column=2, msg="x"),
            UnsupportedJsonShape(root_type="int", hint="x"),
            InvalidSheetName(name="x", reason="y"),
            TimezoneNotSupported(value="x", sheet="S", row=1, column="A", tz_offset="+00:00"),
            InvalidDateString(value="x", sheet="S", row=1, column="A"),
            SelfOverwriteRefused(input_path="a", output_path="b"),
            PostValidateFailed(validator_output=""),
        ]
        for exc in samples:
            self.assertIsInstance(exc, _AppError)
            self.assertIsInstance(exc.message, str)
            self.assertIsInstance(exc.code, int)
            self.assertIsInstance(exc.error_type, str)
            self.assertIsInstance(exc.details, dict)
            # Cross-5 contract: code in {1, 2, 6, 7} for taxonomy errors.
            self.assertIn(exc.code, {1, 2, 6, 7})


# ===========================================================================
# CLI helpers (F6 enable-check + F8 same-path guard, owned by 004.03)
# ===========================================================================
class TestCliHelpers(unittest.TestCase):
    """Tests for `json2xlsx.cli_helpers` (F6 + F8)."""

    def test_smoke(self) -> None:
        _smoke_import("json2xlsx.cli_helpers")

    # ----- assert_distinct_paths (F8) --------------------------------------

    def test_assert_distinct_paths_collision_exit_6(self) -> None:
        """TC-UNIT-CLIH-03 — input/output resolve to same path → exit 6."""
        from json2xlsx.cli_helpers import assert_distinct_paths
        from json2xlsx.exceptions import SelfOverwriteRefused
        with tempfile.TemporaryDirectory() as td:
            p = Path(td) / "same.xlsx"
            p.touch()
            with self.assertRaises(SelfOverwriteRefused) as cm:
                assert_distinct_paths(str(p), p)
            self.assertEqual(cm.exception.code, 6)
            self.assertEqual(cm.exception.error_type, "SelfOverwriteRefused")

    def test_assert_distinct_paths_stdin_skipped(self) -> None:
        """TC-UNIT-CLIH-02 — stdin sentinel `-` never trips the guard."""
        from json2xlsx.cli_helpers import assert_distinct_paths
        # Should NOT raise regardless of the output path.
        assert_distinct_paths("-", Path("/tmp/anything.xlsx"))
        assert_distinct_paths("-", Path("/dev/null"))

    def test_assert_distinct_paths_symlink_follow(self) -> None:
        """TC-UNIT-CLIH-04 — symlink chain to same target trips guard
        (Path.resolve follows symlinks)."""
        from json2xlsx.cli_helpers import assert_distinct_paths
        from json2xlsx.exceptions import SelfOverwriteRefused
        with tempfile.TemporaryDirectory() as td:
            target = Path(td) / "real.xlsx"
            target.touch()
            link = Path(td) / "link.xlsx"
            try:
                link.symlink_to(target)
            except OSError:  # pragma: no cover (Windows w/o privileges)
                self.skipTest("symlink creation not permitted on this platform")
            with self.assertRaises(SelfOverwriteRefused):
                assert_distinct_paths(str(link), target)

    def test_assert_distinct_paths_different_paths_ok(self) -> None:
        """TC-UNIT-CLIH-05 — different paths are NOT a collision."""
        from json2xlsx.cli_helpers import assert_distinct_paths
        with tempfile.TemporaryDirectory() as td:
            a = Path(td) / "in.json"
            b = Path(td) / "out.xlsx"
            a.touch()
            assert_distinct_paths(str(a), b)  # must not raise

    # ----- post_validate_enabled (F6) --------------------------------------

    def test_post_validate_enabled_truthy_allowlist(self) -> None:
        """TC-UNIT-CLIH-01 — every truthy form enables the hook."""
        from json2xlsx.cli_helpers import post_validate_enabled
        truthy_forms = ["1", "true", "TRUE", "True", "yes", "Yes", "YES", "on", "On"]
        for v in truthy_forms:
            prev = os.environ.get("XLSX_JSON2XLSX_POST_VALIDATE")
            os.environ["XLSX_JSON2XLSX_POST_VALIDATE"] = v
            try:
                self.assertTrue(post_validate_enabled(), f"truthy={v!r}")
            finally:
                if prev is None:
                    os.environ.pop("XLSX_JSON2XLSX_POST_VALIDATE", None)
                else:
                    os.environ["XLSX_JSON2XLSX_POST_VALIDATE"] = prev

    def test_post_validate_enabled_falsy_off(self) -> None:
        """TC-UNIT-CLIH-01 negative — anything outside the allowlist is off."""
        from json2xlsx.cli_helpers import post_validate_enabled
        falsy_forms = ["0", "", "false", "FALSE", "no", "off", "anything-else", "  "]
        for v in falsy_forms:
            prev = os.environ.get("XLSX_JSON2XLSX_POST_VALIDATE")
            os.environ["XLSX_JSON2XLSX_POST_VALIDATE"] = v
            try:
                self.assertFalse(post_validate_enabled(), f"falsy={v!r}")
            finally:
                if prev is None:
                    os.environ.pop("XLSX_JSON2XLSX_POST_VALIDATE", None)
                else:
                    os.environ["XLSX_JSON2XLSX_POST_VALIDATE"] = prev
        # And the missing-env case.
        prev = os.environ.pop("XLSX_JSON2XLSX_POST_VALIDATE", None)
        try:
            self.assertFalse(post_validate_enabled())
        finally:
            if prev is not None:
                os.environ["XLSX_JSON2XLSX_POST_VALIDATE"] = prev

    # ----- read_stdin_utf8 (F1 helper) -------------------------------------

    def test_read_stdin_utf8(self) -> None:
        """TC-UNIT-CLIH-06 — bytes pass through the stdin buffer.

        Uses a real subprocess (no mock) per the task spec: pipe a
        small JSON payload to a child that imports `read_stdin_utf8`
        and emits the byte length on stdout.
        """
        payload = b'{"name":"Alice","age":30}\n'
        child = (
            "import sys, os; "
            "sys.path.insert(0, os.environ['XLSX2_SCRIPTS']); "
            "from json2xlsx.cli_helpers import read_stdin_utf8; "
            "data = read_stdin_utf8(); "
            "sys.stdout.write(str(len(data)))"
        )
        proc = subprocess.run(
            [sys.executable, "-c", child],
            input=payload,
            capture_output=True,
            timeout=15,
            env={**os.environ, "XLSX2_SCRIPTS": str(_SCRIPTS_DIR)},
        )
        self.assertEqual(proc.returncode, 0, proc.stderr.decode("utf-8", "replace"))
        self.assertEqual(int(proc.stdout.decode()), len(payload))


# ===========================================================================
# Round-trip with xlsx-8 (UC-5; owned by 004.08 synthetic, xlsx-8 merge live)
# ===========================================================================
class TestRoundTripXlsx8(unittest.TestCase):
    """Synthetic xlsx-8 round-trip — locks UC-5 contract in v1.

    Live wiring (`test_live_roundtrip`) is gated by AQ-5
    `@unittest.skipUnless` until xlsx-8 lands.
    """

    GOLDEN = Path(__file__).resolve().parent / "golden" / "json2xlsx_xlsx8_shape.json"

    def test_golden_fixture_exists(self) -> None:
        """Lock the contract fixture's presence and JSON-parseability.

        Runs immediately in 004.02 — proves the synthetic xlsx-8 output
        shape is committed to the repo and well-formed JSON. The
        downstream `test_synthetic_roundtrip` (004.08) consumes the
        same file.
        """
        import json
        self.assertTrue(self.GOLDEN.is_file(), f"missing fixture: {self.GOLDEN}")
        doc = json.loads(self.GOLDEN.read_text("utf-8"))
        self.assertEqual(list(doc.keys()), ["Employees", "Departments"])
        self.assertEqual(len(doc["Employees"]), 3)
        self.assertEqual(len(doc["Departments"]), 2)
        # Sanity on the locked attributes (sheet names, null preservation,
        # bool, ISO-date strings).
        emp = doc["Employees"]
        self.assertEqual(emp[0]["Name"], "Alice")
        self.assertEqual(emp[0]["Hired"], "2024-01-15")
        self.assertIs(emp[0]["Active"], True)
        self.assertIs(emp[1]["Active"], True)
        self.assertIs(emp[2]["Active"], False)
        self.assertIsNone(emp[2]["Salary"])

    def test_synthetic_roundtrip(self) -> None:
        """Synthetic xlsx-8 round-trip — UC-5 contract lock.

        Read the golden JSON (committed to repo); produce a workbook
        via the public helper; reload via openpyxl; verify sheet
        names + headers + key cell values match expected. Formatting
        / styles NOT compared (xlsx-2 produces fresh styling).
        """
        from datetime import date, datetime
        import openpyxl
        from json2xlsx import convert_json_to_xlsx
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            out = Path(tmp.name)
        try:
            rc = convert_json_to_xlsx(str(self.GOLDEN), str(out))
            self.assertEqual(rc, 0)
            wb = openpyxl.load_workbook(out)
            self.assertEqual(wb.sheetnames, ["Employees", "Departments"])

            emp = wb["Employees"]
            self.assertEqual(
                [c.value for c in emp[1]],
                ["Name", "Hired", "Salary", "Active"],
            )
            self.assertEqual(emp["A2"].value, "Alice")
            # openpyxl round-trips date → datetime (midnight).
            self.assertIn(
                emp["B2"].value,
                (date(2024, 1, 15), datetime(2024, 1, 15, 0, 0)),
            )
            self.assertEqual(emp["C2"].value, 100000)
            self.assertIs(emp["D2"].value, True)
            self.assertIsNone(emp["C4"].value)  # Carol's null salary
            self.assertIs(emp["D4"].value, False)

            dept = wb["Departments"]
            self.assertEqual(
                [c.value for c in dept[1]],
                ["Dept", "Head", "HC"],
            )
            self.assertEqual(dept["A2"].value, "Eng")
            self.assertEqual(dept["C2"].value, 12)
        finally:
            out.unlink(missing_ok=True)

    @unittest.skipUnless(_xlsx2json_available(), "xlsx-8 not landed yet")
    def test_live_roundtrip(self) -> None:
        """Live test: xlsx2json(original) → JSON → json2xlsx → assert
        structural equivalence. Activated automatically once xlsx-8
        lands and `import xlsx2json` succeeds.

        Implementation body is reserved for the xlsx-8 merge commit
        (per O1 closure — xlsx-2 owns the contract; xlsx-8 wires the
        live invocation).
        """
        self.skipTest("Implementation deferred to xlsx-8 merge commit.")


# ===========================================================================
# Post-validate hook (F6 subprocess body, owned by 004.08)
# ===========================================================================
class TestPostValidate(unittest.TestCase):
    """Tests for `cli_helpers.run_post_validate` + the CLI cleanup-on-
    failure path. The hook is gated by `XLSX_JSON2XLSX_POST_VALIDATE`
    and exits the CLI with code 7 on failure, unlinking the output.
    """

    def _write_valid_workbook(self, td: str) -> Path:
        """Produce a structurally-valid .xlsx via the public helper
        inside the given temp directory (caller owns cleanup)."""
        from json2xlsx import convert_json_to_xlsx
        inp = Path(td) / "in.json"
        inp.write_text('[{"a": 1}]', encoding="utf-8")
        out = Path(td) / "out.xlsx"
        rc = convert_json_to_xlsx(str(inp), str(out))
        self.assertEqual(rc, 0)
        return out

    def test_run_post_validate_success_on_valid_workbook(self) -> None:
        """On a structurally-valid xlsx produced by xlsx-2, the
        cross-format validator should exit 0 → (True, True, ...)."""
        from json2xlsx.cli_helpers import run_post_validate
        with tempfile.TemporaryDirectory() as td:
            out = self._write_valid_workbook(td)
            passed, hook_ok, captured = run_post_validate(out)
            self.assertTrue(
                passed,
                f"post-validate failed on a valid xlsx: {captured[:1000]}",
            )
            self.assertTrue(hook_ok, "hook_ok must be True when validator ran")

    def test_run_post_validate_missing_validator_returns_hook_failure(self) -> None:
        """If `office/validate.py` is missing, return
        (False, False, '...not found...') — VDD-multi H1 fix: the
        hook_ok=False flag tells the CLI orchestrator NOT to unlink
        the (presumably valid) output workbook on this code path.
        """
        import json2xlsx.cli_helpers as ch
        with tempfile.TemporaryDirectory() as td:
            out = self._write_valid_workbook(td)
            original_is_file = Path.is_file

            def fake_is_file(self_path):
                # Fake only the validator path; everything else real.
                if self_path.name == "validate.py" and "office" in str(self_path):
                    return False
                return original_is_file(self_path)
            Path.is_file = fake_is_file  # type: ignore[method-assign]
            try:
                passed, hook_ok, captured = ch.run_post_validate(out)
            finally:
                Path.is_file = original_is_file  # type: ignore[method-assign]
            self.assertFalse(passed)
            self.assertFalse(hook_ok, "hook_ok must be False on missing-validator path")
            self.assertIn("not found", captured)

    def test_post_validate_off_by_default(self) -> None:
        """Without the env var, `cli._run` MUST NOT invoke the hook
        (assertion: file existence + non-7 exit on a happy path)."""
        from json2xlsx import convert_json_to_xlsx
        # Ensure the env var is OFF.
        prev = os.environ.pop("XLSX_JSON2XLSX_POST_VALIDATE", None)
        try:
            with tempfile.TemporaryDirectory() as td:
                inp = Path(td) / "in.json"
                inp.write_text('[{"a": 1}]', encoding="utf-8")
                out = Path(td) / "out.xlsx"
                rc = convert_json_to_xlsx(str(inp), str(out))
                self.assertEqual(rc, 0)
                self.assertTrue(out.is_file())
        finally:
            if prev is not None:
                os.environ["XLSX_JSON2XLSX_POST_VALIDATE"] = prev

    def test_post_validate_workbook_failure_unlinks_output(self) -> None:
        """VDD-multi H1 fix: when the hook ran to completion AND
        reported a real workbook problem (`passed=False, hook_ok=True`),
        the CLI MUST:
          (a) emit a PostValidateFailed envelope (exit 7),
          (b) unlink the output (cleanup-on-failure).
        """
        import json2xlsx.cli as cli_mod
        original = cli_mod.run_post_validate
        cli_mod.run_post_validate = lambda _out: (False, True, "synthetic workbook failure")
        os.environ["XLSX_JSON2XLSX_POST_VALIDATE"] = "1"
        try:
            with tempfile.TemporaryDirectory() as td:
                inp = Path(td) / "in.json"
                inp.write_text('[{"a": 1}]', encoding="utf-8")
                out = Path(td) / "out.xlsx"
                rc = cli_mod.main([str(inp), str(out)])
                self.assertEqual(rc, 7)
                # Cleanup-on-failure: output file should be unlinked.
                self.assertFalse(
                    out.is_file(),
                    "output should be unlinked when post-validate fails",
                )
        finally:
            cli_mod.run_post_validate = original
            os.environ.pop("XLSX_JSON2XLSX_POST_VALIDATE", None)

    def test_post_validate_hook_failure_keeps_output(self) -> None:
        """VDD-multi H1 fix: when the hook itself broke (validator
        missing / timeout / etc. — `hook_ok=False`), the CLI MUST:
          (a) emit a PostValidateHookError envelope (exit 7),
          (b) LEAVE the output workbook intact — it is presumed valid.
        Prevents the worst-case regression where a missing validator
        silently deletes a freshly-produced, correct workbook.
        """
        import json2xlsx.cli as cli_mod
        original = cli_mod.run_post_validate
        cli_mod.run_post_validate = lambda _out: (
            False, False, "validator not found: synthetic"
        )
        os.environ["XLSX_JSON2XLSX_POST_VALIDATE"] = "1"
        try:
            with tempfile.TemporaryDirectory() as td:
                inp = Path(td) / "in.json"
                inp.write_text('[{"a": 1}]', encoding="utf-8")
                out = Path(td) / "out.xlsx"
                rc = cli_mod.main([str(inp), str(out)])
                self.assertEqual(rc, 7)
                # CRITICAL: output MUST survive hook-infrastructure failure.
                self.assertTrue(
                    out.is_file(),
                    "workbook must NOT be deleted when the hook itself fails",
                )
        finally:
            cli_mod.run_post_validate = original
            os.environ.pop("XLSX_JSON2XLSX_POST_VALIDATE", None)


# ===========================================================================
# Public surface invariants (locked by 004.01 — re-asserted here)
# ===========================================================================
class TestPublicSurface(unittest.TestCase):
    """Re-locks the 004.01 public-surface contract at the unit-test level
    so accidental package-init rewrites are caught early.
    """

    def test_convert_helper_kwargs_with_flag_lookalike_value(self) -> None:
        """VDD-multi Logic M4 fix: a kwarg whose VALUE starts with `--`
        must NOT poison the argparse parse. The helper uses
        `--flag=value` form so the value cannot be misread as a flag.
        """
        from json2xlsx import convert_json_to_xlsx
        with tempfile.TemporaryDirectory() as td:
            inp = Path(td) / "in.json"
            inp.write_text('[{"a": 1}]', encoding="utf-8")
            out = Path(td) / "out.xlsx"
            # `date_format` starts with `--` — must NOT be swallowed as
            # a separate `--strict-dates` flag.
            rc = convert_json_to_xlsx(
                str(inp), str(out),
                date_format="--strict-dates",
            )
            self.assertEqual(rc, 0)
            self.assertTrue(out.is_file())
            # The date_format string is stored on the date cell — verify
            # by writing a date column to confirm the value applied.
            inp.write_text('[{"d": "2024-01-15"}]', encoding="utf-8")
            out2 = Path(td) / "out2.xlsx"
            rc2 = convert_json_to_xlsx(
                str(inp), str(out2),
                date_format="--strict-dates",
            )
            self.assertEqual(rc2, 0)
            import openpyxl
            wb = openpyxl.load_workbook(out2)
            self.assertEqual(wb.active["A2"].number_format, "--strict-dates")

    def test_re_export_set(self) -> None:
        from json2xlsx import (
            convert_json_to_xlsx, main, _run,
            _AppError, EmptyInput, NoRowsToWrite, JsonDecodeError,
            UnsupportedJsonShape, InvalidSheetName, TimezoneNotSupported,
            InvalidDateString, SelfOverwriteRefused, PostValidateFailed,
        )
        # All callable / class-typed
        self.assertTrue(callable(convert_json_to_xlsx))
        self.assertTrue(callable(main))
        self.assertTrue(callable(_run))
        # Exception inheritance lock.
        for cls in (EmptyInput, NoRowsToWrite, JsonDecodeError,
                    UnsupportedJsonShape, InvalidSheetName,
                    TimezoneNotSupported, InvalidDateString,
                    SelfOverwriteRefused, PostValidateFailed):
            self.assertTrue(issubclass(cls, _AppError))
            self.assertTrue(issubclass(_AppError, Exception))

    def test_shim_help_exits_zero(self) -> None:
        """Lock TC-E2E-01 at unit level — `--help` short-circuits."""
        shim = _SCRIPTS_DIR / "json2xlsx.py"
        proc = subprocess.run(
            [sys.executable, str(shim), "--help"],
            capture_output=True, timeout=15,
        )
        self.assertEqual(proc.returncode, 0)
        self.assertIn(b"usage:", proc.stdout)
        self.assertIn(b"input", proc.stdout)
        self.assertIn(b"output", proc.stdout)


@unittest.skipUnless(
    os.environ.get("RUN_PERF_TESTS") == "1",
    "Set RUN_PERF_TESTS=1 to run perf-budget tests "
    "(no CI gate — reviewer runs locally before merge).",
)
class TestPerfBudget(unittest.TestCase):
    """Informal perf budget (NOT CI-gated). 100K-row JSONL budget is
    deferred to v2 per honest scope §11.3 / O4 — the openpyxl
    `write_only=True` mode is out of scope v1; current normal-write
    mode meets the 10K-row × 6-col ≤ 3 s anchor.
    """

    def test_10k_rows_6_cols_under_3s(self) -> None:
        import json
        import time
        from json2xlsx import convert_json_to_xlsx
        rows = [
            {
                "a": i, "b": i * 2, "c": "x" * 20,
                "d": True, "e": 3.14, "f": "2024-01-15",
            }
            for i in range(10_000)
        ]
        with tempfile.TemporaryDirectory() as td:
            inp = Path(td) / "perf.json"
            inp.write_text(json.dumps(rows), encoding="utf-8")
            out = Path(td) / "perf.xlsx"
            t0 = time.perf_counter()
            rc = convert_json_to_xlsx(str(inp), str(out))
            elapsed = time.perf_counter() - t0
        self.assertEqual(rc, 0)
        self.assertLess(
            elapsed, 3.0,
            f"10K × 6-col write took {elapsed:.2f}s, informal budget 3.0s",
        )


if __name__ == "__main__":
    unittest.main()
