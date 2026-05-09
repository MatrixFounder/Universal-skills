#!/usr/bin/env python3
"""Comprehensive xlsx-7 RULE-SEMANTICS regression-coverage fixture.

Builds `tests/golden/inputs/comprehensive-rules.xlsx` with one sheet
per SPEC §4–§5 rule family, each ≥ 20 rows. Paired with
`comprehensive-rules.rules.json` and exercised by
`TestComprehensiveRulesCoverage` in `tests/test_xlsx_check_rules.py`.

**SCOPE (Sarcasmotron iter-1 hygiene per S1):** this fixture covers
RULE SEMANTICS — comparison/type-guard/regex/date/aggregate/composite
predicates over the SPEC §4 scope vocabulary. It does NOT cover:
  - YAML-specific hardening (this fixture is JSON only — no
    billion-laughs / alias rejection / dup-key smuggling).
  - Pathological regex (no nested quantifiers; one well-formed
    `^PRJ-\\d{4}$`). ReDoS reject-list lives in dedicated tests.
  - 1 MiB rules-file cap (this rules file is ~6 KB).
  - cross-7 H1 same-path guard (no `--output` flag exercised).
  - `string.Template` format-string injection guard (rule messages
    use defaults).
  - Encrypted / macro-enabled inputs (cross-3 / cross-4).
Those hardening surfaces are tested separately in
`test_xlsx_check_rules.py::TestSecurity*` / `TestHonestScope*` /
`TestEnvelope*` etc.

Goal: prove every rule type fires (or correctly does NOT fire) on a
deterministic dataset, and lock the behaviour against future
regressions. Each sheet mixes compliant + violating rows with the
violations placed at known cell coordinates so the test can assert
on both summary counts AND specific cell-level findings.

Run manually when the SPEC or rules.json drifts:
    cd skills/xlsx/scripts
    .venv/bin/python tests/golden/inputs/make_comprehensive_fixture.py

The output is committed (Q5 hybrid pattern, see ../README.md). The
companion rules.json is hand-authored so the rule-text stays
reviewable.
"""
from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


OUT_DIR = Path(__file__).resolve().parent
OUT_XLSX = OUT_DIR / "comprehensive-rules.xlsx"


# === Shared helpers ========================================================

def _write_header(ws: Worksheet, headers: list[str]) -> None:
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)


def _set_error_cell(ws: Worksheet, row: int, col: int, code: str) -> None:
    """Write a cell with `t="e"` data type (Excel error cell). The
    auto-detecting `Cell.value` setter would coerce '#N/A' to text;
    we bypass via the lower-level `_value` + `data_type` to produce a
    `<c t="e"><v>#N/A</v></c>` round-trip exactly as Excel does."""
    cell = ws.cell(row=row, column=col)
    cell._value = code
    cell.data_type = "e"


# === Per-sheet builders ====================================================
#
# Each builder appends 20+ rows and returns the list of (row, col_letter,
# expected_rule_id) tuples that the validator MUST flag. The test
# walks this manifest and asserts each predicted finding is present.

def build_compare_scalar(ws: Worksheet) -> list[tuple[int, str, str]]:
    """SPEC §5.1 — scalar comparisons (==, !=, <, <=, >, >=). One
    column per operator; each column has a rule asserting the
    corresponding predicate."""
    _write_header(ws, ["EqVal", "NeqVal", "LtVal", "LteVal", "GtVal", "GteVal"])
    expected: list[tuple[int, str, str]] = []
    for row in range(2, 22):
        # Row 2-11: compliant; row 12-21: violations.
        violation = row >= 12
        # EqVal: rule is `value == 100`. Compliant=100, violation=99.
        ws.cell(row=row, column=1, value=99 if violation else 100)
        # NeqVal: rule is `value != 0`. Compliant=anything-non-zero, violation=0.
        ws.cell(row=row, column=2, value=0 if violation else (row - 1))
        # LtVal: rule is `value < 50`. Compliant=row-1, violation=99.
        ws.cell(row=row, column=3, value=99 if violation else (row - 1))
        # LteVal: rule is `value <= 50`. Compliant=50, violation=51.
        ws.cell(row=row, column=4, value=51 if violation else 50)
        # GtVal: rule is `value > 0`. Compliant=row, violation=-1.
        ws.cell(row=row, column=5, value=-1 if violation else row)
        # GteVal: rule is `value >= 10`. Compliant=10, violation=5.
        ws.cell(row=row, column=6, value=5 if violation else 10)
        if violation:
            for col_letter, rid in [
                ("A", "compare-eq"), ("B", "compare-neq"),
                ("C", "compare-lt"), ("D", "compare-lte"),
                ("E", "compare-gt"), ("F", "compare-gte"),
            ]:
                expected.append((row, col_letter, rid))
    return expected


def build_between(ws: Worksheet) -> list[tuple[int, str, str]]:
    """SPEC §5.1 — between (inclusive) and between_excl."""
    _write_header(ws, ["IncVal", "ExclVal"])
    expected: list[tuple[int, str, str]] = []
    for row in range(2, 22):
        violation = row >= 14
        # between:0,24 — compliant in [0,24], violation outside.
        ws.cell(row=row, column=1, value=99 if violation else (row % 25))
        # between_excl:0,1 — compliant in (0,1), violation at 0 or 1.
        ws.cell(row=row, column=2,
                value=(0 if row % 2 == 0 else 1) if violation else 0.5)
        if violation:
            expected.append((row, "A", "between-inc"))
            expected.append((row, "B", "between-excl"))
    return expected


def build_in_list(ws: Worksheet) -> list[tuple[int, str, str]]:
    """SPEC §5.1 — `value in [...]` and `value not in [...]`."""
    _write_header(ws, ["Status", "Forbidden"])
    expected: list[tuple[int, str, str]] = []
    allowed = ["Approved", "Pending", "Rejected"]
    for row in range(2, 22):
        violation = row >= 15
        ws.cell(row=row, column=1, value="Cancelled" if violation else allowed[row % 3])
        # Forbidden: rule is `value not in [Draft]`. Violation=Draft.
        ws.cell(row=row, column=2, value="Draft" if violation else "Final")
        if violation:
            expected.append((row, "A", "status-allowlist"))
            expected.append((row, "B", "status-denylist"))
    return expected


def build_type_guards(ws: Worksheet) -> list[tuple[int, str, str]]:
    """SPEC §5.2 — is_number/is_text/is_date/is_bool/is_error/required.
    Each column has rule asserting the corresponding type guard;
    violations populate the wrong type."""
    _write_header(ws, ["NumCol", "TextCol", "DateCol", "BoolCol", "ReqCol"])
    expected: list[tuple[int, str, str]] = []
    for row in range(2, 22):
        violation = row >= 13
        # NumCol: rule is_number; violation = text in numeric column.
        ws.cell(row=row, column=1, value="not-a-num" if violation else (row * 10))
        # TextCol: rule is_text; violation = number where text expected.
        ws.cell(row=row, column=2, value=42 if violation else f"text-{row}")
        # DateCol: rule is_date; violation = text "abc".
        if violation:
            ws.cell(row=row, column=3, value="not-a-date")
        else:
            ws.cell(row=row, column=3, value=date(2026, 5, row % 28 + 1))
        # BoolCol: rule is_bool; violation = string.
        ws.cell(row=row, column=4, value="not-bool" if violation else (row % 2 == 0))
        # ReqCol: rule required; violation = empty.
        if not violation:
            ws.cell(row=row, column=5, value=f"present-{row}")
        # else: leave cell empty
        if violation:
            expected.append((row, "A", "type-num"))
            expected.append((row, "B", "type-text"))
            expected.append((row, "C", "type-date"))
            expected.append((row, "D", "type-bool"))
            expected.append((row, "E", "required-non-empty"))
    return expected


def build_regex(ws: Worksheet) -> list[tuple[int, str, str]]:
    """SPEC §5.3 — regex:PATTERN. PRJ-#### project codes."""
    _write_header(ws, ["ProjectCode"])
    expected: list[tuple[int, str, str]] = []
    for row in range(2, 22):
        violation = row >= 14
        if violation:
            # Various violation shapes — wrong prefix, wrong digit count, etc.
            shapes = ["XXX-1234", "PRJ-12345", "prj-1234", "PRJ-AAAA", "PRJ123"]
            ws.cell(row=row, column=1, value=shapes[(row - 14) % len(shapes)])
            expected.append((row, "A", "project-format"))
        else:
            ws.cell(row=row, column=1, value=f"PRJ-{1000 + row:04d}")
    return expected


def build_len_text(ws: Worksheet) -> list[tuple[int, str, str]]:
    """SPEC §5.3 — `len OP N`, `not_empty`, `starts_with`, `ends_with`."""
    _write_header(ws, ["LenCapped", "Notes", "StartsTok", "EndsTok"])
    expected: list[tuple[int, str, str]] = []
    for row in range(2, 22):
        violation = row >= 14
        # LenCapped: `len <= 10`; violation = >10 chars.
        ws.cell(row=row, column=1,
                value=("X" * 25) if violation else f"row-{row}")
        # Notes: `not_empty`; violation = empty / whitespace-only.
        if not violation:
            ws.cell(row=row, column=2, value=f"comment {row}")
        else:
            ws.cell(row=row, column=2, value="")
        # StartsTok: `starts_with:OK-`; violation = wrong prefix.
        ws.cell(row=row, column=3,
                value=f"BAD-{row}" if violation else f"OK-{row}")
        # EndsTok: `ends_with:.csv`; violation = wrong suffix.
        ws.cell(row=row, column=4,
                value=f"file{row}.txt" if violation else f"file{row}.csv")
        if violation:
            expected.append((row, "A", "len-cap"))
            expected.append((row, "B", "notes-required"))
            expected.append((row, "C", "starts-with"))
            expected.append((row, "D", "ends-with"))
    return expected


def build_dates(ws: Worksheet) -> list[tuple[int, str, str]]:
    """SPEC §5.4 — date_in_month / date_in_range / date_before / date_after / date_weekday."""
    _write_header(ws, ["InMonth", "InRange", "Before", "After", "Weekday"])
    expected: list[tuple[int, str, str]] = []
    for row in range(2, 22):
        violation = row >= 14
        # InMonth: rule is `date_in_month:2026-05`; violation = April or June.
        ws.cell(row=row, column=1,
                value=date(2026, 4, 15) if violation else date(2026, 5, (row % 28) + 1))
        # InRange: rule is `date_in_range:2026-01-01,2026-06-30`; violation = July+.
        ws.cell(row=row, column=2,
                value=date(2026, 7, 15) if violation else date(2026, 3, (row % 28) + 1))
        # Before: rule is `date_before:2026-06-01`; violation = on/after that date.
        ws.cell(row=row, column=3,
                value=date(2026, 6, 5) if violation else date(2026, 5, 20))
        # After: rule is `date_after:2026-01-01`; violation = on/before.
        ws.cell(row=row, column=4,
                value=date(2025, 12, 25) if violation else date(2026, 3, 5))
        # Weekday: rule is `date_weekday:Mon,Tue,Wed,Thu,Fri`; violation = Sat/Sun.
        # 2026-05-02 is Saturday, 2026-05-03 is Sunday.
        if violation:
            ws.cell(row=row, column=5, value=date(2026, 5, 2))  # Sat
        else:
            # Pick a weekday deterministically.
            base = date(2026, 5, 4)  # Monday
            ws.cell(row=row, column=5, value=base + timedelta(days=row % 5))
        if violation:
            expected.append((row, "A", "date-in-month"))
            expected.append((row, "B", "date-in-range"))
            expected.append((row, "C", "date-before"))
            expected.append((row, "D", "date-after"))
            expected.append((row, "E", "date-weekday"))
    return expected


def build_aggregates_basic(ws: Worksheet) -> list[tuple[int, str, str]]:
    """SPEC §5.5 — sum / avg / min / max / median / stdev / count* aggregates.
    Single 'Hours' column; H1-H7 hold scalar cells that compare against
    the column's aggregates so the validator exercises both directions
    (cell vs aggregate)."""
    _write_header(ws, [
        "Hours",  # A
        "Note",   # B (text — skipped by numeric aggregates per §5.5.1)
    ])
    # Data rows: 20 numeric values 1..20 in col A; sum=210, avg=10.5,
    # min=1, max=20, median=10.5.
    for row in range(2, 22):
        ws.cell(row=row, column=1, value=row - 1)  # 1..20
        ws.cell(row=row, column=2, value=f"r{row}")
    # Scalar comparison cells in row 23-29 (so they're below the data
    # range and don't pollute the column's data values). The rules.json
    # uses cell:H23..H29 with the asserted aggregate.
    # Use violations on purpose so each rule fires exactly once.
    ws.cell(row=23, column=8, value=999)   # H23: should equal sum (210) — fail
    ws.cell(row=24, column=8, value=999)   # H24: avg (10.5) — fail
    ws.cell(row=25, column=8, value=999)   # H25: min (1) — fail
    ws.cell(row=26, column=8, value=0)     # H26: max (20) — fail (uses >=)
    ws.cell(row=27, column=8, value=0)     # H27: median (10.5) — fail
    ws.cell(row=28, column=8, value=0)     # H28: count (20) — fail
    ws.cell(row=29, column=8, value=0)     # H29: count_nonempty (20) — fail
    return [
        (23, "H", "agg-sum"),
        (24, "H", "agg-avg"),
        (25, "H", "agg-min"),
        (26, "H", "agg-max"),
        (27, "H", "agg-median"),
        (28, "H", "agg-count"),
        (29, "H", "agg-count-nonempty"),
    ]


def build_aggregate_arith(ws: Worksheet) -> list[tuple[int, str, str]]:
    """SPEC §5.5.2 — aggregate vs cell comparison. The rule
    `value >= sum(col:Cost)` (no `* 1.05` factor — parser doesn't
    compose `value * N` arithmetic with aggregate operands; documented
    honest-scope). H1 holds half the actual sum to guarantee the rule
    fires."""
    _write_header(ws, ["Cost", "_pad"])
    # 20 numeric rows.
    values = list(range(50, 110, 3))[:20]  # 50, 53, ..., 107
    for idx, v in enumerate(values, start=2):
        ws.cell(row=idx, column=1, value=v)
    # H1 = half of actual sum. Rule asserts H1 >= sum → fires.
    actual_sum = sum(values)
    cap = int(actual_sum / 2)
    ws.cell(row=1, column=8, value=cap)
    return [(1, "H", "budget-cap")]


def build_group_by(ws: Worksheet) -> list[tuple[int, str, str]]:
    """SPEC §5.6 — sum_by/count_by/avg_by. WeekNum partitions the data
    into 4 groups, each violating exactly one group-by rule:
      - Week 1: sum=50 (>40, fires sum-by); avg=10 (≤10, count=5).
      - Week 2: count=1 (<2, fires count-by); sum=5; avg=5.
      - Week 3: avg=12 (>10, fires avg-by); sum=72 also fires sum-by.
      - Week 4: 4 rows of 8 = sum=32 (≤40); avg=8; count=4. PASSES.
    Net: sum-by-week fires twice (Weeks 1, 3), count-by-week once
    (Week 2), avg-by-week once (Week 3). 4 grouped findings total —
    matches the manifest predictions."""
    _write_header(ws, ["Hours", "WeekNum", "Project"])
    plan = [
        # (hours, week, project)
        (12, 1, "PRJ-A"), (10, 1, "PRJ-A"), (10, 1, "PRJ-A"),
        (10, 1, "PRJ-A"), (8, 1, "PRJ-A"),                    # week 1: sum=50, avg=10
        (5, 2, "PRJ-B"),                                       # week 2: count=1
        (12, 3, "PRJ-A"), (12, 3, "PRJ-A"), (12, 3, "PRJ-A"),
        (12, 3, "PRJ-A"), (12, 3, "PRJ-A"), (12, 3, "PRJ-A"),  # week 3: sum=72, avg=12
        (8, 4, "PRJ-A"), (8, 4, "PRJ-A"), (8, 4, "PRJ-A"),
        (8, 4, "PRJ-A"),                                       # week 4: sum=32, count=4
    ]
    # Pad to ≥ 20 data rows with values that fall into Week 4 and
    # remain compliant: 4 rows of value 0 (number, but skipped from
    # avg/sum because they're zero — actually NOT skipped, but contribute
    # 0 to the sum). Use NUM 0 to keep counts stable; week-4 sum stays
    # 32, count grows to 8.
    plan.extend([(0, 4, "PRJ-A")] * 4)
    for idx, (h, w, p) in enumerate(plan, start=2):
        ws.cell(row=idx, column=1, value=h)
        ws.cell(row=idx, column=2, value=w)
        ws.cell(row=idx, column=3, value=p)
    # GroupBy findings target the SHEET (not specific cells) per §7.1.3.
    # Sentinel: row=0 (no real cell at row 0; Excel rows start at 1).
    # Documented at the manifest level — see expected.json `_comment`.
    # The 4 tuples below match the predicted firing breakdown above.
    return [
        (0, "", "sum-by-week"),    # Week 1
        (0, "", "sum-by-week"),    # Week 3
        (0, "", "count-by-week"),  # Week 2
        (0, "", "avg-by-week"),    # Week 3
    ]


def build_composite(ws: Worksheet) -> list[tuple[int, str, str]]:
    """SPEC §5.7 — composite and / or / not."""
    _write_header(ws, ["Hours", "Status"])
    expected: list[tuple[int, str, str]] = []
    for row in range(2, 22):
        # Rule: (Hours > 0 AND Hours <= 24) — violation if either fails.
        # Row 12: Hours = -2 → fails Hours>0
        # Row 13: Hours = 30 → fails Hours<=24
        # Row 14: Hours = 0 → fails Hours>0
        # Others: 8 hours.
        if row == 12:
            ws.cell(row=row, column=1, value=-2)
            expected.append((row, "A", "and-hours-realistic"))
        elif row == 13:
            ws.cell(row=row, column=1, value=30)
            expected.append((row, "A", "and-hours-realistic"))
        elif row == 14:
            ws.cell(row=row, column=1, value=0)
            expected.append((row, "A", "and-hours-realistic"))
        else:
            ws.cell(row=row, column=1, value=8)
        # Status: rule is `not (value == 'Cancelled')`.
        if row in (15, 16):
            ws.cell(row=row, column=2, value="Cancelled")
            expected.append((row, "B", "not-cancelled"))
        else:
            ws.cell(row=row, column=2, value="Approved")
    return expected


def build_or_composite(ws: Worksheet) -> list[tuple[int, str, str]]:
    """SPEC §5.7 — explicit OR composite. The predicate
    `value > 50 OR value < 10` MUST be satisfied (an OK row has an
    extreme value). Mid-range values (10..50) violate. Predicate
    semantics: the rule's check expresses what compliant cells look
    like; `or` returns False (= finding) when neither disjunct holds."""
    _write_header(ws, ["Amount"])
    expected: list[tuple[int, str, str]] = []
    for row in range(2, 22):
        # 6 violation rows scattered in the 20-row dataset.
        violation_rows = {7, 11, 15, 17, 19, 21}
        if row in violation_rows:
            ws.cell(row=row, column=1, value=25)  # in (10, 50) — fails OR
            expected.append((row, "A", "or-suspect"))
        else:
            # Alternate extreme values: low (<10) and high (>50).
            ws.cell(row=row, column=1, value=5 if row % 2 == 0 else 100)
    return expected


def build_when_filter(ws: Worksheet) -> list[tuple[int, str, str]]:
    """SPEC §3 — `when` field pre-filters rows. The when predicate
    operates on the SAME cell as the main check (current xlsx-7 v1
    implementation; cross-column when clauses are SPEC-aspirational
    but not yet wired — see L5 honest-scope). The fixture exercises
    the same-column when path: only run `value <= 24` when the cell
    is a number (text rows are filtered out by `when: is_number`)."""
    _write_header(ws, ["Hours"])
    expected: list[tuple[int, str, str]] = []
    for row in range(2, 22):
        # Row 5, 6: Hours=30 (number, when passes, check fails) → fires.
        # Row 7, 8: Hours='off-day' (text, when fails) → skipped despite
        #          violating "value <= 24" (string vs 24 comparison would
        #          have failed anyway, but the when-skip is what we test).
        # Others: Hours=8 (number, when passes, check passes).
        if row in (5, 6):
            ws.cell(row=row, column=1, value=30)
            expected.append((row, "A", "when-hours-realistic"))
        elif row in (7, 8):
            ws.cell(row=row, column=1, value="off-day")  # text — when filter excludes
        else:
            ws.cell(row=row, column=1, value=8)
    return expected


def build_tolerance(ws: Worksheet) -> list[tuple[int, str, str]]:
    """SPEC §5.5.4 — equality with tolerance. Rule asserts
    `value == 100` with tolerance 0.5; values within ±0.5 pass."""
    _write_header(ws, ["Measured"])
    expected: list[tuple[int, str, str]] = []
    for row in range(2, 22):
        if row == 19:
            ws.cell(row=row, column=1, value=101.0)  # outside tolerance
            expected.append((row, "A", "exact-100"))
        elif row == 20:
            ws.cell(row=row, column=1, value=98.0)   # outside tolerance
            expected.append((row, "A", "exact-100"))
        else:
            # Within ±0.5: 99.6, 100.0, 100.3, 99.9, ...
            offsets = [0.0, 0.3, -0.3, 0.4, -0.2, 0.1, -0.4, 0.5, -0.5,
                       0.0, 0.0, 0.2, -0.1, 0.0, 0.0, 0.0, 0.0]
            ws.cell(row=row, column=1, value=100 + offsets[(row - 2) % len(offsets)])
    return expected


def build_merged_data(ws: Worksheet) -> list[tuple[int, str, str]]:
    """SPEC §4.4 — merged cells in DATA range (not header). Anchor
    A2:A4 (merged 3 rows); B is normal."""
    _write_header(ws, ["MergedAnchor", "Side"])
    # 21 data rows.
    for row in range(2, 23):
        ws.cell(row=row, column=2, value=row)
    # Merge A2:A4 with anchor value "MergedHead" (passes is_text rule).
    # Then a violating cell at A12 (number where text expected).
    ws.cell(row=2, column=1, value="MergedHead")
    ws.merge_cells(start_row=2, end_row=4, start_column=1, end_column=1)
    # Other anchors that pass.
    for row in (5, 6, 7, 8, 9, 10, 11):
        ws.cell(row=row, column=1, value=f"text-{row}")
    # Violations:
    ws.cell(row=12, column=1, value=42)  # is_text rule fires
    for row in range(13, 23):
        ws.cell(row=row, column=1, value=f"text-{row}")
    return [(12, "A", "merged-text-only")]


def build_range_ref(ws: Worksheet) -> list[tuple[int, str, str]]:
    """SPEC §4 — RANGE scope (`A2:B25`)."""
    _write_header(ws, ["X", "Y"])
    expected: list[tuple[int, str, str]] = []
    for row in range(2, 22):
        # Rule on A2:B25: value > 0; violations in row 14-15.
        if row in (14, 15):
            ws.cell(row=row, column=1, value=-1)
            ws.cell(row=row, column=2, value=-2)
            expected.append((row, "A", "range-positive"))
            expected.append((row, "B", "range-positive"))
        else:
            ws.cell(row=row, column=1, value=row)
            ws.cell(row=row, column=2, value=row * 2)
    return expected


def build_row_ref(ws: Worksheet) -> list[tuple[int, str, str]]:
    """SPEC §4 — ROW scope (`row:5`). Row 5 has 6 cells; rule asserts
    each is text. Violations: row 5 cols B and D (numeric)."""
    _write_header(ws, ["A", "B", "C", "D", "E", "F"])
    # Fill rows 2-21 with text by default.
    for row in range(2, 22):
        for col in range(1, 7):
            ws.cell(row=row, column=col, value=f"r{row}c{col}")
    # Row 5 violations:
    ws.cell(row=5, column=2, value=42)
    ws.cell(row=5, column=4, value=99)
    return [
        (5, "B", "row5-text"),
        (5, "D", "row5-text"),
    ]


def build_sheet_scope(ws: Worksheet) -> list[tuple[int, str, str]]:
    """SPEC §4 — `sheet:NAME` scope. Asserts every non-empty cell
    below header is is_number; violations on rows 16, 17 col B."""
    _write_header(ws, ["A", "B"])
    expected: list[tuple[int, str, str]] = []
    for row in range(2, 22):
        ws.cell(row=row, column=1, value=row)
        if row in (16, 17):
            ws.cell(row=row, column=2, value="not-num")
            expected.append((row, "B", "sheet-numbers-only"))
        else:
            ws.cell(row=row, column=2, value=row * 10)
    return expected


def build_excel_table(ws: Worksheet) -> list[tuple[int, str, str]]:
    """SPEC §4.3 — Excel Table auto-detect. The rule uses
    `col:Hours` which resolves through the Table header (Tables take
    precedence over row-1 lookup per §4.3)."""
    from openpyxl.worksheet.table import Table, TableStyleInfo
    _write_header(ws, ["Date", "Hours", "Project"])
    expected: list[tuple[int, str, str]] = []
    for row in range(2, 22):
        ws.cell(row=row, column=1, value=date(2026, 5, (row % 28) + 1))
        if row in (10, 11):
            ws.cell(row=row, column=2, value=99)  # > 24 violation
            expected.append((row, "B", "table-hours-realistic"))
        else:
            ws.cell(row=row, column=2, value=8)
        ws.cell(row=row, column=3, value=f"PRJ-{1000 + row}")
    # Define the Table covering A1:C21.
    tab = Table(displayName="Timesheet1", ref="A1:C21")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2",
                                          showRowStripes=True)
    ws.add_table(tab)
    return expected


def build_error_cells(ws: Worksheet) -> list[tuple[int, str, str]]:
    """SPEC §5.0 — auto-emit `cell-error` finding. Excel error cells
    on column A short-circuit OTHER rules and emit synthetic
    `cell-error` finding instead. Error cells: A5 (#REF!), A10 (#N/A),
    A15 (#DIV/0!).

    L3 (Sarcasmotron iter-1 on this fixture): row 7 deliberately
    holds value=-5 (negative) so the user rule `error-cells-visit`
    (`value > 0`) fires its OWN finding on a non-error cell. Without
    this, the rule has zero predicted violations and a regression
    that silently disables the rule wouldn't be caught (only the
    auto-emitted `cell-error` finding would canary, and that's a
    different code path)."""
    _write_header(ws, ["MaybeError"])
    expected: list[tuple[int, str, str]] = []
    for row in range(2, 22):
        if row == 5:
            _set_error_cell(ws, row=row, col=1, code="#REF!")
            expected.append((row, "A", "cell-error"))
        elif row == 10:
            _set_error_cell(ws, row=row, col=1, code="#N/A")
            expected.append((row, "A", "cell-error"))
        elif row == 15:
            _set_error_cell(ws, row=row, col=1, code="#DIV/0!")
            expected.append((row, "A", "cell-error"))
        elif row == 7:
            ws.cell(row=row, column=1, value=-5)  # negative — fires error-cells-visit
            expected.append((row, "A", "error-cells-visit"))
        else:
            ws.cell(row=row, column=1, value=row * 10)
    return expected


def build_hidden_rows(ws: Worksheet) -> list[tuple[int, str, str]]:
    """SPEC §4.5 — hidden rows. By default included; the rule on this
    sheet uses default behaviour, so all violations fire. (The
    --visible-only behaviour is tested at the run level, not here.)"""
    _write_header(ws, ["Val"])
    expected: list[tuple[int, str, str]] = []
    for row in range(2, 22):
        # Rule: value > 0; violations at rows 6, 12 (one will be hidden).
        if row in (6, 12):
            ws.cell(row=row, column=1, value=-1)
            expected.append((row, "A", "hidden-positive"))
        else:
            ws.cell(row=row, column=1, value=row)
    # Hide row 12.
    ws.row_dimensions[12].hidden = True
    return expected


def build_required_skip_empty(ws: Worksheet) -> list[tuple[int, str, str]]:
    """SPEC §3 / §5.2 — `required` overrides `skip_empty: true`. The
    rule asserts `required` on column A; empty rows fire findings."""
    _write_header(ws, ["MustHave"])
    expected: list[tuple[int, str, str]] = []
    for row in range(2, 22):
        if row in (8, 13, 17):
            # Empty cell — REQUIRED rule fires.
            expected.append((row, "A", "must-have"))
        else:
            ws.cell(row=row, column=1, value=f"v-{row}")
    return expected


def build_apostrophe_sheet(ws: Worksheet) -> list[tuple[int, str, str]]:
    """SPEC §4.1 — apostrophe-escaped sheet name with spaces."""
    _write_header(ws, ["Hours"])
    expected: list[tuple[int, str, str]] = []
    for row in range(2, 22):
        if row in (15, 16, 17):
            ws.cell(row=row, column=1, value=99)
            expected.append((row, "A", "apostrophe-hours-realistic"))
        else:
            ws.cell(row=row, column=1, value=8)
    return expected


def build_cross_sheet_a(ws: Worksheet) -> list[tuple[int, str, str]]:
    """SPEC §5.5 — cross-sheet aggregate. The rule on CrossSheetA
    asserts `value <= sum(CrossSheetB!col:Cost)` (no `* 0.5` factor —
    parser doesn't compose `aggregate * N`; documented honest-scope).
    CrossSheetB's Cost column sums to 2000; rows with Spend > 2000
    violate."""
    _write_header(ws, ["Spend"])
    expected: list[tuple[int, str, str]] = []
    for row in range(2, 22):
        if row in (12, 13):
            ws.cell(row=row, column=1, value=99999)  # > 2000
            expected.append((row, "A", "cross-sheet-cap"))
        else:
            ws.cell(row=row, column=1, value=50)  # ≤ 2000
    return expected


def build_cross_sheet_b(ws: Worksheet) -> list[tuple[int, str, str]]:
    """Companion to CrossSheetA. Cost column sums to 100×20 = 2000;
    CrossSheetA's rule references it as the cap."""
    _write_header(ws, ["Cost"])
    for row in range(2, 22):
        ws.cell(row=row, column=1, value=100)  # sum = 2000
    return []


def build_count_distinct_errors(ws: Worksheet) -> list[tuple[int, str, str]]:
    """SPEC §5.5 — count_distinct, count_errors. Status column has
    duplicate values + a few error cells. Rules on H1/H2 cells assert
    exact counts that fail."""
    _write_header(ws, ["Status"])
    statuses = ["A", "A", "B", "B", "C", "A", "B", "C", "A", "B",
                "A", "A", "B", "C", "A", "B", "C", "A", "B", "C"]
    for idx, s in enumerate(statuses, start=2):
        ws.cell(row=idx, column=1, value=s)
    # Inject 2 error cells.
    _set_error_cell(ws, row=23, col=1, code="#N/A")
    _set_error_cell(ws, row=24, col=1, code="#REF!")
    # Scalar comparison cells.
    ws.cell(row=1, column=8, value=99)   # H1: should equal count_distinct=3 → fail
    ws.cell(row=2, column=8, value=99)   # H2: should equal count_errors=2 → fail
    return [
        (1, "H", "distinct-count"),
        (2, "H", "errors-count"),
    ]


# === Build orchestration ====================================================

# Manifest of (sheet_name, builder_fn). Order matters: it determines
# the workbook's xl/workbook.xml `<sheets>` order, which the validator
# uses for "first visible sheet" defaults (irrelevant here — every
# rule qualifies its scope explicitly).
_SHEET_BUILDERS = [
    ("CompareScalar", build_compare_scalar),
    ("Between", build_between),
    ("InList", build_in_list),
    ("TypeGuards", build_type_guards),
    ("Regex", build_regex),
    ("LenText", build_len_text),
    ("Dates", build_dates),
    ("AggregatesBasic", build_aggregates_basic),
    ("AggregateArith", build_aggregate_arith),
    ("GroupBy", build_group_by),
    ("Composite", build_composite),
    ("OrComposite", build_or_composite),
    ("WhenFilter", build_when_filter),
    ("Tolerance", build_tolerance),
    ("MergedData", build_merged_data),
    ("RangeRef", build_range_ref),
    ("RowRef", build_row_ref),
    ("SheetScope", build_sheet_scope),
    ("ExcelTable", build_excel_table),
    ("ErrorCells", build_error_cells),
    ("HiddenRows", build_hidden_rows),
    ("RequiredSkipEmpty", build_required_skip_empty),
    ("Apo's Sheet", build_apostrophe_sheet),  # apostrophe in name
    ("CrossSheetA", build_cross_sheet_a),
    ("CrossSheetB", build_cross_sheet_b),
    ("CountDistinct", build_count_distinct_errors),
]


def build_workbook() -> tuple[Workbook, dict[str, list[tuple[int, str, str]]]]:
    wb = Workbook()
    # Drop the default 'Sheet'.
    default = wb.active
    wb.remove(default)
    expected_per_sheet: dict[str, list[tuple[int, str, str]]] = {}
    for sheet_name, builder in _SHEET_BUILDERS:
        ws = wb.create_sheet(title=sheet_name)
        expected_per_sheet[sheet_name] = builder(ws)
    return wb, expected_per_sheet


OUT_EXPECTED = OUT_DIR / "comprehensive-rules.expected.json"


def _rule_severity_index() -> dict[str, str]:
    """Read `comprehensive-rules.rules.json` and build {rule_id:
    severity} so the manifest's summary_floor can be derived from
    actual rule severities — not from string-prefix heuristics on
    rule_id names. L4 (Sarcasmotron iter-1).

    Honors `defaults.severity` (SPEC §3 default-block) before falling
    back to "error" — addresses the iter-1 verifier's note that the
    floor would mis-classify if a rules-file ever set
    `defaults: {severity: warning}` and omitted per-rule severity."""
    import json
    rules_path = OUT_DIR / "comprehensive-rules.rules.json"
    doc = json.loads(rules_path.read_text())
    default_severity = doc.get("defaults", {}).get("severity", "error")
    out: dict[str, str] = {}
    for rule in doc.get("rules", []):
        rid = rule.get("id")
        sev = rule.get("severity", default_severity)
        if rid:
            out[rid] = sev
    # Engine auto-emit: cell-error has fixed severity=error.
    out["cell-error"] = "error"
    return out


def write_expected_manifest(expected: dict[str, list[tuple[int, str, str]]]) -> None:
    """Dump the generator's predicted findings for the test to assert
    against. Format documented inline; consumed by
    `TestComprehensiveRulesCoverage.test_all_rule_families_fire`.

    L8 (Sarcasmotron iter-1): grouped findings (SPEC §7.1.3) carry
    `row=null`, `column=null` in the validator output. The generator
    emits `row=0, column=""` as a sentinel; the test method
    `test_sampled_cells_flagged` translates `row=0` → `None` and falls
    back to (sheet, rule_id) match for these tuples."""
    import json
    sampled = []
    rule_id_counts: dict[str, int] = {}
    severity = _rule_severity_index()
    min_errors = 0
    min_warnings = 0
    for sheet, viols in expected.items():
        for row, col, rid in viols:
            rule_id_counts[rid] = rule_id_counts.get(rid, 0) + 1
            sampled.append({
                "sheet": sheet, "row": row, "column": col, "rule_id": rid,
            })
            sev = severity.get(rid, "error")
            if sev == "error":
                min_errors += 1
            elif sev == "warning":
                min_warnings += 1
    OUT_EXPECTED.write_text(json.dumps({
        "_comment": (
            "Auto-generated by make_comprehensive_fixture.py — do NOT "
            "edit by hand. Re-run the generator to refresh. The "
            "`sampled_findings` list uses row=0 / column='' as a "
            "SENTINEL for grouped findings (SPEC §7.1.3 row=null / "
            "column=null in validator output); the test's "
            "`test_sampled_cells_flagged` method translates these "
            "sentinels and matches by (sheet, rule_id) only. The "
            "summary_floor values are derived by iterating the rules "
            "file and counting predicted findings per `severity` field "
            "(L4 fix)."
        ),
        "fixture": OUT_XLSX.name,
        "rules_file": "comprehensive-rules.rules.json",
        "rule_id_counts": rule_id_counts,
        "summary_floor": {
            "min_errors": min_errors,
            "min_warnings": min_warnings,
            "min_cell_errors": 3,
        },
        "sampled_findings": sampled,
    }, indent=2, sort_keys=False))
    print(f"Wrote {OUT_EXPECTED} ({len(sampled)} predicted findings; "
          f"errors≥{min_errors}, warnings≥{min_warnings})")


def _verify_error_cells_round_tripped() -> None:
    """S2 (Sarcasmotron iter-1): the error-cell write uses
    `cell._value = code; cell.data_type = 'e'` (private API). Re-open
    the saved workbook and verify each error cell round-tripped as
    `t="e"` — guards against silent openpyxl regressions where
    `_value` becomes property-validated and downgrades the cells to
    text. Without this assertion, the engine's `cell-error` auto-emit
    would silently stop firing and the test would still pass on
    NUMERIC rules."""
    from openpyxl import load_workbook
    wb = load_workbook(OUT_XLSX)
    if "ErrorCells" not in wb.sheetnames:
        raise AssertionError("ErrorCells sheet missing post-write")
    ws = wb["ErrorCells"]
    expected_codes = {(5, "#REF!"), (10, "#N/A"), (15, "#DIV/0!")}
    actual: set[tuple[int, str]] = set()
    for row, _expected_code in expected_codes:
        cell = ws.cell(row=row, column=1)
        if cell.data_type != "e":
            raise AssertionError(
                f"ErrorCells!A{row} round-tripped with data_type="
                f"{cell.data_type!r} (expected 'e'); openpyxl private-API "
                f"regression — error-cell auto-emit will fail silently"
            )
        actual.add((row, str(cell.value)))
    if actual != expected_codes:
        raise AssertionError(
            f"ErrorCells codes mismatch: actual={actual} expected={expected_codes}"
        )


def main() -> int:
    wb, expected = build_workbook()
    wb.save(OUT_XLSX)
    _verify_error_cells_round_tripped()
    write_expected_manifest(expected)
    total_violations = sum(len(v) for v in expected.values())
    print(f"Wrote {OUT_XLSX} ({OUT_XLSX.stat().st_size:,} bytes)")
    print(f"Sheets: {len(expected)} | Predicted violations: {total_violations}")
    for name, viols in expected.items():
        print(f"  {name:<22s} {len(viols):3d} expected findings")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
