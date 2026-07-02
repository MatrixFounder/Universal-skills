"""Unit tests for xlsx_recalc.verify_cached_values — no LibreOffice needed.

Regression suite for the LibreOffice 26.2 silent no-op: soffice exited
0 while every formula cell stayed uncached, and xlsx_recalc.py happily
printed "Recalculated.". The script now verifies the output itself and
must raise RecalcVerificationError when recalculation clearly did not
run (formula cells exist, none carries a cached value).

Run:
    cd skills/xlsx/scripts
    ./.venv/bin/python -m unittest tests.test_xlsx_recalc_verify
"""

from __future__ import annotations

import sys
import tempfile
import unittest
import zipfile
from pathlib import Path

HERE = Path(__file__).resolve().parent
SCRIPTS = HERE.parent  # skills/xlsx/scripts
sys.path.insert(0, str(SCRIPTS))

from openpyxl import Workbook  # noqa: E402

from xlsx_recalc import (  # noqa: E402
    RECALC_PROFILE_SEED,
    RecalcVerificationError,
    recalc,
    verify_cached_values,
)


class TestVerifyCachedValues(unittest.TestCase):
    def setUp(self) -> None:
        self._tmp = tempfile.TemporaryDirectory(prefix="recalc-verify-")
        self.tmp = Path(self._tmp.name)

    def tearDown(self) -> None:
        self._tmp.cleanup()

    def _save(self, wb: Workbook, name: str) -> Path:
        path = self.tmp / name
        wb.save(str(path))
        return path

    def test_uncached_formulas_raise(self) -> None:
        """openpyxl output (formulas, no <v>) = recalc did not happen."""
        wb = Workbook()
        ws = wb.active
        ws.append([1, 2, 3])
        ws["D1"] = "=SUM(A1:C1)"
        path = self._save(wb, "uncached.xlsx")
        with self.assertRaises(RecalcVerificationError) as ctx:
            verify_cached_values(path)
        self.assertIn("recalculation did not happen", str(ctx.exception))

    def test_cached_formulas_pass(self) -> None:
        """A workbook whose formula cells carry <v> passes verification."""
        wb = Workbook()
        ws = wb.active
        ws.append([1, 2, 3])
        ws["D1"] = "=SUM(A1:C1)"
        path = self._save(wb, "cached.xlsx")
        # Inject a cached value the way a real spreadsheet engine
        # would. NB: openpyxl itself already writes an EMPTY <v></v>
        # after <f> (the uncached signature) — replace the whole pair.
        patched = self.tmp / "cached_patched.xlsx"
        zin = zipfile.ZipFile(str(path))
        with zipfile.ZipFile(str(patched), "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == "xl/worksheets/sheet1.xml":
                    text = data.decode()
                    self.assertIn("<f>SUM(A1:C1)</f><v></v>", text)
                    data = text.replace(
                        "<f>SUM(A1:C1)</f><v></v>", "<f>SUM(A1:C1)</f><v>6</v>"
                    ).encode()
                zout.writestr(item, data)
        zin.close()
        verify_cached_values(patched)  # must not raise

    def test_no_formulas_pass(self) -> None:
        """Constants-only workbooks have nothing to verify."""
        wb = Workbook()
        wb.active.append([1, 2, 3])
        path = self._save(wb, "constants.xlsx")
        verify_cached_values(path)  # must not raise

    def test_multi_sheet_uncached_raise(self) -> None:
        """Formulas on a non-active sheet are checked too."""
        wb = Workbook()
        wb.active.append([1])
        ws2 = wb.create_sheet("Calc")
        ws2["A1"] = "=1+1"
        path = self._save(wb, "multi.xlsx")
        with self.assertRaises(RecalcVerificationError):
            verify_cached_values(path)

    def test_empty_string_cache_passes(self) -> None:
        """LibreOffice caches a formula evaluating to "" as
        <c t="str"><f>…</f><v></v></c> — an empty <v> WITH t="str".
        openpyxl's data_only reports that as None, so a value-based
        check would false-fail an all-blank-producing template; the
        XML-level check must accept it (VDD finding, 2026-07-02)."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "x"
        ws["B1"] = '=IF(A1="y","match","")'
        path = self._save(wb, "blank.xlsx")
        patched = self.tmp / "blank_cached.xlsx"
        zin = zipfile.ZipFile(str(path))
        with zipfile.ZipFile(str(patched), "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == "xl/worksheets/sheet1.xml":
                    text = data.decode()
                    # openpyxl writes <c r="B1"><f>…</f><v></v></c>;
                    # LibreOffice's empty-string cache adds t="str".
                    self.assertIn('<c r="B1">', text)
                    data = text.replace('<c r="B1">', '<c r="B1" t="str">').encode()
                zout.writestr(item, data)
        zin.close()
        verify_cached_values(patched)  # must not raise

    def test_openpyxl_empty_v_still_raises(self) -> None:
        """openpyxl's own output has an EMPTY <v></v> after every <f>
        (no t attribute) — that is the UNCACHED signature and must not
        be mistaken for LibreOffice's t="str" empty-string cache."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = 1
        ws["B1"] = "=A1*2"
        path = self._save(wb, "openpyxl_raw.xlsx")
        raw = zipfile.ZipFile(str(path)).read("xl/worksheets/sheet1.xml").decode()
        self.assertIn("<f>A1*2</f><v></v>", raw)  # lock the premise
        with self.assertRaises(RecalcVerificationError):
            verify_cached_values(path)

    def test_output_directory_rejected(self) -> None:
        """--output pointing at an existing directory must fail fast:
        shutil.move would otherwise silently drop the file INTO the
        directory (VDD finding, 2026-07-02). Raised before LibreOffice
        is ever invoked, so no soffice is needed here."""
        wb = Workbook()
        wb.active["A1"] = 1
        src = self._save(wb, "in.xlsx")
        out_dir = self.tmp / "results"
        out_dir.mkdir()
        with self.assertRaises(IsADirectoryError):
            recalc(src, out_dir, timeout=5)


class TestRecalcProfileSeed(unittest.TestCase):
    def test_seed_forces_ooxml_recalc_always(self) -> None:
        """The throwaway-profile seed must pin OOXMLRecalcMode=0
        ("recalculate always on load") — without it LibreOffice keeps
        stale cached values on .xlsx load."""
        xcu = RECALC_PROFILE_SEED["user/registrymodifications.xcu"]
        self.assertIn("OOXMLRecalcMode", xcu)
        self.assertIn("<value>0</value>", xcu)
        self.assertIn("/org.openoffice.Office.Calc/Formula/Load", xcu)


if __name__ == "__main__":
    unittest.main()
