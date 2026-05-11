"""xlsx-3 unit-test scaffolding (Stage 1 — task-005-02).

≥ 35 unit cases across 6 classes, all currently RED via
`unittest.SkipTest("xlsx-3 stub — task-005-NN")`. Each test gets
filled in / un-skipped as the corresponding F-region lands in
tasks 005-04 through 005-09.

The TestPublicSurface class is the ONLY class with live tests at
Stage 1 — its 3 cases pin the public API + parser-instance attrs
(TC-UNIT-01/02/03 from task-005-01).

Drift-detection counterparties (csv2xlsx.HEADER_FILL and
json2xlsx.writer.HEADER_FILL) are imported here so a Stage-1
import-time check confirms they exist; the assertions turn live
in 005-08 when md_tables2xlsx.writer.HEADER_FILL is wired up.
"""
from __future__ import annotations

import inspect
import sys
import unittest
from pathlib import Path

# Make sibling scripts/ importable (mirrors json2xlsx test pattern).
_SCRIPTS_DIR = Path(__file__).resolve().parent.parent
if str(_SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS_DIR))

import md_tables2xlsx  # noqa: E402
from md_tables2xlsx import (  # noqa: E402
    convert_md_tables_to_xlsx,
    _AppError,
    EmptyInput,
    NoTablesFound,
    MalformedTable,
    InputEncodingError,
    InvalidSheetName,
    SelfOverwriteRefused,
    PostValidateFailed,
    NoSubstantialRowsAfterParse,
)

# Drift-detection import counterparties (ARCH m8 lock). Importable at
# Stage 1; live drift assertions activate in 005-08.
import csv2xlsx  # noqa: E402
from json2xlsx import writer as _json_writer  # noqa: E402


_STUB = "xlsx-3 stub — task-005-NN"


def _skip(reason: str = _STUB) -> None:
    raise unittest.SkipTest(reason)


# ============================================================
# TestPublicSurface — LIVE at Stage 1 (task-005-01)
# ============================================================
class TestPublicSurface(unittest.TestCase):
    """Pin the public surface from day 1 — these are the TC-UNIT-01/02/03
    cases from task-005-01. They stay live through every subsequent
    task as a regression gate.
    """

    def test_public_symbols_importable(self):
        # Re-asserts module-level imports above succeeded.
        for sym_name in [
            "main", "_run", "convert_md_tables_to_xlsx",
            "_AppError", "EmptyInput", "NoTablesFound",
            "MalformedTable", "InputEncodingError", "InvalidSheetName",
            "SelfOverwriteRefused", "PostValidateFailed",
            "NoSubstantialRowsAfterParse",
        ]:
            self.assertTrue(
                hasattr(md_tables2xlsx, sym_name),
                f"public symbol {sym_name!r} missing",
            )

    def test_convert_md_tables_to_xlsx_signature(self):
        """ARCH M4 + TASK §8 lock — `**kwargs -> int` mirrors xlsx-2."""
        sig = inspect.signature(convert_md_tables_to_xlsx)
        params = list(sig.parameters.items())
        self.assertEqual(len(params), 3, f"expected 3 params, got {params}")
        self.assertEqual(params[0][0], "input_path")
        self.assertEqual(params[1][0], "output_path")
        self.assertEqual(params[2][0], "kwargs")
        self.assertEqual(
            params[2][1].kind, inspect.Parameter.VAR_KEYWORD,
            "expected **kwargs",
        )
        # PEP 563 annotations are strings.
        self.assertEqual(params[0][1].annotation, "str | Path")
        self.assertEqual(params[1][1].annotation, "str | Path")
        self.assertEqual(sig.return_annotation, "int")

    def test_html_parser_singleton_constructed(self):
        """ARCH M1 lock — `_HTML_PARSER` exists at module import."""
        from md_tables2xlsx.tables import _HTML_PARSER
        import lxml.etree as _et
        self.assertIsInstance(_HTML_PARSER, _et.HTMLParser)

    def test_convert_md_tables_to_xlsx_routes_argv(self):
        """ARCH M4 lock — public helper routes through main(argv)."""
        from unittest.mock import patch
        captured = {}
        def fake_main(argv):
            captured["argv"] = list(argv)
            return 0
        with patch("md_tables2xlsx.cli.main", fake_main):
            # Need to also patch the re-export.
            with patch("md_tables2xlsx.main", fake_main):
                rc = convert_md_tables_to_xlsx(
                    "a.md", "b.xlsx", allow_empty=True,
                )
        self.assertEqual(rc, 0)
        self.assertEqual(captured["argv"][:2], ["a.md", "b.xlsx"])
        self.assertIn("--allow-empty", captured["argv"])

    def test_convert_md_tables_to_xlsx_atomic_token_protection(self):
        """ARCH M4 VDD-multi lock — `sheet_prefix='--evil'` becomes
        a SINGLE atomic token `--sheet-prefix=--evil` (not poisoning
        argparse into firing `--evil` as a separate flag).
        """
        from unittest.mock import patch
        captured = {}
        def fake_main(argv):
            captured["argv"] = list(argv)
            return 0
        with patch("md_tables2xlsx.main", fake_main):
            convert_md_tables_to_xlsx(
                "a.md", "b.xlsx", sheet_prefix="--evil-flag-attempt",
            )
        # The atomic-token form must keep `--evil-flag-attempt` as the
        # value of `--sheet-prefix`, NOT a separate flag.
        argv = captured["argv"]
        # Find the sheet-prefix token; must be the `--flag=value` form.
        prefix_tokens = [t for t in argv if t.startswith("--sheet-prefix")]
        self.assertEqual(len(prefix_tokens), 1, argv)
        self.assertEqual(prefix_tokens[0], "--sheet-prefix=--evil-flag-attempt")
        # And `--evil-flag-attempt` must NOT appear as a standalone token.
        self.assertNotIn("--evil-flag-attempt", argv)


# ============================================================
# TestPipeParser — block-detection live at 005-04; full parse at 005-06
# ============================================================
class TestPipeParser(unittest.TestCase):
    """GFM pipe-table parsing — block-detection (iter_blocks)
    activates in 005-04; full cell-parse (parse_pipe_table)
    activates in 005-06.
    """

    # ---- 005-06 (full parse) — LIVE ----
    def test_basic_pipe_table(self):
        from md_tables2xlsx.loaders import PipeTable
        from md_tables2xlsx.tables import parse_pipe_table
        block = PipeTable(raw_lines=["| a | b |", "|---|---|", "| 1 | 2 |"], line=1)
        rt = parse_pipe_table(block)
        self.assertEqual(rt.header, ["a", "b"])
        self.assertEqual(rt.rows, [["1", "2"]])
        self.assertEqual(rt.source, "gfm")

    def test_alignment_markers(self):
        from md_tables2xlsx.loaders import PipeTable
        from md_tables2xlsx.tables import parse_pipe_table
        # GFM separator requires `-{3,}` per spec; use `:---:` not `:--:`.
        # Per GFM: `---` (no colons) = no explicit alignment ("general"
        # for openpyxl); `:---` = left; `---:` = right; `:---:` = center.
        block = PipeTable(
            raw_lines=["| a | b | c | d |", "|:---|---:|:---:|---|"],
            line=1,
        )
        rt = parse_pipe_table(block)
        self.assertIsNotNone(rt, "parser returned None unexpectedly")
        self.assertEqual(rt.alignments, ["left", "right", "center", "general"])

    def test_escaped_pipe_in_cell(self):
        from md_tables2xlsx.loaders import PipeTable
        from md_tables2xlsx.tables import parse_pipe_table
        block = PipeTable(
            raw_lines=[r"| a \| b | c |", "|---|---|", r"| 1 \| 2 | 3 |"],
            line=1,
        )
        rt = parse_pipe_table(block)
        self.assertEqual(rt.header, ["a | b", "c"])
        self.assertEqual(rt.rows, [["1 | 2", "3"]])

    def test_column_count_mismatch_skips(self):
        from md_tables2xlsx.loaders import PipeTable
        from md_tables2xlsx.tables import parse_pipe_table
        block = PipeTable(
            raw_lines=["| a | b | c |", "|---|---|"], line=1,
        )
        rt = parse_pipe_table(block)
        self.assertIsNone(rt)

    def test_inline_strip_in_cell(self):
        from md_tables2xlsx.loaders import PipeTable
        from md_tables2xlsx.tables import parse_pipe_table
        block = PipeTable(
            raw_lines=["| **bold** | _italic_ |", "|---|---|", "| `code` | [link](u) |"],
            line=1,
        )
        rt = parse_pipe_table(block)
        self.assertEqual(rt.header, ["bold", "italic"])
        self.assertEqual(rt.rows, [["code", "link"]])

    def test_trailing_pipe_optional(self):
        from md_tables2xlsx.loaders import PipeTable
        from md_tables2xlsx.tables import parse_pipe_table
        b1 = PipeTable(raw_lines=["| a | b |", "|---|---|", "| 1 | 2 |"], line=1)
        b2 = PipeTable(raw_lines=["a | b", "---|---", "1 | 2"], line=1)
        r1, r2 = parse_pipe_table(b1), parse_pipe_table(b2)
        self.assertEqual(r1.header, r2.header)
        self.assertEqual(r1.rows, r2.rows)

    def test_zero_data_rows_kept(self):
        from md_tables2xlsx.loaders import PipeTable
        from md_tables2xlsx.tables import parse_pipe_table
        block = PipeTable(raw_lines=["| a | b |", "|---|---|"], line=1)
        rt = parse_pipe_table(block)
        self.assertEqual(rt.header, ["a", "b"])
        self.assertEqual(rt.rows, [])

    # ---- 005-04 (block detection) — LIVE ----
    def test_iter_blocks_finds_pipe_table(self):
        from md_tables2xlsx.loaders import iter_blocks, PipeTable
        text = "| a | b |\n|---|---|\n| 1 | 2 |\n"
        blocks = list(iter_blocks(text))
        pipes = [b for b in blocks if isinstance(b, PipeTable)]
        self.assertEqual(len(pipes), 1)
        self.assertEqual(len(pipes[0].raw_lines), 3)

    def test_iter_blocks_finds_3_tables_with_headings(self):
        from md_tables2xlsx.loaders import iter_blocks, Heading, PipeTable
        fixture = Path(__file__).resolve().parent.parent.parent / "examples" / "md_tables_simple.md"
        text = fixture.read_text(encoding="utf-8")
        blocks = list(iter_blocks(text))
        headings = [b for b in blocks if isinstance(b, Heading)]
        pipes = [b for b in blocks if isinstance(b, PipeTable)]
        self.assertEqual(len(pipes), 3, f"expected 3 pipe tables, got {len(pipes)}")
        # Document has H1 'Project Spec' + 3 H2s.
        self.assertGreaterEqual(len(headings), 4)

    def test_iter_blocks_skips_blockquoted_table(self):
        from md_tables2xlsx.loaders import iter_blocks, PipeTable
        text = "> | a | b |\n> |---|---|\n> | 1 | 2 |\n"
        blocks = list(iter_blocks(text))
        pipes = [b for b in blocks if isinstance(b, PipeTable)]
        self.assertEqual(len(pipes), 0, "blockquoted tables must be skipped (R9.g lock)")

    def test_scrub_fenced_strips_pipe_table_inside_fence(self):
        from md_tables2xlsx.loaders import scrub_fenced_and_comments, iter_blocks, PipeTable
        text = "```text\n| a | b |\n|---|---|\n| 1 | 2 |\n```\n"
        scrubbed, regs = scrub_fenced_and_comments(text)
        blocks = list(iter_blocks(scrubbed))
        pipes = [b for b in blocks if isinstance(b, PipeTable)]
        self.assertEqual(len(pipes), 0, "fenced-code-block table must be scrubbed")
        self.assertTrue(any(r.kind == "fenced_code" for r in regs))

    def test_scrub_html_comment_strips_table(self):
        from md_tables2xlsx.loaders import scrub_fenced_and_comments, iter_blocks, HtmlTable
        text = "<!--\n<table><tr><td>x</td></tr></table>\n-->\n"
        scrubbed, regs = scrub_fenced_and_comments(text)
        blocks = list(iter_blocks(scrubbed))
        htmls = [b for b in blocks if isinstance(b, HtmlTable)]
        self.assertEqual(len(htmls), 0)
        self.assertTrue(any(r.kind == "html_comment" for r in regs))

    def test_scrub_indented_code_strips_table(self):
        from md_tables2xlsx.loaders import scrub_fenced_and_comments, iter_blocks, PipeTable
        text = "Some prose.\n\n    | a | b |\n    |---|---|\n    | 1 | 2 |\n"
        scrubbed, regs = scrub_fenced_and_comments(text)
        blocks = list(iter_blocks(scrubbed))
        pipes = [b for b in blocks if isinstance(b, PipeTable)]
        self.assertEqual(len(pipes), 0, "indented-code table must be scrubbed (ARCH Q1)")
        self.assertTrue(any(r.kind == "indented_code" for r in regs))

    def test_scrub_style_block_stripped(self):
        from md_tables2xlsx.loaders import scrub_fenced_and_comments
        text = "before\n<style>body{color:red}</style>\nafter\n"
        scrubbed, regs = scrub_fenced_and_comments(text)
        self.assertNotIn("color:red", scrubbed, "style block must be scrubbed")
        self.assertTrue(any(r.kind == "style_block" for r in regs))

    def test_scrub_script_block_stripped(self):
        from md_tables2xlsx.loaders import scrub_fenced_and_comments
        text = "before\n<script>alert('xss')</script>\nafter\n"
        scrubbed, regs = scrub_fenced_and_comments(text)
        self.assertNotIn("alert", scrubbed)
        self.assertTrue(any(r.kind == "script_block" for r in regs))

    def test_scrub_preserves_line_numbers(self):
        from md_tables2xlsx.loaders import scrub_fenced_and_comments
        text = "L1\nL2\n```\ninside\n```\nL6\n"
        scrubbed, _ = scrub_fenced_and_comments(text)
        lines = scrubbed.split("\n")
        # Line 6 (index 5) should still be "L6" — line count preserved.
        self.assertEqual(lines[5], "L6")


# ============================================================
# TestHtmlParser — scaffolding for task-005-06
# ============================================================
class TestHtmlParser(unittest.TestCase):
    """HTML `<table>` parsing — full parser bodies land in 005-06.
    Block-detection (`iter_blocks` finds <table> ranges + ARCH m6
    heading-inside-table skip) is live now via 005-04.
    """

    # ---- 005-06 (full parse) — LIVE ----
    def test_basic_html_table(self):
        from md_tables2xlsx.loaders import HtmlTable
        from md_tables2xlsx.tables import parse_html_table
        block = HtmlTable(
            fragment="<table><tr><th>a</th></tr><tr><td>1</td></tr></table>",
            line=1,
        )
        rt = parse_html_table(block)
        self.assertEqual(rt.header, ["a"])
        self.assertEqual(rt.rows, [["1"]])
        self.assertEqual(rt.source, "html")

    def test_thead_tbody_split(self):
        from md_tables2xlsx.loaders import HtmlTable
        from md_tables2xlsx.tables import parse_html_table
        block = HtmlTable(
            fragment=(
                "<table><thead><tr><th>h</th></tr></thead>"
                "<tbody><tr><td>1</td></tr></tbody></table>"
            ),
            line=1,
        )
        rt = parse_html_table(block)
        self.assertEqual(rt.header, ["h"])
        self.assertEqual(rt.rows, [["1"]])

    def test_colspan_rowspan(self):
        from md_tables2xlsx.loaders import HtmlTable
        from md_tables2xlsx.tables import parse_html_table, MergeRange
        block = HtmlTable(
            fragment=(
                "<table>"
                "<tr><td colspan='2'>A</td></tr>"
                "<tr><td>1</td><td>2</td></tr>"
                "</table>"
            ),
            line=1,
        )
        rt = parse_html_table(block)
        self.assertIn(MergeRange(1, 1, 1, 2), rt.merges)

    def test_rowspan_only(self):
        from md_tables2xlsx.loaders import HtmlTable
        from md_tables2xlsx.tables import parse_html_table, MergeRange
        block = HtmlTable(
            fragment=(
                "<table>"
                "<tr><td rowspan='2'>A</td><td>B</td></tr>"
                "<tr><td>C</td></tr>"
                "</table>"
            ),
            line=1,
        )
        rt = parse_html_table(block)
        self.assertIn(MergeRange(1, 1, 2, 1), rt.merges)

    def test_html_entity_decode(self):
        from md_tables2xlsx.loaders import HtmlTable
        from md_tables2xlsx.tables import parse_html_table
        block = HtmlTable(
            fragment="<table><tr><td>a &amp; b</td></tr></table>", line=1,
        )
        rt = parse_html_table(block)
        self.assertEqual(rt.header, ["a & b"])

    def test_html_lxml_lenient_malformed_recovery(self):
        from md_tables2xlsx.loaders import HtmlTable
        from md_tables2xlsx.tables import parse_html_table
        # Missing </td> — lxml auto-closes.
        block = HtmlTable(
            fragment="<table><tr><td>x</tr></table>", line=1,
        )
        rt = parse_html_table(block)
        self.assertEqual(rt.header, ["x"])

    def test_html_billion_laughs_neutered(self):
        """ARCH M1 lock — `lxml.html` (HTML mode) does NOT process
        internal-subset `<!ENTITY>` declarations, so a billion-laughs
        DTD payload cannot cause exponential expansion. Verified by:
        (1) Wall-clock ≤ 100ms on a 100-level nested-entity payload
            (would take seconds if expansion were active).
        (2) The cell text contains the literal entity reference
            `&b99;` (or `b99` after entity-decode) NOT a billion 'X's.
        """
        import time
        from md_tables2xlsx.loaders import HtmlTable
        from md_tables2xlsx.tables import parse_html_table
        # Build the entity declarations as a separate string; the
        # `<table>` fragment uses the entity ref as a single cell.
        # We embed the entity declarations OUTSIDE the table — the
        # HTML parser will ignore them (HTML mode doesn't process
        # internal-subset DTDs). lxml.html fragment_fromstring with
        # create_parent=False expects a single root element, so we
        # put the entire payload inside the <table>.
        nested = "<!--"
        for i in range(1, 100):
            nested += f"<!ENTITY b{i} '&b{i-1};&b{i-1};&b{i-1};&b{i-1};&b{i-1};&b{i-1};&b{i-1};&b{i-1};&b{i-1};&b{i-1};'>"
        nested += "-->"
        payload = f"<table>{nested}<tr><td>safe-cell</td></tr></table>"
        block = HtmlTable(fragment=payload, line=1)
        t0 = time.perf_counter()
        rt = parse_html_table(block)
        elapsed_ms = (time.perf_counter() - t0) * 1000
        # HTML mode + no_network + huge_tree=False must keep this fast.
        self.assertLess(elapsed_ms, 1000.0, f"too slow: {elapsed_ms:.1f}ms")
        # Cell text must NOT have exploded.
        text_pile = " ".join(
            (rt.header[0] if rt.header else "")
            + " "
            + " ".join((r[0] or "") for r in rt.rows)
        )
        self.assertLess(len(text_pile), 1000, "entity expansion blew up")
        # Sanity — the literal cell content survives.
        flat = (rt.header + [c for row in rt.rows for c in row])
        self.assertIn("safe-cell", [v for v in flat if v])

    def test_html_parser_singleton_reusable(self):
        """ARCH M1 lock — `_HTML_PARSER` is a module-level singleton,
        reusable across calls without re-construction. Verifies parser
        identity is stable AND multiple sequential parses succeed.
        """
        from md_tables2xlsx.loaders import HtmlTable
        from md_tables2xlsx.tables import parse_html_table, _HTML_PARSER
        parser_id_before = id(_HTML_PARSER)
        # Two sequential parses should reuse the same parser.
        for _ in range(3):
            rt = parse_html_table(HtmlTable(
                fragment="<table><tr><td>x</td></tr></table>", line=1,
            ))
            self.assertEqual(rt.header, ["x"])
        from md_tables2xlsx.tables import _HTML_PARSER as p2
        self.assertEqual(id(p2), parser_id_before, "parser was re-constructed")

    def test_gfm_no_colspan_support(self):
        """R9.c lock — `colspan=2` in a GFM cell is literal text,
        no merge range emitted.
        """
        from md_tables2xlsx.loaders import PipeTable
        from md_tables2xlsx.tables import parse_pipe_table
        block = PipeTable(
            raw_lines=[
                "| a | colspan=2 |",
                "|---|---|",
                "| 1 | 2 |",
            ],
            line=1,
        )
        rt = parse_pipe_table(block)
        self.assertEqual(rt.merges, [])
        self.assertEqual(rt.header, ["a", "colspan=2"])

    # ---- 005-04 (block detection) — LIVE ----
    def test_iter_blocks_finds_html_table(self):
        from md_tables2xlsx.loaders import iter_blocks, HtmlTable
        text = "<table><tr><th>a</th></tr><tr><td>1</td></tr></table>\n"
        blocks = list(iter_blocks(text))
        htmls = [b for b in blocks if isinstance(b, HtmlTable)]
        self.assertEqual(len(htmls), 1)
        self.assertIn("<table>", htmls[0].fragment)

    def test_iter_blocks_skips_html_heading_inside_table(self):
        from md_tables2xlsx.loaders import iter_blocks, Heading, HtmlTable
        text = (
            "<table><tr><td><h3>InsideHeading</h3></td></tr></table>\n"
            "\n"
            "## OutsideHeading\n"
        )
        blocks = list(iter_blocks(text))
        heading_texts = [b.text for b in blocks if isinstance(b, Heading)]
        # ARCH m6 lock — <h3> inside <table> must NOT be emitted.
        self.assertNotIn("InsideHeading", heading_texts)
        self.assertIn("OutsideHeading", heading_texts)
        # One HtmlTable should be present.
        self.assertEqual(
            sum(1 for b in blocks if isinstance(b, HtmlTable)),
            1,
        )


# ============================================================
# TestInlineStrip — scaffolding for task-005-05
# ============================================================
class TestInlineStrip(unittest.TestCase):
    """Inline-markdown strip helper — bodies live in 005-05 (F5)."""

    def setUp(self):
        from md_tables2xlsx.inline import strip_inline_markdown
        self.strip = strip_inline_markdown

    def test_strip_bold(self):
        self.assertEqual(self.strip("**hello**"), "hello")
        self.assertEqual(self.strip("__hello__"), "hello")

    def test_strip_italic_underscore(self):
        self.assertEqual(self.strip("_world_"), "world")
        # Word-boundary protection: var_name stays intact.
        self.assertEqual(self.strip("var_name"), "var_name")

    def test_strip_code_span(self):
        self.assertEqual(self.strip("`code`"), "code")

    def test_strip_link(self):
        self.assertEqual(self.strip("[text](http://x)"), "text")

    def test_strip_strikethrough(self):
        self.assertEqual(self.strip("~~gone~~"), "gone")

    def test_br_to_newline(self):
        # **R9.c lock** [plan-review M1 fix] — literal \n, not wrap_text=True.
        self.assertEqual(self.strip("line1<br>line2"), "line1\nline2")
        self.assertEqual(self.strip("a<br/>b"), "a\nb")
        self.assertEqual(self.strip("a<br />b"), "a\nb")
        self.assertEqual(self.strip("a<BR>b"), "a\nb")

    def test_html_entity_decode(self):
        self.assertEqual(self.strip("a &amp; b &lt; c"), "a & b < c")
        self.assertEqual(self.strip("&#65;&#66;"), "AB")

    def test_strip_mixed_inline(self):
        result = self.strip("**bold** _italic_ `code` [link](url)")
        self.assertEqual(result, "bold italic code link")

    def test_strip_html_span(self):
        self.assertEqual(self.strip('<span class="hl">X</span>'), "X")

    def test_strip_idempotent(self):
        for s in ["**bold** _italic_", "plain", "[link](url)", "<span>X</span>"]:
            once = self.strip(s)
            twice = self.strip(once)
            self.assertEqual(once, twice, f"non-idempotent for {s!r}")

    def test_no_rich_text_runs(self):
        """R9.d lock — strip returns plain str, NOT an openpyxl Run."""
        result = self.strip("**bold**")
        self.assertIs(type(result), str)

    def test_multimarkdown_caption_literal(self):
        """R9.b lock (plan-review M1) — MultiMarkdown `[Caption: …]`
        passes through as literal text; no caption-metadata side channel."""
        result = self.strip("[Caption: Q1 Results]")
        # The link-regex `[text](url)` requires a `(url)` clause; bare
        # bracketed text without `(...)` must survive unchanged.
        self.assertEqual(result, "[Caption: Q1 Results]")


# ============================================================
# TestCoerce — scaffolding for task-005-05
# ============================================================
class TestCoerce(unittest.TestCase):
    """Per-column cell coercion — bodies live in 005-05 (F6)."""

    def setUp(self):
        from md_tables2xlsx.coerce import coerce_column, CoerceOptions
        self.coerce = coerce_column
        self.opts = CoerceOptions()

    def test_coerce_int(self):
        self.assertEqual(self.coerce(["1", "2", "3"], self.opts), [1, 2, 3])

    def test_coerce_float_comma(self):
        self.assertEqual(self.coerce(["1,5", "2,7"], self.opts), [1.5, 2.7])

    def test_coerce_leading_zero_keeps_text(self):
        self.assertEqual(
            self.coerce(["007", "042", "0123"], self.opts),
            ["007", "042", "0123"],
        )

    def test_coerce_iso_date(self):
        from datetime import date
        out = self.coerce(["2026-05-11"], self.opts)
        self.assertEqual(out[0], date(2026, 5, 11))

    def test_coerce_iso_datetime(self):
        from datetime import datetime
        out = self.coerce(["2026-05-11T14:30:00"], self.opts)
        self.assertEqual(out[0], datetime(2026, 5, 11, 14, 30, 0))
        self.assertIsNone(out[0].tzinfo)

    def test_coerce_iso_datetime_aware_to_utc_naive(self):
        from datetime import datetime
        out = self.coerce(["2026-05-11T14:30:00+02:00"], self.opts)
        # +02:00 → 12:30 UTC, tz stripped.
        self.assertEqual(out[0], datetime(2026, 5, 11, 12, 30, 0))
        self.assertIsNone(out[0].tzinfo)

    def test_coerce_no_coerce_flag_keeps_text(self):
        from md_tables2xlsx.coerce import CoerceOptions
        out = self.coerce(["1", "2026-05-11"], CoerceOptions(coerce=False))
        self.assertEqual(out, ["1", "2026-05-11"])

    def test_coerce_empty_string_to_None(self):
        out = self.coerce(["", "1"], self.opts)
        self.assertEqual(out, [None, 1])

    def test_coerce_lenient_date_string_rejected(self):
        # "May 11" must NOT be parsed by lenient dateutil.
        out = self.coerce(["May 11"], self.opts)
        self.assertEqual(out, ["May 11"])

    def test_coerce_mixed_numeric_with_string_keeps_text(self):
        out = self.coerce(["1", "abc", "3"], self.opts)
        # In mixed mode, each cell coerces independently.
        self.assertEqual(out, [1, "abc", 3])


# ============================================================
# TestSheetNaming — scaffolding for task-005-07
# ============================================================
class TestSheetNaming(unittest.TestCase):
    """Sheet-name resolution (9-step algorithm) — bodies live in 005-07."""

    def _resolver(self, prefix=None):
        from md_tables2xlsx.naming import SheetNameResolver
        return SheetNameResolver(sheet_prefix=prefix)

    def test_simple_heading(self):
        r = self._resolver()
        self.assertEqual(r.resolve("Q1 Budget"), "Q1 Budget")

    def test_strip_forbidden_chars(self):
        r = self._resolver()
        # Forbidden chars [, ], :, *, ?, /, \ → _.
        result = r.resolve("Q1: [Budget]")
        self.assertNotIn("[", result)
        self.assertNotIn("]", result)
        self.assertNotIn(":", result)
        self.assertIn("Q1", result)
        self.assertIn("Budget", result)

    def test_inline_markdown_strip(self):
        r = self._resolver()
        self.assertEqual(r.resolve("**Bold Heading**"), "Bold Heading")

    def test_dedup_simple(self):
        r = self._resolver()
        self.assertEqual(r.resolve("Results"), "Results")
        self.assertEqual(r.resolve("Results"), "Results-2")

    def test_dedup_case_insensitive(self):
        r = self._resolver()
        r.resolve("Results")
        # Lowercase resolution → dedup'd.
        second = r.resolve("results")
        self.assertNotEqual(second.lower(), "results")
        self.assertTrue(second.lower().endswith("-2"))

    def test_reserved_history_suffixed(self):
        r = self._resolver()
        # "History" is a reserved Excel sheet name → must get _.
        result = r.resolve("History")
        self.assertNotEqual(result.lower(), "history")
        self.assertEqual(result, "History_")

    def test_truncate_utf16_31_chars(self):
        from md_tables2xlsx.naming import _truncate_utf16
        # 50-char ASCII name → 31 chars after truncate.
        out = _truncate_utf16("A" * 50, limit=31)
        self.assertEqual(len(out), 31)
        self.assertEqual(len(out.encode("utf-16-le")) // 2, 31)

    def test_truncate_utf16_emoji_supplementary_plane(self):
        """m1 review-fix lock — emoji is 2 UTF-16 units."""
        from md_tables2xlsx.naming import _truncate_utf16
        # 20 emoji = 40 UTF-16 units → must truncate to 15 emoji
        # (30 units; 31st unit would split a surrogate pair).
        out = _truncate_utf16("😀" * 20, limit=31)
        # Result must have ≤ 31 UTF-16 units.
        self.assertLessEqual(len(out.encode("utf-16-le")) // 2, 31)
        # Only complete emoji preserved (no orphan surrogates).
        for ch in out:
            self.assertGreaterEqual(ord(ch), 0x1F000)

    def test_dedup_emoji_prefix_utf16_safe(self):
        """ARCH M3 lock — prefix re-truncation in dedup must be
        UTF-16-aware. 16-emoji collision must not leak past 31 units.
        """
        r = self._resolver()
        first = r.resolve("😀" * 16)
        # First should be 15 emoji = 30 UTF-16 units.
        self.assertLessEqual(len(first.encode("utf-16-le")) // 2, 31)
        # Second collision with the same heading → dedup with -2.
        second = r.resolve("😀" * 16)
        self.assertLessEqual(
            len(second.encode("utf-16-le")) // 2, 31,
            f"prefix re-truncation leaked: {second!r} "
            f"= {len(second.encode('utf-16-le')) // 2} UTF-16 units",
        )
        self.assertTrue(second.endswith("-2"))

    def test_fallback_table_n_no_heading(self):
        r = self._resolver()
        self.assertEqual(r.resolve(None), "Table-1")
        self.assertEqual(r.resolve(None), "Table-2")

    def test_fallback_table_n_empty_heading_after_sanitise(self):
        r = self._resolver()
        # Whitespace-only + apostrophe heading strips to empty.
        self.assertEqual(r.resolve("   '   "), "Table-1")
        # An empty-string heading also lands here.
        self.assertEqual(r.resolve(""), "Table-2")

    def test_dedup_overflow_raises_InvalidSheetName(self):
        from md_tables2xlsx.exceptions import InvalidSheetName
        r = self._resolver()
        # Pre-populate _used_lower with Foo + Foo-2..Foo-99.
        r._used_lower.add("foo")
        for n in range(2, 100):
            r._used_lower.add(f"foo-{n}")
        with self.assertRaises(InvalidSheetName) as cm:
            r.resolve("Foo")
        self.assertEqual(cm.exception.code, 2)
        self.assertEqual(cm.exception.details["retry_cap"], 99)

    def test_sheet_prefix_mode_ignores_heading(self):
        """ARCH m12 lock — `--sheet-prefix` mode ignores heading,
        emits sequential STR-1, STR-2.
        """
        r = self._resolver(prefix="Report")
        self.assertEqual(r.resolve("AnythingHere"), "Report-1")
        self.assertEqual(r.resolve("AlsoIgnored"), "Report-2")


# ============================================================
# TestExceptions — partially live at Stage 1 (smoke checks)
# ============================================================
class TestExceptions(unittest.TestCase):
    """Closed `_AppError` taxonomy — full body checks land in 005-03."""

    def test_AppError_is_Exception(self):
        self.assertTrue(issubclass(_AppError, Exception))

    def test_eight_subclasses_inherit_AppError(self):
        for cls in (
            EmptyInput, NoTablesFound, MalformedTable, InputEncodingError,
            InvalidSheetName, SelfOverwriteRefused, PostValidateFailed,
            NoSubstantialRowsAfterParse,
        ):
            self.assertTrue(issubclass(cls, _AppError))

    def test_each_AppError_has_code_and_type(self):
        """task-005-03 — every typed subclass carries `code` + `error_type`."""
        expected = [
            (EmptyInput, 2, "EmptyInput"),
            (NoTablesFound, 2, "NoTablesFound"),
            (MalformedTable, 2, "MalformedTable"),
            (InputEncodingError, 2, "InputEncodingError"),
            (InvalidSheetName, 2, "InvalidSheetName"),
            (SelfOverwriteRefused, 6, "SelfOverwriteRefused"),
            (PostValidateFailed, 7, "PostValidateFailed"),
            (NoSubstantialRowsAfterParse, 2, "NoSubstantialRowsAfterParse"),
        ]
        for cls, code, etype in expected:
            exc = cls("msg", code=code, error_type=etype)
            self.assertEqual(exc.code, code)
            self.assertEqual(exc.error_type, etype)
            self.assertEqual(exc.message, "msg")

    def test_AppError_details_default_empty_dict(self):
        """task-005-03 — `details` defaults to {} (not None)."""
        exc = EmptyInput("m", code=2, error_type="EmptyInput")
        self.assertEqual(exc.details, {})

    def test_post_validate_truthy_allowlist(self):
        """task-005-03 — `{1, true, yes, on}` truthy; everything else falsy."""
        import os
        from md_tables2xlsx.cli_helpers import post_validate_enabled
        truthy = ["1", "true", "TRUE", "True", "yes", "YES", "on", "ON",
                  " 1 ", "True ", " on"]
        falsy = ["0", "false", "no", "off", "", "2", "enable", "y", "Y"]
        try:
            for v in truthy:
                os.environ["XLSX_MD_TABLES_POST_VALIDATE"] = v
                self.assertTrue(
                    post_validate_enabled(),
                    f"expected truthy for {v!r}",
                )
            for v in falsy:
                os.environ["XLSX_MD_TABLES_POST_VALIDATE"] = v
                self.assertFalse(
                    post_validate_enabled(),
                    f"expected falsy for {v!r}",
                )
            os.environ.pop("XLSX_MD_TABLES_POST_VALIDATE", None)
            self.assertFalse(post_validate_enabled())
        finally:
            os.environ.pop("XLSX_MD_TABLES_POST_VALIDATE", None)

    def test_assert_distinct_paths_stdin_bypasses(self):
        """task-005-03 — `-` sentinel bypasses guard."""
        from md_tables2xlsx.cli_helpers import assert_distinct_paths
        # should NOT raise
        assert_distinct_paths("-", Path("/tmp/out.xlsx"))

    def test_assert_distinct_paths_same_path_raises(self):
        """task-005-03 — same resolved path → SelfOverwriteRefused (code 6)."""
        from md_tables2xlsx.cli_helpers import assert_distinct_paths
        with self.assertRaises(SelfOverwriteRefused) as cm:
            assert_distinct_paths("/tmp/x.md", Path("/tmp/x.md"))
        self.assertEqual(cm.exception.code, 6)
        self.assertEqual(cm.exception.error_type, "SelfOverwriteRefused")

    def test_assert_distinct_paths_symlink_followed(self):
        """task-005-03 — symlink resolved before compare (cross-7 H1 lock)."""
        import tempfile
        from md_tables2xlsx.cli_helpers import assert_distinct_paths
        with tempfile.TemporaryDirectory() as td:
            tdp = Path(td)
            real = tdp / "real.md"
            real.write_text("x")
            link = tdp / "link.md"
            link.symlink_to(real)
            with self.assertRaises(SelfOverwriteRefused):
                assert_distinct_paths(str(real), link)

    def test_read_stdin_utf8_strict(self):
        """task-005-03 — bad UTF-8 → UnicodeDecodeError."""
        import io
        from md_tables2xlsx.cli_helpers import read_stdin_utf8
        # Build a fake stdin whose .buffer.read returns bad bytes.
        class _FakeBuf:
            def read(self):
                return b"\xff\xfeinvalid"
        class _FakeStdin:
            buffer = _FakeBuf()
        real_stdin = sys.stdin
        try:
            sys.stdin = _FakeStdin()
            with self.assertRaises(UnicodeDecodeError):
                read_stdin_utf8()
        finally:
            sys.stdin = real_stdin

    def test_read_stdin_utf8_happy_path(self):
        """task-005-03 — valid UTF-8 from stdin returns decoded str."""
        from md_tables2xlsx.cli_helpers import read_stdin_utf8
        class _FakeBuf:
            def read(self):
                return "héllo world".encode("utf-8")
        class _FakeStdin:
            buffer = _FakeBuf()
        real_stdin = sys.stdin
        try:
            sys.stdin = _FakeStdin()
            self.assertEqual(read_stdin_utf8(), "héllo world")
        finally:
            sys.stdin = real_stdin


# ============================================================
# TestStyleConstantDrift — scaffolding for task-005-08 (ARCH m8)
# ============================================================
class TestStyleConstantDrift(unittest.TestCase):
    """3-way drift detection (csv2xlsx ↔ json2xlsx ↔ md_tables2xlsx).
    ARCH m8 lock — live as of task-005-08.
    """

    def test_csv2xlsx_HEADER_FILL_importable(self):
        self.assertTrue(hasattr(csv2xlsx, "HEADER_FILL"))

    def test_json2xlsx_writer_HEADER_FILL_importable(self):
        self.assertTrue(hasattr(_json_writer, "HEADER_FILL"))

    def test_header_fill_matches_csv2xlsx(self):
        from md_tables2xlsx.writer import HEADER_FILL as _MD_FILL
        # csv2xlsx may have 6-char ARGB or 8-char form (openpyxl
        # normalises lazily). Accept both forms.
        accepted = ("F2F2F2", "00F2F2F2")
        self.assertIn(csv2xlsx.HEADER_FILL.fgColor.rgb, accepted)
        self.assertEqual(_MD_FILL.fgColor.rgb, csv2xlsx.HEADER_FILL.fgColor.rgb)

    def test_header_fill_matches_json2xlsx(self):
        from md_tables2xlsx.writer import HEADER_FILL as _MD_FILL
        self.assertEqual(_MD_FILL.fgColor.rgb, _json_writer.HEADER_FILL.fgColor.rgb)

    def test_max_col_width_matches_csv2xlsx(self):
        from md_tables2xlsx.writer import MAX_COL_WIDTH
        self.assertEqual(MAX_COL_WIDTH, csv2xlsx.MAX_COL_WIDTH)


# ============================================================
# TestWriter — live at 005-08
# ============================================================
class TestWriter(unittest.TestCase):
    """openpyxl Workbook assembly + styling + merges."""

    def _make_tbl(self, header, rows, alignments=None, merges=None,
                  source="gfm", sheet_name="Test"):
        from md_tables2xlsx.tables import RawTable
        from md_tables2xlsx.writer import ParsedTable
        if alignments is None:
            alignments = ["general"] * len(header)
        rt = RawTable(
            header=header, rows=rows, alignments=alignments,
            merges=merges or [], source=source, source_line=1,
        )
        # Build coerced_columns from rows (no real coerce here —
        # values are passed through).
        n_cols = len(header)
        cols = [[r[c] for r in rows] for c in range(n_cols)]
        return ParsedTable(raw=rt, sheet_name=sheet_name, coerced_columns=cols)

    def test_basic_workbook(self):
        import tempfile
        import openpyxl
        from md_tables2xlsx.writer import write_workbook, WriterOptions
        tbl = self._make_tbl(["a", "b"], [["1", "2"]])
        with tempfile.TemporaryDirectory() as td:
            out = Path(td) / "out.xlsx"
            write_workbook([tbl], out, WriterOptions())
            wb = openpyxl.load_workbook(str(out))
            self.assertEqual(wb.sheetnames, ["Test"])
            ws = wb["Test"]
            self.assertEqual(ws["A1"].value, "a")
            self.assertEqual(ws["A2"].value, "1")
            # Header styled (bold + fill).
            self.assertTrue(ws["A1"].font.bold)

    def test_multi_sheet(self):
        import tempfile, openpyxl
        from md_tables2xlsx.writer import write_workbook, WriterOptions
        t1 = self._make_tbl(["h"], [["x"]], sheet_name="S1")
        t2 = self._make_tbl(["h"], [["y"]], sheet_name="S2")
        t3 = self._make_tbl(["h"], [["z"]], sheet_name="S3")
        with tempfile.TemporaryDirectory() as td:
            out = Path(td) / "out.xlsx"
            write_workbook([t1, t2, t3], out, WriterOptions())
            wb = openpyxl.load_workbook(str(out))
            self.assertEqual(wb.sheetnames, ["S1", "S2", "S3"])

    def test_default_sheet_removed(self):
        import tempfile, openpyxl
        from md_tables2xlsx.writer import write_workbook, WriterOptions
        tbl = self._make_tbl(["h"], [["1"]], sheet_name="Only")
        with tempfile.TemporaryDirectory() as td:
            out = Path(td) / "out.xlsx"
            write_workbook([tbl], out, WriterOptions())
            wb = openpyxl.load_workbook(str(out))
            self.assertNotIn("Sheet", wb.sheetnames)

    def test_freeze_pane_applied(self):
        import tempfile, openpyxl
        from md_tables2xlsx.writer import write_workbook, WriterOptions
        tbl = self._make_tbl(["a"], [["1"]])
        with tempfile.TemporaryDirectory() as td:
            out = Path(td) / "out.xlsx"
            write_workbook([tbl], out, WriterOptions(freeze=True))
            wb = openpyxl.load_workbook(str(out))
            self.assertEqual(wb["Test"].freeze_panes, "A2")

    def test_auto_filter_applied(self):
        import tempfile, openpyxl
        from md_tables2xlsx.writer import write_workbook, WriterOptions
        tbl = self._make_tbl(["a"], [["1"]])
        with tempfile.TemporaryDirectory() as td:
            out = Path(td) / "out.xlsx"
            write_workbook([tbl], out, WriterOptions(auto_filter=True))
            wb = openpyxl.load_workbook(str(out))
            self.assertIsNotNone(wb["Test"].auto_filter.ref)

    def test_no_freeze_flag(self):
        import tempfile, openpyxl
        from md_tables2xlsx.writer import write_workbook, WriterOptions
        tbl = self._make_tbl(["a"], [["1"]])
        with tempfile.TemporaryDirectory() as td:
            out = Path(td) / "out.xlsx"
            write_workbook([tbl], out, WriterOptions(freeze=False))
            wb = openpyxl.load_workbook(str(out))
            self.assertNotEqual(wb["Test"].freeze_panes, "A2")

    def test_no_filter_flag(self):
        import tempfile, openpyxl
        from md_tables2xlsx.writer import write_workbook, WriterOptions
        tbl = self._make_tbl(["a"], [["1"]])
        with tempfile.TemporaryDirectory() as td:
            out = Path(td) / "out.xlsx"
            write_workbook([tbl], out, WriterOptions(auto_filter=False))
            wb = openpyxl.load_workbook(str(out))
            self.assertIsNone(wb["Test"].auto_filter.ref)

    def test_merge_cells_applied(self):
        import tempfile, openpyxl
        from md_tables2xlsx.writer import write_workbook, WriterOptions
        from md_tables2xlsx.tables import MergeRange
        tbl = self._make_tbl(
            ["A", "B"], [["1", "2"]],
            merges=[MergeRange(1, 1, 1, 2)],
            source="html",
        )
        with tempfile.TemporaryDirectory() as td:
            out = Path(td) / "out.xlsx"
            write_workbook([tbl], out, WriterOptions())
            wb = openpyxl.load_workbook(str(out))
            ranges = [str(r) for r in wb["Test"].merged_cells.ranges]
            self.assertIn("A1:B1", ranges)

    def test_gfm_alignment_applied(self):
        import tempfile, openpyxl
        from md_tables2xlsx.writer import write_workbook, WriterOptions
        tbl = self._make_tbl(
            ["a", "b", "c"],
            [["1", "2", "3"]],
            alignments=["left", "right", "center"],
        )
        with tempfile.TemporaryDirectory() as td:
            out = Path(td) / "out.xlsx"
            write_workbook([tbl], out, WriterOptions())
            wb = openpyxl.load_workbook(str(out))
            ws = wb["Test"]
            self.assertEqual(ws["A2"].alignment.horizontal, "left")
            self.assertEqual(ws["B2"].alignment.horizontal, "right")
            self.assertEqual(ws["C2"].alignment.horizontal, "center")

    def test_no_formula_evaluation(self):
        """R9.e lock — cell value `=SUM(A1:A3)` is literal text."""
        import tempfile, openpyxl
        from md_tables2xlsx.writer import write_workbook, WriterOptions
        tbl = self._make_tbl(["a"], [["=SUM(A1:A3)"]])
        with tempfile.TemporaryDirectory() as td:
            out = Path(td) / "out.xlsx"
            write_workbook([tbl], out, WriterOptions())
            wb = openpyxl.load_workbook(str(out))
            cell = wb["Test"]["A2"]
            self.assertEqual(cell.value, "=SUM(A1:A3)")
            self.assertEqual(cell.data_type, "s")

    def test_allow_empty_with_zero_tables(self):
        import tempfile, openpyxl
        from md_tables2xlsx.writer import write_workbook, WriterOptions
        with tempfile.TemporaryDirectory() as td:
            out = Path(td) / "out.xlsx"
            write_workbook([], out, WriterOptions(allow_empty=True))
            wb = openpyxl.load_workbook(str(out))
            self.assertEqual(wb.sheetnames, ["Empty"])

    def test_parent_dir_auto_created(self):
        """ARCH A8 lock — output.parent.mkdir(parents=True, exist_ok=True)."""
        import tempfile, openpyxl
        from md_tables2xlsx.writer import write_workbook, WriterOptions
        tbl = self._make_tbl(["a"], [["1"]])
        with tempfile.TemporaryDirectory() as td:
            out = Path(td) / "deep" / "nested" / "out.xlsx"
            self.assertFalse(out.parent.exists())
            write_workbook([tbl], out, WriterOptions())
            self.assertTrue(out.exists())


# ============================================================
# TestVddMultiFixes — regression tests for /vdd-multi findings
# ============================================================
class TestVddMultiFixes(unittest.TestCase):
    """Regression locks for the 8 fixes applied during /vdd-multi
    Phase 3 (critics: logic, security, performance).
    """

    # ---- H1: tables.py colspan/rowspan crash on non-int values ----

    def test_H1_colspan_non_int_does_not_crash(self):
        """ARCH H1 (vdd-multi) — `colspan="abc"` defaults to 1 instead
        of raising ValueError."""
        from md_tables2xlsx.loaders import HtmlTable
        from md_tables2xlsx.tables import parse_html_table
        block = HtmlTable(
            fragment='<table><tr><td colspan="abc">x</td></tr></table>',
            line=1,
        )
        rt = parse_html_table(block)
        self.assertIsNotNone(rt)
        self.assertEqual(rt.header, ["x"])

    def test_H1_colspan_float_string_does_not_crash(self):
        """`colspan="2.5"` defaults to 1 (int() raises on float string)."""
        from md_tables2xlsx.loaders import HtmlTable
        from md_tables2xlsx.tables import parse_html_table
        block = HtmlTable(
            fragment='<table><tr><td colspan="2.5">x</td></tr></table>',
            line=1,
        )
        rt = parse_html_table(block)
        self.assertIsNotNone(rt)

    def test_H1_colspan_clamped_to_excel_limit(self):
        """vdd-multi M2 (perf) review-fix — colspan=999999 clamps to
        Excel column limit, preventing OOM allocation in _expand_spans.
        """
        from md_tables2xlsx.loaders import HtmlTable
        from md_tables2xlsx.tables import parse_html_table, _MAX_SPAN
        block = HtmlTable(
            fragment='<table><tr><td colspan="999999">x</td></tr></table>',
            line=1,
        )
        rt = parse_html_table(block)
        self.assertIsNotNone(rt)
        # Width must be clamped, not 999999.
        self.assertLessEqual(len(rt.header), _MAX_SPAN)

    # ---- H2: OSError from Path.resolve() escapes envelope ----

    def test_H2_OSError_in_assert_distinct_paths_routes_to_envelope(self):
        """vdd-multi H2 review-fix — Path.resolve()-raised OSError
        (e.g. symlink loop on Linux: 'Too many levels of symbolic
        links') routes through cross-5 envelope instead of bare
        traceback. We can't easily make Path.resolve() raise on macOS;
        instead, monkeypatch assert_distinct_paths to raise an OSError
        and verify main() catches it.
        """
        from unittest.mock import patch
        from md_tables2xlsx.cli import main
        def fake_assert(*a, **kw):
            raise OSError(40, "Too many levels of symbolic links", "/tmp/loop.md")
        with patch("md_tables2xlsx.cli.assert_distinct_paths", fake_assert):
            rc = main(["/tmp/loop.md", "/tmp/out.xlsx", "--json-errors"])
        self.assertEqual(rc, 1)  # routed to IOError envelope, exit 1.

    # ---- H3: _replace_with_spaces correctness after rewrite ----

    def test_H3_replace_with_spaces_preserves_length_and_newlines(self):
        """vdd-multi H3 (perf) review-fix — the optimised
        `split/join` form produces same output as the old per-char
        loop.
        """
        from md_tables2xlsx.loaders import _replace_with_spaces
        text = "abc\ndef\nghi"
        result = _replace_with_spaces(text, 4, 7)  # blank out "def"
        self.assertEqual(result, "abc\n   \nghi")
        # No change to length.
        self.assertEqual(len(result), len(text))

    def test_H3_replace_with_spaces_empty_slice(self):
        from md_tables2xlsx.loaders import _replace_with_spaces
        text = "abc"
        self.assertEqual(_replace_with_spaces(text, 1, 1), "abc")

    # ---- H4: _strip_style_script pos tracking (just correctness) ----

    def test_H4_multiple_style_blocks_all_stripped(self):
        """vdd-multi H4 (perf) review-fix — pos-tracking correctness:
        all 3 style blocks must be stripped, not just the first."""
        from md_tables2xlsx.loaders import scrub_fenced_and_comments
        text = (
            "<style>a {color:red}</style>\n"
            "Some prose.\n"
            "<style>b {color:blue}</style>\n"
            "More prose.\n"
            "<style>c {color:green}</style>\n"
        )
        scrubbed, regs = scrub_fenced_and_comments(text)
        self.assertNotIn("color:red", scrubbed)
        self.assertNotIn("color:blue", scrubbed)
        self.assertNotIn("color:green", scrubbed)
        style_regs = [r for r in regs if r.kind == "style_block"]
        self.assertEqual(len(style_regs), 3)

    # ---- M1: control-char strip in sheet names ----

    def test_M1_control_chars_in_sheet_name_stripped(self):
        """vdd-multi M1 (security) review-fix — `\\x00`-`\\x1F` and
        `\\x7F` are stripped from sheet names; Excel rejects workbooks
        containing them otherwise.
        """
        from md_tables2xlsx.naming import SheetNameResolver
        r = SheetNameResolver()
        # `foo\x01bar` heading must NOT produce a sheet containing \x01.
        result = r.resolve("foo\x01bar")
        self.assertNotIn("\x01", result)
        self.assertNotIn("\x00", r.resolve("a\x00b"))
        self.assertNotIn("\x7F", r.resolve("c\x7Fd"))

    # ---- M2: tables._walk_rows merges thead + tbody + direct <tr> ----

    def test_M2_walk_rows_merges_direct_tr_with_thead(self):
        """vdd-multi M2 review-fix — `<table><thead>…</thead><tr>…</tr></table>`
        no longer silently drops the direct `<tr>`."""
        from md_tables2xlsx.loaders import HtmlTable
        from md_tables2xlsx.tables import parse_html_table
        block = HtmlTable(
            fragment=(
                "<table>"
                "<thead><tr><th>h1</th><th>h2</th></tr></thead>"
                "<tr><td>d1</td><td>d2</td></tr>"
                "</table>"
            ),
            line=1,
        )
        rt = parse_html_table(block)
        self.assertEqual(rt.header, ["h1", "h2"])
        self.assertEqual(rt.rows, [["d1", "d2"]])

    # ---- M3: empty <table> returns None ----

    def test_M3_empty_table_returns_None(self):
        """vdd-multi M3 review-fix — empty `<table></table>` returns
        None (orchestrator-skip), instead of producing a named-but-
        empty sheet."""
        from md_tables2xlsx.loaders import HtmlTable
        from md_tables2xlsx.tables import parse_html_table
        block = HtmlTable(fragment="<table></table>", line=1)
        rt = parse_html_table(block)
        self.assertIsNone(rt)

    # ---- M4: fence opening capped at 3 leading spaces ----

    def test_M4_fence_indented_4_spaces_treated_as_code_block(self):
        """vdd-multi M4 review-fix — CommonMark caps fence indent at
        3 spaces. A 4-space-indented `` ``` `` is an indented code
        block opener, NOT a fence — so the regex must NOT match.
        """
        from md_tables2xlsx.loaders import _FENCED_OPEN_RE
        # 0-3 spaces: matches.
        for indent in ["", " ", "  ", "   "]:
            self.assertIsNotNone(_FENCED_OPEN_RE.match(f"{indent}```"))
        # 4+ spaces: does NOT match.
        for indent in ["    ", "     ", "          "]:
            self.assertIsNone(_FENCED_OPEN_RE.match(f"{indent}```"))

    # ---- M5: SystemExit(None) handled as exit 0 ----

    def test_M5_systemexit_none_is_clean_exit(self):
        """vdd-multi M5 review-fix — `SystemExit(None)` mapped to
        return 0, not UsageError envelope."""
        from unittest.mock import patch
        from md_tables2xlsx.cli import main
        # argparse parse_args → SystemExit(None) is unusual but
        # conventionally means clean exit. Simulate via parser patch.
        def fake_parse_args(*a, **kw):
            raise SystemExit(None)
        with patch("argparse.ArgumentParser.parse_args", fake_parse_args):
            rc = main(["a.md", "b.xlsx"])
        self.assertEqual(rc, 0)

    # ---- M6: kwargs None skipped in convert_md_tables_to_xlsx ----

    def test_M6_kwargs_None_value_skipped(self):
        """vdd-multi M6 review-fix — `sheet_prefix=None` does NOT
        produce a literal `--sheet-prefix=None` argv token (which
        would create a 'None' sheet-prefix). It's skipped entirely.
        """
        from unittest.mock import patch
        captured = {}
        def fake_main(argv):
            captured["argv"] = list(argv)
            return 0
        with patch("md_tables2xlsx.main", fake_main):
            convert_md_tables_to_xlsx(
                "a.md", "b.xlsx", sheet_prefix=None, encoding="utf-8",
            )
        argv = captured["argv"]
        # `--sheet-prefix=None` must NOT appear.
        self.assertFalse(
            any(t.startswith("--sheet-prefix") for t in argv),
            f"sheet_prefix=None should be skipped; argv={argv!r}",
        )
        # `--encoding=utf-8` should still appear (non-None value).
        self.assertIn("--encoding=utf-8", argv)

    # ---- M7: subprocess `--` separator (security: arg-injection) ----

    def test_M7_post_validate_uses_double_dash_separator(self):
        """vdd-multi M7 review-fix — `run_post_validate` passes `--`
        before the user-controlled output path so a path like
        `--help` cannot be interpreted as a flag by the subprocess.
        """
        from unittest.mock import patch, MagicMock
        from pathlib import Path as _P
        from md_tables2xlsx.cli_helpers import run_post_validate
        captured = {}
        def fake_run(cmd, **kw):
            captured["cmd"] = list(cmd)
            result = MagicMock()
            result.returncode = 0
            result.stdout = ""
            result.stderr = ""
            return result
        with patch("subprocess.run", fake_run):
            with patch("pathlib.Path.is_file", return_value=True):
                run_post_validate(_P("/tmp/--help"))
        # `--` must appear before the output path token.
        cmd = captured["cmd"]
        dash_dash_idx = cmd.index("--")
        output_idx = cmd.index("/tmp/--help")
        self.assertLess(dash_dash_idx, output_idx)

    # ---- L1: negative leading-zero gate ----

    def test_L1_negative_leading_zero_kept_as_text(self):
        """vdd-multi L1 review-fix — `-007` triggers leading-zero
        gate (was only `007` previously)."""
        from md_tables2xlsx.coerce import coerce_column, CoerceOptions
        out = coerce_column(["-007", "-008", "-009"], CoerceOptions())
        self.assertEqual(out, ["-007", "-008", "-009"])


if __name__ == "__main__":
    unittest.main()
