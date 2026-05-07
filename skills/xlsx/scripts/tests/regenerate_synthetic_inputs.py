"""Regenerate synthetic .xlsx/.xlsm input fixtures for xlsx-6 E2E tests.

Run from inside the skill:
    cd skills/xlsx/scripts
    ./.venv/bin/python tests/regenerate_synthetic_inputs.py

What this script generates (under `tests/golden/inputs/`):
    clean.xlsx          — single sheet, header + 10 data rows
    multi_sheet.xlsx    — 3 sheets ("Sheet1", "Sheet2", "Q1 2026"),
                          last name has whitespace to exercise quoted-
                          cell-syntax (`'Q1 2026'!A1`).
    hidden_first.xlsx   — Sheet1 state="hidden", Sheet2 visible (M2 lock).
    merged.xlsx         — single sheet with <mergeCell ref="A1:C3"/>.
    with_legacy.xlsx    — single sheet with 2 pre-existing legacy
                          comments via openpyxl.comments.Comment.
                          M-E DEVIATION: Excel-365 authoring unavailable
                          at fixture-creation time; this is a synthetic
                          openpyxl-generated equivalent. See
                          tests/golden/README.md §Provenance.
    macro.xlsm          — macro-enabled workbook with a hand-crafted
                          xl/vbaProject.bin. M-E DEVIATION: synthetic;
                          office/validate.py treats vbaProject.bin as
                          opaque, so the structural validator is
                          satisfied. See tests/golden/README.md.

What this script does NOT regenerate:
    encrypted.xlsx      — produced by office_passwd.py from clean.xlsx;
                          regenerate via:
                              ./.venv/bin/python office_passwd.py \
                                  tests/golden/inputs/clean.xlsx \
                                  tests/golden/inputs/encrypted.xlsx \
                                  --encrypt password123
                          Encryption uses a fresh random salt per run, so
                          this output is intentionally non-deterministic.

Per the m-E fallback in `docs/tasks/task-001-04-fixtures.md`, all six
fixtures here are agent-synthetic. If/when an Excel-365-authored
`with_legacy.xlsx` or `macro.xlsm` becomes available, replace those two
files verbatim and remove their generation blocks below; the README's
Provenance section documents the swap procedure.
"""
from __future__ import annotations

import sys
import zipfile
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.comments import Comment

HERE = Path(__file__).resolve().parent
GOLDEN_INPUTS = HERE / "golden" / "inputs"

# Pinned epoch for deterministic regeneration. openpyxl's
# DocumentProperties stamps `created` / `modified` with `datetime.now()`
# on Workbook construction; without override, two consecutive regens
# produce 2-line diffs in `docProps/core.xml`. Pinning to a fixed
# datetime makes the synthetic fixtures byte-deterministic across runs
# (Sarcasmotron M-1 fix; encrypted.xlsx is still non-deterministic by
# design — fresh salt per run).
_FIXED_EPOCH = datetime(2026, 1, 1, 0, 0, 0, tzinfo=timezone.utc)


def _pin_properties(wb) -> None:
    """Pin core.xml created/modified to the fixed epoch for determinism."""
    wb.properties.created = _FIXED_EPOCH
    wb.properties.modified = _FIXED_EPOCH


def _save(wb, name: str) -> Path:
    """Save `wb` to GOLDEN_INPUTS/{name} and return the path."""
    GOLDEN_INPUTS.mkdir(parents=True, exist_ok=True)
    _pin_properties(wb)
    out = GOLDEN_INPUTS / name
    wb.save(out)
    return out


def make_clean() -> Path:
    """Single sheet, header + 10 numeric data rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Col1", "Col2", "Col3"])
    for i in range(1, 11):
        ws.append([i, i * 2, i * 3])
    return _save(wb, "clean.xlsx")


def make_multi_sheet() -> Path:
    """Three sheets, last with whitespace in name (apostrophe-syntax fixture)."""
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1["A1"] = "Sheet1 data"
    ws2 = wb.create_sheet("Sheet2")
    ws2["A1"] = "Sheet2 data"
    ws3 = wb.create_sheet("Q1 2026")
    ws3["A1"] = "quarterly"
    return _save(wb, "multi_sheet.xlsx")


def make_hidden_first() -> Path:
    """Sheet1 state=hidden, Sheet2 visible (M2 first-VISIBLE rule fixture)."""
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.sheet_state = "hidden"
    ws1["A1"] = "hidden"
    ws2 = wb.create_sheet("Sheet2")
    ws2["A1"] = "visible"
    return _save(wb, "hidden_first.xlsx")


def make_merged() -> Path:
    """Single sheet with A1:C3 merged (R6 fixture)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "merged-anchor"
    ws.merge_cells("A1:C3")
    ws["A4"] = "below-merge"
    return _save(wb, "merged.xlsx")


def make_with_legacy() -> Path:
    """Single sheet with 2 pre-existing legacy comments via openpyxl.

    M-E DEVIATION (synthetic): Excel-365 authoring unavailable at fixture
    creation. openpyxl emits legacy <comment> + VML drawing roughly
    matching what Excel itself would emit; the OOXML semantics under
    test (sheet/comments/vml binding via rels, idmap data attribute,
    o:spid allocation) are equivalent for the test purpose.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["a", "b", "c"])
    ws.append([1, 2, 3])
    ws["A2"].comment = Comment(text="first existing comment", author="Original")
    ws["B2"].comment = Comment(text="second existing comment", author="Original")
    return _save(wb, "with_legacy.xlsx")


def make_macro_xlsm() -> Path:
    """Macro-enabled .xlsm with a hand-crafted xl/vbaProject.bin.

    M-E DEVIATION (synthetic): Excel-365 authoring unavailable at
    fixture creation. Strategy mirrors the existing test_e2e.sh recipe
    (cross-4 macro warnings block, lines ~246-268):
      1. openpyxl.save → .xlsm path produces a base workbook.
      2. ZIP-rewrite to (a) swap the main Content_Types entry to
         the macroEnabled MIME type, (b) inject xl/vbaProject.bin.
    office/validate.py treats vbaProject.bin as an opaque binary
    stream; structural validation is satisfied by the swap.
    """
    # Use a real tempdir so a Ctrl-C between save and unlink does not
    # leave an orphan `_macro_base.xlsx` in tests/golden/inputs/
    # (Sarcasmotron m-3 fix).
    import tempfile
    out = GOLDEN_INPUTS / "macro.xlsm"
    with tempfile.TemporaryDirectory() as td:
        base = Path(td) / "_macro_base.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["a", "b", "c"])
        ws.append([1, 2, 3])
        _pin_properties(wb)
        wb.save(base)
        with zipfile.ZipFile(base, "r") as src:
            data = {n: src.read(n) for n in src.namelist()}

    ct = data["[Content_Types].xml"].decode("utf-8")
    ct = ct.replace(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml",
        "application/vnd.ms-excel.sheet.macroEnabled.main+xml",
    )
    data["[Content_Types].xml"] = ct.encode("utf-8")
    # Hand-crafted vbaProject.bin: minimal CFB header (8-byte magic) +
    # padding. Excel would refuse to RUN this VBA, but office._macros
    # only inspects the package shape, not the macro contents.
    data["xl/vbaProject.bin"] = (
        b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"  # CFB magic (MS-CFB §2.2)
        + b"\x00" * 504                       # pad to 512 bytes
    )

    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as z:
        for name, payload in data.items():
            z.writestr(name, payload)
    return out


def main() -> int:
    paths = [
        make_clean(),
        make_multi_sheet(),
        make_hidden_first(),
        make_merged(),
        make_with_legacy(),
        make_macro_xlsm(),
    ]
    for p in paths:
        size = p.stat().st_size
        print(f"  {p.name:24s}  {size:6d} bytes")
        assert size <= 50 * 1024, f"{p.name} exceeds 50 KB cap ({size} bytes)"
    print(f"Regenerated {len(paths)} synthetic fixture(s) under {GOLDEN_INPUTS}/")
    print("NOTE: encrypted.xlsx is produced separately via office_passwd.py — see docstring.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
