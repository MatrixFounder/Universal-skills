"""Unit tests for :mod:`xlsx2csv2json.cli` (010-03).

Covers the full argparse surface, format-lock guard, cross-flag
validation, output-arg reconciliation, and the public-helper kwarg
marshaller.
"""
from __future__ import annotations

import argparse
import io
import json
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

_SCRIPTS_DIR = Path(__file__).resolve().parents[2]
if str(_SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS_DIR))


def _capture_stderr() -> tuple[io.StringIO, callable]:
    """Return ``(buf, restore)``. Use restore() in addCleanup."""
    buf = io.StringIO()
    old = sys.stderr
    sys.stderr = buf

    def restore() -> None:
        sys.stderr = old
    return buf, restore


def _make_workbook(td: Path, with_two_sheets: bool = False) -> Path:
    """Tiny .xlsx for tests that need a real input file."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "id"
    ws["B1"] = "name"
    ws["A2"] = 1
    ws["B2"] = "alice"
    if with_two_sheets:
        ws2 = wb.create_sheet("Sheet2")
        ws2["A1"] = "x"
        ws2["A2"] = 99
    out = td / "in.xlsx"
    wb.save(out)
    return out


# ===========================================================================
# build_parser — flag table coverage + format-lock behaviour
# ===========================================================================
class TestBuildParser(unittest.TestCase):

    def test_defaults_locked(self) -> None:
        from xlsx2csv2json.cli import build_parser
        parser = build_parser(format_lock="csv")
        args = parser.parse_args(["fake.xlsx"])
        self.assertEqual(args.sheet, "all")
        self.assertFalse(args.include_hidden)
        # M1 (vdd-multi): parser-level default is None (mode-aware
        # materialisation happens in _validate_flag_combo).
        self.assertIsNone(args.header_rows)
        self.assertEqual(args.header_flatten_style, "string")
        self.assertEqual(args.merge_policy, "anchor-only")
        self.assertEqual(args.tables, "whole")
        self.assertEqual(args.gap_rows, 2)
        self.assertEqual(args.gap_cols, 1)
        self.assertFalse(args.include_hyperlinks)
        self.assertFalse(args.include_formulas)
        self.assertEqual(args.datetime_format, "ISO")
        self.assertFalse(args.json_errors)

    def test_format_lock_csv_rejects_json(self) -> None:
        from xlsx2csv2json.cli import build_parser
        from xlsx2csv2json import FormatLockedByShim
        parser = build_parser(format_lock="csv")
        with self.assertRaises(FormatLockedByShim):
            parser.parse_args(["fake.xlsx", "--format", "json"])

    def test_format_lock_json_rejects_csv(self) -> None:
        from xlsx2csv2json.cli import build_parser
        from xlsx2csv2json import FormatLockedByShim
        parser = build_parser(format_lock="json")
        with self.assertRaises(FormatLockedByShim):
            parser.parse_args(["fake.xlsx", "--format", "csv"])

    def test_format_lock_csv_accepts_csv(self) -> None:
        from xlsx2csv2json.cli import build_parser
        parser = build_parser(format_lock="csv")
        args = parser.parse_args(["fake.xlsx", "--format", "csv"])
        self.assertEqual(args.format, "csv")

    def test_no_format_lock_requires_format(self) -> None:
        from xlsx2csv2json.cli import build_parser
        parser = build_parser(format_lock=None)
        args = parser.parse_args(["fake.xlsx", "--format", "csv"])
        self.assertEqual(args.format, "csv")

    def test_header_rows_type_auto(self) -> None:
        from xlsx2csv2json.cli import build_parser
        parser = build_parser(format_lock="csv")
        args = parser.parse_args(["fake.xlsx", "--header-rows", "auto"])
        self.assertEqual(args.header_rows, "auto")

    def test_header_rows_type_int(self) -> None:
        from xlsx2csv2json.cli import build_parser
        parser = build_parser(format_lock="csv")
        args = parser.parse_args(["fake.xlsx", "--header-rows", "2"])
        self.assertEqual(args.header_rows, 2)

    def test_header_rows_rejects_negative(self) -> None:
        from xlsx2csv2json.cli import build_parser
        parser = build_parser(format_lock="csv")
        # argparse converts ArgumentTypeError into SystemExit.
        # Suppress the usage banner argparse emits on stderr.
        old = sys.stderr
        sys.stderr = io.StringIO()
        try:
            with self.assertRaises(SystemExit):
                parser.parse_args(["fake.xlsx", "--header-rows", "-1"])
        finally:
            sys.stderr = old

    def test_header_rows_rejects_garbage(self) -> None:
        from xlsx2csv2json.cli import build_parser
        parser = build_parser(format_lock="csv")
        old = sys.stderr
        sys.stderr = io.StringIO()
        try:
            with self.assertRaises(SystemExit):
                parser.parse_args(["fake.xlsx", "--header-rows", "not-a-num"])
        finally:
            sys.stderr = old


# ===========================================================================
# _validate_flag_combo — cross-flag invariants
# ===========================================================================
class TestCliDelimiterAliasResolution(unittest.TestCase):
    """**vdd-adversarial R26 HIGH-1 fix:** the CLI-layer alias resolution
    (`_delimiter_type` argparse type-callable) was untested. The 4
    `test_delimiter_*` tests in `test_emit_csv.py` call `emit_csv` with
    raw `\\t`/`|`/`;`/`,`, bypassing `cli.main` entirely. These tests
    drive the actual CLI via `main([...])` and assert per-alias output.
    """

    def _run_and_read(self, td: Path, delimiter_arg: str) -> str:
        import tempfile
        from xlsx2csv2json.cli import main
        wb = _make_workbook(td)
        out = td / "out.csv"
        rc = main(
            [str(wb), str(out), "--sheet", "Sheet1",
             "--delimiter", delimiter_arg],
            format_lock="csv",
        )
        self.assertEqual(rc, 0)
        return out.read_text(encoding="utf-8")

    def test_cli_e2e_delimiter_tab_alias_writes_tsv(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            text = self._run_and_read(Path(td), "tab")
            self.assertIn("id\tname", text)
            self.assertIn("1\talice", text)
            self.assertNotIn(",", text)

    def test_cli_e2e_delimiter_pipe_alias_writes_pipe_separated(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            text = self._run_and_read(Path(td), "pipe")
            self.assertIn("id|name", text)
            self.assertIn("1|alice", text)

    def test_cli_e2e_delimiter_backslash_t_escape_writes_tsv(self) -> None:
        """The 2-char escape `\\t` (literal backslash + t) is for shells
        that can't easily emit a raw tab. Maps to a real tab character.
        """
        with tempfile.TemporaryDirectory() as td:
            text = self._run_and_read(Path(td), "\\t")
            self.assertIn("id\tname", text)

    def test_cli_e2e_delimiter_semicolon_literal_excel_ru_recipe(self) -> None:
        """The documented Excel-RU recipe: --delimiter ';'."""
        with tempfile.TemporaryDirectory() as td:
            text = self._run_and_read(Path(td), ";")
            self.assertIn("id;name", text)
            self.assertIn("1;alice", text)


class TestCliDelimiterRejection(unittest.TestCase):
    """**vdd-adversarial R26 MED-2 fix:** AC-26.5 ("argparse rejects
    values outside {,, ;, tab, pipe}") was unverified.
    """

    def test_cli_rejects_unknown_delimiter(self) -> None:
        """argparse usage error → `main()` returns exit code 2
        (it converts the underlying `SystemExit` to a rc-return per
        cross-cutting envelope plumbing).
        """
        from xlsx2csv2json.cli import main
        with tempfile.TemporaryDirectory() as td:
            wb = _make_workbook(Path(td))
            _, restore = _capture_stderr()
            try:
                rc = main(
                    [str(wb), "--delimiter", "bogus"], format_lock="csv"
                )
            finally:
                restore()
            self.assertEqual(rc, 2)

    def test_cli_rejects_multi_char_delimiter(self) -> None:
        from xlsx2csv2json.cli import main
        with tempfile.TemporaryDirectory() as td:
            wb = _make_workbook(Path(td))
            _, restore = _capture_stderr()
            try:
                rc = main(
                    [str(wb), "--delimiter", ";;"], format_lock="csv"
                )
            finally:
                restore()
            self.assertEqual(rc, 2)

    def test_delimiter_type_helper_raises_argument_type_error(self) -> None:
        """Direct unit test on the `_delimiter_type` callable — exercises
        the error message contents (which argparse surfaces in usage).
        """
        import argparse
        from xlsx2csv2json.cli import _delimiter_type
        self.assertEqual(_delimiter_type(","), ",")
        self.assertEqual(_delimiter_type(";"), ";")
        self.assertEqual(_delimiter_type("tab"), "\t")
        self.assertEqual(_delimiter_type("\\t"), "\t")
        self.assertEqual(_delimiter_type("pipe"), "|")
        with self.assertRaises(argparse.ArgumentTypeError) as cm:
            _delimiter_type("bogus")
        self.assertIn("tab", str(cm.exception))
        self.assertIn("pipe", str(cm.exception))


class TestStdoutEncodingWarning(unittest.TestCase):
    """**/vdd-multi-3 Logic-LOW-1 fix:** `--encoding utf-8-sig` is
    silently ignored on stdout CSV output (BOM-in-pipe corrupts
    downstream consumers, so emit_csv drops it). Surface a stderr
    warning so the user isn't silently surprised.
    """

    def test_csv_stdout_with_utf8_sig_emits_warning(self) -> None:
        from xlsx2csv2json.cli import main
        with tempfile.TemporaryDirectory() as td:
            wb = _make_workbook(Path(td))
            buf, restore = _capture_stderr()
            try:
                rc = main(
                    [str(wb), "-", "--sheet", "Sheet1",
                     "--encoding", "utf-8-sig"],
                    format_lock="csv",
                )
            finally:
                restore()
            self.assertEqual(rc, 0)
            self.assertIn("no effect on stdout CSV output",
                          buf.getvalue())

    def test_csv_file_output_with_utf8_sig_silent(self) -> None:
        """File output (not stdout) does NOT trigger the warning —
        the BOM is correctly emitted to the file.
        """
        from xlsx2csv2json.cli import main
        with tempfile.TemporaryDirectory() as td:
            wb = _make_workbook(Path(td))
            out = Path(td) / "out.csv"
            buf, restore = _capture_stderr()
            try:
                main(
                    [str(wb), str(out), "--sheet", "Sheet1",
                     "--encoding", "utf-8-sig"],
                    format_lock="csv",
                )
            finally:
                restore()
            self.assertNotIn("no effect on stdout", buf.getvalue())
            self.assertTrue(out.read_bytes().startswith(b"\xef\xbb\xbf"))


class TestJsonDelimiterWarning(unittest.TestCase):
    """**vdd-adversarial R26 HIGH-2 fix:** `xlsx2json.py --delimiter ';'`
    must emit a stderr warning (mirror of the `--encoding utf-8-sig`
    warning we added for the same JSON-vs-CSV inconsistency in R25).
    """

    def test_xlsx2json_with_semicolon_delimiter_warns(self) -> None:
        from xlsx2csv2json.cli import main
        with tempfile.TemporaryDirectory() as td:
            wb = _make_workbook(Path(td))
            out = Path(td) / "out.json"
            buf, restore = _capture_stderr()
            try:
                rc = main(
                    [str(wb), str(out), "--sheet", "Sheet1",
                     "--delimiter", ";"],
                    format_lock="json",
                )
            finally:
                restore()
            self.assertEqual(rc, 0)
            self.assertIn("--delimiter has no effect on JSON output",
                          buf.getvalue())

    def test_xlsx2json_default_delimiter_is_silent(self) -> None:
        from xlsx2csv2json.cli import main
        with tempfile.TemporaryDirectory() as td:
            wb = _make_workbook(Path(td))
            out = Path(td) / "out.json"
            buf, restore = _capture_stderr()
            try:
                main([str(wb), str(out)], format_lock="json")
            finally:
                restore()
            self.assertNotIn("--delimiter", buf.getvalue())


class TestPublicHelperAcceptsDelimiterKwarg(unittest.TestCase):
    """**vdd-adversarial R26 HIGH-3 fix:** the Python public helper
    `convert_xlsx_to_csv(..., delimiter=...)` previously raised
    `TypeError: Unknown kwarg: 'delimiter'` because `_KWARG_TO_FLAG`
    in `__init__.py` was not updated alongside the CLI flag.
    """

    def test_convert_xlsx_to_csv_accepts_delimiter_kwarg(self) -> None:
        from xlsx2csv2json import convert_xlsx_to_csv
        with tempfile.TemporaryDirectory() as td:
            wb = _make_workbook(Path(td))
            out = Path(td) / "out.csv"
            rc = convert_xlsx_to_csv(
                str(wb), str(out), sheet="Sheet1", delimiter=";"
            )
            self.assertEqual(rc, 0)
            self.assertIn("id;name", out.read_text(encoding="utf-8"))

    def test_convert_xlsx_to_csv_accepts_encoding_kwarg(self) -> None:
        """Side-fix: `--encoding` (R25) had the same hole — it was also
        missing from `_KWARG_TO_FLAG`. This test locks the fix that
        landed alongside the delimiter one.
        """
        from xlsx2csv2json import convert_xlsx_to_csv
        with tempfile.TemporaryDirectory() as td:
            wb = _make_workbook(Path(td))
            out = Path(td) / "out.csv"
            rc = convert_xlsx_to_csv(
                str(wb), str(out), sheet="Sheet1", encoding="utf-8-sig"
            )
            self.assertEqual(rc, 0)
            self.assertTrue(out.read_bytes().startswith(b"\xef\xbb\xbf"))

    def test_convert_xlsx_to_csv_rejects_unknown_kwarg(self) -> None:
        """Defensive: typo'd kwarg still raises (no silent regression
        of the strict-mode kwarg validation)."""
        from xlsx2csv2json import convert_xlsx_to_csv
        with tempfile.TemporaryDirectory() as td:
            wb = _make_workbook(Path(td))
            out = Path(td) / "out.csv"
            with self.assertRaisesRegex(TypeError, "Unknown kwarg"):
                convert_xlsx_to_csv(
                    str(wb), str(out), sheet="Sheet1", delim=";"  # typo
                )


class TestValidateFlagCombo(unittest.TestCase):

    def _parse(self, format_lock: str | None, argv: list[str]) -> argparse.Namespace:
        from xlsx2csv2json.cli import build_parser
        return build_parser(format_lock=format_lock).parse_args(argv)

    def test_header_rows_int_with_multi_table_raises(self) -> None:
        from xlsx2csv2json.cli import _validate_flag_combo
        from xlsx2csv2json import HeaderRowsConflict
        args = self._parse(
            "csv",
            ["fake.xlsx", "--header-rows", "2", "--tables", "listobjects"],
        )
        with self.assertRaises(HeaderRowsConflict):
            _validate_flag_combo(args, format_lock="csv")

    def test_header_rows_auto_with_multi_table_ok(self) -> None:
        from xlsx2csv2json.cli import _validate_flag_combo
        args = self._parse(
            "csv",
            ["fake.xlsx", "--header-rows", "auto", "--tables", "listobjects"],
        )
        _validate_flag_combo(args, format_lock="csv")  # no raise

    def test_csv_sheet_all_dash_output_raises(self) -> None:
        from xlsx2csv2json.cli import _validate_flag_combo
        from xlsx2csv2json import MultiSheetRequiresOutputDir
        args = self._parse("csv", ["fake.xlsx", "-"])
        with self.assertRaises(MultiSheetRequiresOutputDir):
            _validate_flag_combo(args, format_lock="csv")

    def test_csv_sheet_all_no_output_no_raise_at_parse_time(self) -> None:
        """Conservative pre-check: with NO output, defer to dispatch."""
        from xlsx2csv2json.cli import _validate_flag_combo
        args = self._parse("csv", ["fake.xlsx"])
        _validate_flag_combo(args, format_lock="csv")  # no raise


# ===========================================================================
# _resolve_output_arg — mutual exclusivity
# ===========================================================================
class TestResolveOutputArg(unittest.TestCase):

    def test_positional_only(self) -> None:
        from xlsx2csv2json.cli import _resolve_output_arg
        ns = argparse.Namespace(output="out.json", output_flag=None)
        self.assertEqual(_resolve_output_arg(ns), "out.json")

    def test_flag_only(self) -> None:
        from xlsx2csv2json.cli import _resolve_output_arg
        ns = argparse.Namespace(output=None, output_flag="out.json")
        self.assertEqual(_resolve_output_arg(ns), "out.json")

    def test_both_raises(self) -> None:
        from xlsx2csv2json.cli import _resolve_output_arg
        from xlsx2csv2json import _AppError
        ns = argparse.Namespace(output="a.json", output_flag="b.json")
        with self.assertRaises(_AppError):
            _resolve_output_arg(ns)

    def test_neither_returns_none(self) -> None:
        from xlsx2csv2json.cli import _resolve_output_arg
        ns = argparse.Namespace(output=None, output_flag=None)
        self.assertIsNone(_resolve_output_arg(ns))


# ===========================================================================
# _build_argv — kwarg → flag marshaller
# ===========================================================================
class TestBuildArgv(unittest.TestCase):

    def test_basic(self) -> None:
        from xlsx2csv2json import _build_argv
        argv = _build_argv("in.xlsx", "out.json", {})
        self.assertEqual(argv, ["in.xlsx", "--output", "out.json"])

    def test_no_output(self) -> None:
        from xlsx2csv2json import _build_argv
        argv = _build_argv("in.xlsx", None, {})
        self.assertEqual(argv, ["in.xlsx"])

    def test_value_kwarg(self) -> None:
        from xlsx2csv2json import _build_argv
        argv = _build_argv("in.xlsx", None, {"sheet": "Sheet1", "tables": "gap"})
        self.assertIn("--sheet", argv)
        self.assertIn("Sheet1", argv)
        self.assertIn("--tables", argv)
        self.assertIn("gap", argv)

    def test_bool_kwarg_true(self) -> None:
        from xlsx2csv2json import _build_argv
        argv = _build_argv("in.xlsx", None, {"include_hyperlinks": True})
        self.assertIn("--include-hyperlinks", argv)

    def test_bool_kwarg_false_dropped(self) -> None:
        from xlsx2csv2json import _build_argv
        argv = _build_argv("in.xlsx", None, {"include_hyperlinks": False})
        self.assertNotIn("--include-hyperlinks", argv)

    def test_unknown_kwarg_raises(self) -> None:
        from xlsx2csv2json import _build_argv
        with self.assertRaises(TypeError):
            _build_argv("in.xlsx", None, {"foo": "bar"})


# ===========================================================================
# main — end-to-end (parse + dispatch trampoline up to emit stub)
# ===========================================================================
class TestMain(unittest.TestCase):

    def test_help_csv_shim_exits_zero(self) -> None:
        """`python3 xlsx2csv.py --help` should exit 0 (subprocess)."""
        result = subprocess.run(
            [sys.executable, str(_SCRIPTS_DIR / "xlsx2csv.py"), "--help"],
            capture_output=True,
            text=True,
        )
        self.assertEqual(result.returncode, 0)
        self.assertIn("INPUT", result.stdout)

    def test_help_json_shim_exits_zero(self) -> None:
        result = subprocess.run(
            [sys.executable, str(_SCRIPTS_DIR / "xlsx2json.py"), "--help"],
            capture_output=True,
            text=True,
        )
        self.assertEqual(result.returncode, 0)
        self.assertIn("INPUT", result.stdout)

    def test_format_lock_violation_exits_2_envelope(self) -> None:
        result = subprocess.run(
            [
                sys.executable,
                str(_SCRIPTS_DIR / "xlsx2csv.py"),
                "/nonexistent.xlsx",
                "--format", "json",
                "--json-errors",
            ],
            capture_output=True,
            text=True,
        )
        self.assertEqual(result.returncode, 2)
        envelope = json.loads(result.stderr.strip().split("\n")[0])
        self.assertEqual(envelope["type"], "FormatLockedByShim")
        self.assertEqual(envelope["code"], 2)

    def test_header_rows_conflict_exits_2_envelope(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            inp = _make_workbook(Path(td))
            result = subprocess.run(
                [
                    sys.executable,
                    str(_SCRIPTS_DIR / "xlsx2json.py"),
                    str(inp),
                    "--header-rows", "2",
                    "--tables", "listobjects",
                    "--json-errors",
                ],
                capture_output=True,
                text=True,
            )
        self.assertEqual(result.returncode, 2)
        envelope = json.loads(result.stderr.strip().split("\n")[0])
        self.assertEqual(envelope["type"], "HeaderRowsConflict")

    def test_main_reaches_emitter_and_writes_json(self) -> None:
        """010-05 wires emit_json — main() now produces real JSON output.

        Prior bump-history: 010-03 sentinel -997; replaced now that
        emit_json is live (returns 0 on success).
        """
        from xlsx2csv2json import main
        with tempfile.TemporaryDirectory() as td:
            inp = _make_workbook(Path(td))
            out = Path(td) / "out.json"
            buf, restore = _capture_stderr()
            self.addCleanup(restore)
            rc = main([str(inp), str(out)], format_lock="json")
            self.assertEqual(rc, 0)
            self.assertTrue(out.exists())
            # Single-sheet single-region: rule 1 → flat array.
            content = out.read_text("utf-8")
            self.assertTrue(content.startswith("["), content[:50])

    def test_main_missing_input_returns_1(self) -> None:
        """FileNotFoundError mapped to exit 1 via envelope."""
        from xlsx2csv2json import main
        buf, restore = _capture_stderr()
        self.addCleanup(restore)
        rc = main(["/nonexistent.xlsx"], format_lock="json")
        self.assertEqual(rc, 1)


if __name__ == "__main__":
    unittest.main()
