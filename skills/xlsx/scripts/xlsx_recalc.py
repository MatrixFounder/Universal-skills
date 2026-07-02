#!/usr/bin/env python3
"""Force formula recalculation in an .xlsx via headless LibreOffice.

Why this is needed: `openpyxl` saves formulas as *strings* with no
cached value. Consumers that read the file with `data_only=True` see
`None` until some spreadsheet engine opens and resaves the workbook.

How it works: the workbook is round-tripped through
`soffice --convert-to` while the throwaway LibreOffice profile is
seeded with `OOXMLRecalcMode=0` ("recalculate always on load"), so
missing AND stale cached values are both recomputed. The result is
then VERIFIED: if the file contains formula cells and none of them
came back with a cached value, the script exits non-zero instead of
pretending success.

History: the previous implementation drove a one-shot StarBasic macro
(`macro:///Standard.Module1.RecalcAndSave`) installed into a second
user profile passed via `-env:UserInstallation=`. LibreOffice 26.2
honours only the FIRST `-env:UserInstallation=` argument and drops
CLI macro dispatch on cold profiles, so the macro never ran — soffice
exited 0, the script printed "Recalculated.", and every formula cell
silently stayed `None`. The `--convert-to` path has no macro-security
or profile-ordering surface and behaves deterministically.

After recalculation the script (optionally) scans every cell for
formula errors (`#REF!`, `#DIV/0!`, `#VALUE!`, `#NAME?`, `#N/A`,
`#NUM!`, `#NULL!`) and reports them.

Usage:
    python3 xlsx_recalc.py input.xlsx [--output out.xlsx]
                                       [--timeout 120]
                                       [--scan-errors]
                                       [--json]

If `--output` is omitted, the input file is rewritten in place.
"""

from __future__ import annotations

import argparse
import json
import shutil
import sys
import tempfile
import zipfile
from pathlib import Path

from defusedxml.ElementTree import iterparse as defused_iterparse  # type: ignore
from openpyxl import load_workbook  # type: ignore

from _errors import add_json_errors_argument, report_error
from _soffice import SofficeError, convert_to as soffice_convert_to
from office._encryption import EncryptedFileError, assert_not_encrypted
from office._macros import warn_if_macros_will_be_dropped


# Written into the throwaway profile that `_soffice.run` owns.
# OOXMLRecalcMode=0 → "Recalculate always" when loading Excel formats;
# LibreOffice's default keeps whatever cached values are in the file,
# so a stale (wrong) cache would survive a plain --convert-to pass.
RECALC_PROFILE_SEED = {
    "user/registrymodifications.xcu": (
        '<?xml version="1.0" encoding="UTF-8"?>\n'
        '<oor:items xmlns:oor="http://openoffice.org/2001/registry"'
        ' xmlns:xs="http://www.w3.org/2001/XMLSchema">\n'
        ' <item oor:path="/org.openoffice.Office.Calc/Formula/Load">'
        '<prop oor:name="OOXMLRecalcMode" oor:op="fuse">'
        "<value>0</value></prop></item>\n"
        "</oor:items>\n"
    ),
}

ERROR_TOKENS = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NUM!", "#NULL!"}


class RecalcVerificationError(RuntimeError):
    """soffice exited 0 but the output workbook was not recalculated."""


def _local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def verify_cached_values(path: Path) -> None:
    """Raise RecalcVerificationError when recalculation clearly did not run.

    A recalculated workbook stores a cached result — a `<v>` element —
    for every formula cell. If the file has formula cells and NOT ONE
    of them carries a `<v>`, the recalculation pass was a silent no-op
    (the LibreOffice 26.2 failure mode this script used to miss).

    The check is XML-level rather than openpyxl's `data_only` values,
    and has to thread a needle:

    - openpyxl writes formula cells as `<c><f>…</f><v></v></c>` — an
      EMPTY `<v>` with no `t` attribute. That is the uncached state
      (a numeric cache can never be empty).
    - LibreOffice caches a formula evaluating to an empty string as
      `<c t="str"><f>…</f><v></v></c>` — an empty `<v>` WITH
      `t="str"`. openpyxl's `data_only` reports that as None, so a
      value-based check would false-fail an all-blank-producing
      template even though LibreOffice did recalculate it.

    A formula cell therefore counts as cached iff its `<v>` has
    non-empty text, or is empty while the cell is string-typed.
    """
    total_formula_cells = 0
    with zipfile.ZipFile(str(path)) as z:
        sheet_parts = [
            name for name in z.namelist()
            if name.startswith("xl/worksheets/")
            and name.endswith(".xml")
            and "/_rels/" not in name
        ]
        for part in sheet_parts:
            with z.open(part) as fh:
                for _event, elem in defused_iterparse(fh):
                    if _local_name(elem.tag) != "c":
                        continue
                    children = {_local_name(child.tag): child for child in elem}
                    if "f" not in children:
                        elem.clear()
                        continue
                    v_elem = children.get("v")
                    if v_elem is not None and (
                        (v_elem.text or "") != ""
                        or elem.get("t") in ("str", "inlineStr")
                    ):
                        return  # at least one cached formula → recalc ran
                    total_formula_cells += 1
                    elem.clear()
    if total_formula_cells == 0:
        return
    raise RecalcVerificationError(
        f"LibreOffice exited 0 but none of the {total_formula_cells} "
        f"formula cell(s) in {path.name} carries a cached <v> value — "
        "recalculation did not happen. Check `soffice --version` (this "
        "script is validated against LibreOffice 26.2) and retry with a "
        "larger --timeout; the input file was left untouched."
    )


def recalc(input_path: Path, output_path: Path, *, timeout: int) -> None:
    input_path = input_path.resolve()
    output_path = output_path.resolve()
    if not input_path.is_file():
        raise FileNotFoundError(input_path)
    if output_path.is_dir():
        # shutil.move would silently drop the file INTO the directory
        # as <dir>/<input stem>.xlsx instead of at the requested path.
        raise IsADirectoryError(
            f"--output points at an existing directory: {output_path}"
        )
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # .xlsm output keeps the VBA project through LibreOffice's
    # macro-enabled filter; anything else goes through the plain
    # xlsx filter (macro drop is warned about upstream in main()).
    target_format = "xlsm" if output_path.suffix.lower() == ".xlsm" else "xlsx"

    with tempfile.TemporaryDirectory(prefix="lo-recalc-") as tmp:
        produced = soffice_convert_to(
            input_path,
            tmp,
            target_format,
            timeout=timeout,
            profile_seed=RECALC_PROFILE_SEED,
        )
        # Verify BEFORE touching output_path: on failure the
        # destination (== the input in in-place mode) stays intact.
        verify_cached_values(produced)
        shutil.move(str(produced), str(output_path))


def scan_errors(path: Path) -> dict[str, list[str]]:
    """Report cells whose data_type is 'e' (Excel error) only.

    A cell can hold the literal string "#REF!" without being a formula
    error; we use the Excel-level data type to distinguish the two.
    """
    wb = load_workbook(str(path), data_only=True, read_only=True)
    hits: dict[str, list[str]] = {}
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.data_type != "e":
                        continue
                    value = cell.value
                    if isinstance(value, str) and value in ERROR_TOKENS:
                        hits.setdefault(value, []).append(f"{sheet_name}!{cell.coordinate}")
    finally:
        wb.close()
    return hits


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("input", type=Path, help="Source .xlsx")
    parser.add_argument("--output", type=Path, default=None, help="Destination .xlsx (default: rewrite in place)")
    parser.add_argument("--timeout", type=int, default=120, help="soffice timeout (default 120s)")
    parser.add_argument("--scan-errors", action="store_true", help="Scan for formula-error cells after recalc")
    parser.add_argument("--json", action="store_true", help="Emit JSON summary")
    add_json_errors_argument(parser)
    args = parser.parse_args(argv)
    je = args.json_errors

    output = args.output or args.input
    warn_if_macros_will_be_dropped(args.input, output, sys.stderr)

    try:
        assert_not_encrypted(args.input)
        recalc(args.input, output, timeout=args.timeout)
    except EncryptedFileError as exc:
        return report_error(
            str(exc), code=3, error_type="EncryptedFileError",
            details={"path": str(args.input)}, json_mode=je,
        )
    except IsADirectoryError as exc:
        return report_error(
            str(exc), code=1, error_type="OutputIsADirectory",
            details={"path": str(output)}, json_mode=je,
        )
    except FileNotFoundError as exc:
        return report_error(
            f"Input not found: {exc}",
            code=1, error_type="FileNotFound", json_mode=je,
        )
    except SofficeError as exc:
        return report_error(
            str(exc), code=1, error_type="SofficeError", json_mode=je,
        )
    except RecalcVerificationError as exc:
        return report_error(
            str(exc), code=1, error_type="RecalcVerificationError",
            details={"path": str(args.input)}, json_mode=je,
        )

    if not args.scan_errors:
        if args.json:
            print(json.dumps({"ok": True, "errors": {}}, ensure_ascii=False))
        else:
            print("Recalculated.")
        return 0

    errors = scan_errors(output)
    if args.json:
        print(json.dumps({"ok": not errors, "errors": errors}, ensure_ascii=False, indent=2))
    else:
        if not errors:
            print("Recalculated, no formula errors.")
        else:
            for code, cells in errors.items():
                print(f"{code}: {len(cells)} cells — {', '.join(cells[:10])}"
                      + ("..." if len(cells) > 10 else ""))
    return 1 if errors else 0


if __name__ == "__main__":
    sys.exit(main())
