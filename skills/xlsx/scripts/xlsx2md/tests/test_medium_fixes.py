"""Regression tests for the 8 Sarcasmotron MEDIUM findings on xlsx-9.

M1: ``--include-formulas`` now wires ``keep_formulas=True`` into
    ``open_workbook`` and the HTML emitter surfaces formula strings as
    ``data-formula`` attributes (was a no-op).
M2: ``PostValidateFailed`` is wired to the env-flag re-parse gate
    (``XLSX_XLSX2MD_POST_VALIDATE=1``) instead of being dead code.
M3: Library auto-streaming (size threshold ≥ 100 MiB) now produces the
    hyperlink-unreliability warning via the post-open ``reader._read_only``
    detection (was bypassed when ``--memory-mode=auto`` triggered streaming).
M5: Output is written to a sibling tempfile and atomically renamed on
    success; the tempfile is unlinked on any exception (was leaving a
    partial ``.md`` on disk when ``emit_workbook_md`` raised mid-write).
M6: ``cell_addr`` for warning messages is computed lazily — only for
    cells that have a hyperlink (was unconditional per-cell ~3M wasted
    ops at R20a scale).
M7: ``--hyperlink-scheme-allowlist='*'`` help text now names
    ``javascript:``/``data:``/``vbscript:``/``file:`` explicitly (was a
    vague "NOT recommended").

(M4 lives in ``test_public_api.py``; M8 lives in ``test_inline.py``.)
"""
from __future__ import annotations

import io
import os
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

_SCRIPTS_ROOT = Path(__file__).resolve().parent.parent.parent
if str(_SCRIPTS_ROOT) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS_ROOT))

import xlsx2md  # noqa: E402
from xlsx2md.cli import (  # noqa: E402
    _post_validate_output,
    _resolve_output_stream,
    main,
)
from xlsx2md.exceptions import PostValidateFailed  # noqa: E402

_FIXTURES = Path(__file__).resolve().parent / "fixtures"
_SHIM = _SCRIPTS_ROOT / "xlsx2md.py"


# ---------------------------------------------------------------------------
# M1 — --include-formulas wiring
# ---------------------------------------------------------------------------

class TestIncludeFormulasWiring(unittest.TestCase):
    """M1: ``--include-formulas`` must open the workbook with
    ``keep_formulas=True`` so formula strings reach the emit layer.
    """

    def test_open_workbook_called_with_keep_formulas_true_when_flag_set(
        self,
    ) -> None:
        """cli.main() must propagate args.include_formulas → keep_formulas."""
        captured: dict = {}

        def _spy_open(path, *, read_only_mode=None, keep_formulas=False):
            captured["keep_formulas"] = keep_formulas
            # Raise to short-circuit downstream — we only care about the
            # kwarg propagation, not the actual workbook contents.
            raise RuntimeError("spy short-circuit")

        with patch("xlsx_read.open_workbook", side_effect=_spy_open):
            main([
                str(_FIXTURES / "single_cell.xlsx"),
                "--format=html",  # M7 lock allows formulas in HTML/hybrid
                "--include-formulas",
            ])
        self.assertTrue(
            captured.get("keep_formulas"),
            "keep_formulas=True must be passed when --include-formulas is set",
        )

    def test_open_workbook_called_with_keep_formulas_false_when_flag_absent(
        self,
    ) -> None:
        """Default (no flag): keep_formulas=False (data_only=True)."""
        captured: dict = {}

        def _spy_open(path, *, read_only_mode=None, keep_formulas=False):
            captured["keep_formulas"] = keep_formulas
            raise RuntimeError("spy short-circuit")

        with patch("xlsx_read.open_workbook", side_effect=_spy_open):
            main([str(_FIXTURES / "single_cell.xlsx"), "--format=html"])
        self.assertFalse(
            captured.get("keep_formulas", True),
            "keep_formulas=False is the default when --include-formulas is absent",
        )

    def test_emit_html_formula_string_emits_data_formula_attr(self) -> None:
        """Direct emit-side test: cell value `"=A1+B1"` + include_formulas
        → ``<td data-formula="=A1+B1" ...></td>`` with empty content.
        """
        from xlsx_read import TableData, TableRegion  # noqa: PLC0415
        from xlsx2md.emit_html import emit_html_table  # noqa: PLC0415

        region = TableRegion(
            sheet="Sheet1", top_row=1, left_col=1,
            bottom_row=2, right_col=1, source="gap_detect",
        )
        td = TableData(
            region=region,
            headers=["Result"],
            rows=[["=A1+B1"]],
        )
        out = io.StringIO()
        emit_html_table(
            td, out,
            include_formulas=True,
            hyperlink_allowlist=None,
        )
        output = out.getvalue()
        self.assertIn('data-formula="=A1+B1"', output)
        # Cell visible content is empty (cached value unavailable when
        # keep_formulas=True; honest-scope §1.4(g)).
        self.assertIn("<td", output)


# ---------------------------------------------------------------------------
# M2 — PostValidateFailed env-flag re-parse gate
# ---------------------------------------------------------------------------

class TestPostValidateGate(unittest.TestCase):
    """M2: ``XLSX_XLSX2MD_POST_VALIDATE=1`` env flag activates a re-parse
    gate on the produced Markdown; on failure raises PostValidateFailed
    (CODE=7) which the cli envelope routes to exit 7.
    """

    def test_post_validate_flag_disabled_by_default_no_op(self) -> None:
        """No env var set → no-op (returns silently)."""
        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".md", delete=False, encoding="utf-8",
        ) as tmp:
            tmp.write("")  # empty file
            out_path = Path(tmp.name)
        try:
            # Env not set → no-op, no exception.
            _post_validate_output(out_path)
        finally:
            out_path.unlink(missing_ok=True)

    def test_post_validate_flag_enabled_empty_file_raises(self) -> None:
        """Env flag on + empty file → PostValidateFailed."""
        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".md", delete=False, encoding="utf-8",
        ) as tmp:
            tmp.write("")
            out_path = Path(tmp.name)
        try:
            with patch.dict(os.environ, {"XLSX_XLSX2MD_POST_VALIDATE": "1"}):
                with self.assertRaises(PostValidateFailed):
                    _post_validate_output(out_path)
        finally:
            out_path.unlink(missing_ok=True)

    def test_post_validate_flag_enabled_no_markdown_structure_raises(
        self,
    ) -> None:
        """Env flag on + text without ## / |---| / <table → raises."""
        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".md", delete=False, encoding="utf-8",
        ) as tmp:
            tmp.write("just plain text without any markdown markers\n")
            out_path = Path(tmp.name)
        try:
            with patch.dict(os.environ, {"XLSX_XLSX2MD_POST_VALIDATE": "1"}):
                with self.assertRaises(PostValidateFailed):
                    _post_validate_output(out_path)
        finally:
            out_path.unlink(missing_ok=True)

    def test_post_validate_flag_enabled_valid_markdown_passes(self) -> None:
        """Env flag on + valid markdown with ## H2 → no-op."""
        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".md", delete=False, encoding="utf-8",
        ) as tmp:
            tmp.write("## Sheet1\n\n| a |\n|---|\n| b |\n")
            out_path = Path(tmp.name)
        try:
            with patch.dict(os.environ, {"XLSX_XLSX2MD_POST_VALIDATE": "1"}):
                _post_validate_output(out_path)
        finally:
            out_path.unlink(missing_ok=True)

    def test_post_validate_via_cli_end_to_end_returns_code_7_on_failure(
        self,
    ) -> None:
        """End-to-end: monkey-patch emit_workbook_md to produce empty
        output → env-flag gate fires → exit 7 PostValidateFailed.
        """
        # Patch emit_workbook_md to produce empty output but return 0.
        def _empty_emit(reader, args, out):
            return 0  # success exit but no markdown written

        with tempfile.TemporaryDirectory() as td:
            out_path = Path(td) / "out.md"
            with patch(
                "xlsx2md.emit_hybrid.emit_workbook_md",
                side_effect=_empty_emit,
            ), patch.dict(
                os.environ, {"XLSX_XLSX2MD_POST_VALIDATE": "1"},
            ):
                result = main([
                    str(_FIXTURES / "single_cell.xlsx"),
                    str(out_path),
                ])
        self.assertEqual(result, 7, "PostValidateFailed → exit 7")


# ---------------------------------------------------------------------------
# M3 — Auto-streaming hyperlink-unreliability warning
# ---------------------------------------------------------------------------

class TestAutoStreamingWarning(unittest.TestCase):
    """M3: Warning fires for EFFECTIVE streaming mode, not just explicit.
    """

    def test_args_read_only_effective_set_by_cli_main_from_reader(
        self,
    ) -> None:
        """cli.main() must read reader._read_only post-open and set
        args._read_only_effective so dispatch's gate fires regardless
        of how streaming was decided.
        """
        from xlsx_read import WorkbookReader  # noqa: PLC0415

        captured_args: dict = {}

        def _spy_emit(reader, args, out):
            captured_args["_read_only_effective"] = getattr(
                args, "_read_only_effective", "MISSING"
            )
            return 0

        # Build a fake reader with _read_only=True (simulates the case
        # where library auto-chose streaming on a >100 MiB workbook).
        class _FakeReader:
            _read_only = True

            def __enter__(self):
                return self

            def __exit__(self, *_exc):
                return None

            def sheets(self):
                return []

        with patch(
            "xlsx_read.open_workbook", return_value=_FakeReader(),
        ), patch(
            "xlsx2md.emit_hybrid.emit_workbook_md", side_effect=_spy_emit,
        ):
            # Default --memory-mode=auto (so _read_only_mode_resolved=None).
            with tempfile.TemporaryDirectory() as td:
                out_path = Path(td) / "out.md"
                main([
                    str(_FIXTURES / "single_cell.xlsx"),
                    str(out_path),
                ])
        # M3 critical assertion: even with --memory-mode=auto (request=None),
        # the EFFECTIVE flag reflects what the library actually chose.
        self.assertEqual(
            captured_args.get("_read_only_effective"), True,
            "_read_only_effective must be True when reader._read_only is True "
            "(detected post-open from the actual library decision)",
        )


# ---------------------------------------------------------------------------
# M5 — Temp-file atomic write
# ---------------------------------------------------------------------------

class TestAtomicWrite(unittest.TestCase):
    """M5: Output written via sibling tempfile + atomic os.replace on
    success; unlink-on-failure when emit raises.
    """

    def test_resolve_output_stream_returns_temp_path_for_file_mode(
        self,
    ) -> None:
        """File mode: stream wraps a sibling ``.partial`` tempfile."""
        with tempfile.TemporaryDirectory() as td:
            out_path = Path(td) / "out.md"
            stream, temp_path = _resolve_output_stream(out_path)
            self.assertIsNotNone(temp_path)
            self.assertTrue(temp_path.name.endswith(".partial"))
            self.assertEqual(temp_path.parent, out_path.parent)
            stream.close()
            temp_path.unlink(missing_ok=True)

    def test_resolve_output_stream_returns_none_temp_for_stdout(
        self,
    ) -> None:
        """Stdout mode: no tempfile (stream is sys.stdout, temp_path=None)."""
        stream, temp_path = _resolve_output_stream(None)
        self.assertIsNone(temp_path)
        self.assertIs(stream, sys.stdout)

    def test_emit_failure_unlinks_temp_no_partial_file(self) -> None:
        """When emit_workbook_md raises, the tempfile is unlinked AND
        the final output_path does NOT exist (no partial file).
        """
        def _raising_emit(reader, args, out):
            out.write("## Sheet1\n\n### Table-1\n\n")  # partial write
            out.flush()
            raise RuntimeError("simulated mid-write failure")

        with tempfile.TemporaryDirectory() as td:
            out_path = Path(td) / "out.md"
            with patch(
                "xlsx2md.emit_hybrid.emit_workbook_md", side_effect=_raising_emit,
            ):
                result = main([
                    str(_FIXTURES / "single_cell.xlsx"),
                    str(out_path),
                ])
            self.assertEqual(result, 7, "RuntimeError → InternalError code 7")
            # M5 ASSERTION: no partial output_path on disk.
            self.assertFalse(
                out_path.exists(),
                f"M5 FIX: output path must NOT exist after a failed emit; "
                f"got {out_path} present",
            )
            # And no orphan .partial file either.
            partial = out_path.with_suffix(out_path.suffix + ".partial")
            self.assertFalse(
                partial.exists(),
                f"M5 FIX: tempfile must be unlinked on failure; "
                f"got {partial} present",
            )

    def test_emit_success_atomic_rename(self) -> None:
        """On success, the tempfile is renamed to the final output_path."""
        with tempfile.TemporaryDirectory() as td:
            out_path = Path(td) / "out.md"
            result = main([
                str(_FIXTURES / "single_cell.xlsx"),
                str(out_path),
            ])
            self.assertEqual(result, 0)
            self.assertTrue(out_path.exists())
            # No leftover .partial.
            partial = out_path.with_suffix(out_path.suffix + ".partial")
            self.assertFalse(partial.exists())


# ---------------------------------------------------------------------------
# M6 — Lazy cell_addr (perf, asserted indirectly via mock counter)
# ---------------------------------------------------------------------------

class TestLazyCellAddr(unittest.TestCase):
    """M6: ``get_column_letter`` invoked only for cells WITH a hyperlink
    (warning-path only), not for every cell unconditionally.
    """

    def test_get_column_letter_not_called_for_cells_without_hyperlinks(
        self,
    ) -> None:
        """Direct emit-side: 3-cell row with NO hyperlinks → 0 calls
        to ``get_column_letter``.
        """
        from xlsx_read import TableData, TableRegion  # noqa: PLC0415
        from xlsx2md.emit_gfm import emit_gfm_table  # noqa: PLC0415

        region = TableRegion(
            sheet="Sheet1", top_row=1, left_col=1,
            bottom_row=2, right_col=3, source="gap_detect",
        )
        td = TableData(
            region=region,
            headers=["a", "b", "c"],
            rows=[["x", "y", "z"]],
        )
        out = io.StringIO()
        call_count = 0
        from openpyxl import utils as _xl  # noqa: PLC0415
        real_glc = _xl.get_column_letter

        def _counting_glc(idx: int) -> str:
            nonlocal call_count
            call_count += 1
            return real_glc(idx)

        with patch.object(
            _xl, "get_column_letter", side_effect=_counting_glc,
        ):
            emit_gfm_table(
                td, out,
                hyperlink_allowlist=None,
                hyperlinks_map={},  # no hyperlinks → no warning-path
            )
        self.assertEqual(
            call_count, 0,
            f"M6 FIX: get_column_letter must NOT be called for cells "
            f"without hyperlinks (warning-path only); got {call_count} "
            f"calls for a 3-cell hyperlink-free row.",
        )

    def test_get_column_letter_called_for_cell_with_hyperlink(self) -> None:
        """Sanity: a cell WITH a hyperlink → 1 call (so the warning-path
        DOES build a cell address when needed)."""
        from xlsx_read import TableData, TableRegion  # noqa: PLC0415
        from xlsx2md.emit_gfm import emit_gfm_table  # noqa: PLC0415

        region = TableRegion(
            sheet="Sheet1", top_row=1, left_col=1,
            bottom_row=2, right_col=1, source="gap_detect",
        )
        td = TableData(
            region=region,
            headers=["link"],
            rows=[["click"]],
        )
        out = io.StringIO()
        call_count = 0
        from openpyxl import utils as _xl  # noqa: PLC0415
        real_glc = _xl.get_column_letter

        def _counting_glc(idx: int) -> str:
            nonlocal call_count
            call_count += 1
            return real_glc(idx)

        with patch.object(
            _xl, "get_column_letter", side_effect=_counting_glc,
        ):
            emit_gfm_table(
                td, out,
                hyperlink_allowlist=None,
                # Key (header_band+r_idx, c_idx) = (1, 0).
                hyperlinks_map={(1, 0): "https://ok.com"},
            )
        self.assertEqual(
            call_count, 1,
            "Cell with hyperlink → exactly 1 get_column_letter call",
        )


# ---------------------------------------------------------------------------
# M7 — Hardened --hyperlink-scheme-allowlist help text
# ---------------------------------------------------------------------------

class TestFormulaPlusHyperlinkPreservation(unittest.TestCase):
    """Iter-2 M-NEW-1 fix: cell with BOTH formula AND hyperlink in HTML
    mode must preserve the link (don't blindly overwrite `cell_text=""`
    when the formula path fires).
    """

    def test_formula_plus_hyperlink_keeps_link_and_data_formula(self) -> None:
        """Cell value `"=A1+B1"` + hyperlink `https://docs.example.com`
        + ``--include-formulas`` → `<td data-formula="..."><a href=...>=A1+B1</a></td>`
        with both pieces preserved.
        """
        from xlsx_read import TableData, TableRegion  # noqa: PLC0415
        from xlsx2md.emit_html import emit_html_table  # noqa: PLC0415

        region = TableRegion(
            sheet="Sheet1", top_row=1, left_col=1,
            bottom_row=2, right_col=1, source="gap_detect",
        )
        td = TableData(
            region=region,
            headers=["Result"],
            rows=[["=A1+B1"]],
        )
        out = io.StringIO()
        emit_html_table(
            td, out,
            include_formulas=True,
            hyperlink_allowlist=None,
            # Hyperlink keyed by region-relative offset (header_band + r_idx, c_idx)
            # = (1, 0) for a single-row-header + first body row.
            hyperlinks_map={(1, 0): "https://docs.example.com"},
        )
        output = out.getvalue()
        # M-NEW-1 CRITICAL: link MUST be present (not silently dropped).
        self.assertIn('<a href="https://docs.example.com"', output)
        self.assertIn('data-formula="=A1+B1"', output)
        # Formula string survives as the link's display text (escaped).
        self.assertIn('=A1+B1</a>', output)

    def test_formula_without_hyperlink_emits_blank_cell_content(self) -> None:
        """Regression for the original M1 contract — no hyperlink path."""
        from xlsx_read import TableData, TableRegion  # noqa: PLC0415
        from xlsx2md.emit_html import emit_html_table  # noqa: PLC0415

        region = TableRegion(
            sheet="Sheet1", top_row=1, left_col=1,
            bottom_row=2, right_col=1, source="gap_detect",
        )
        td = TableData(
            region=region,
            headers=["Result"],
            rows=[["=A1+B1"]],
        )
        out = io.StringIO()
        emit_html_table(
            td, out,
            include_formulas=True,
            hyperlink_allowlist=None,
            hyperlinks_map={},  # NO hyperlink
        )
        output = out.getvalue()
        self.assertIn('data-formula="=A1+B1"', output)
        self.assertIn('class="stale-cache"', output)
        # No <a> tag because no hyperlink.
        self.assertNotIn("<a href", output)


class TestPostValidateRunsBeforeAtomicReplace(unittest.TestCase):
    """Iter-2 M-NEW-2 fix: ``_post_validate_output`` runs against the
    TEMP file BEFORE ``os.replace`` publishes it. On validation failure,
    the temp is unlinked and the final ``output_path`` does NOT exist
    on disk (preserving M5's "never leave a partial file" contract at
    the M2 validation boundary).
    """

    def test_validation_failure_no_partial_file_at_output_path(self) -> None:
        """Env-flag on + empty emit → validation fails → output_path
        must NOT exist (no bogus file published).
        """
        def _empty_emit(reader, args, out):
            return 0  # success exit but empty content

        with tempfile.TemporaryDirectory() as td:
            out_path = Path(td) / "out.md"
            with patch(
                "xlsx2md.emit_hybrid.emit_workbook_md",
                side_effect=_empty_emit,
            ), patch.dict(
                os.environ, {"XLSX_XLSX2MD_POST_VALIDATE": "1"},
            ):
                result = main([
                    str(_FIXTURES / "single_cell.xlsx"),
                    str(out_path),
                ])
            self.assertEqual(result, 7)
            # M-NEW-2 ASSERTION: no bogus output_path on disk.
            self.assertFalse(
                out_path.exists(),
                f"M-NEW-2 FIX: validation failure must NOT publish "
                f"the bogus temp via os.replace; got {out_path} present",
            )
            # And no leftover .partial either.
            partial = out_path.with_suffix(out_path.suffix + ".partial")
            self.assertFalse(partial.exists())

    def test_validation_success_atomic_publish(self) -> None:
        """Sanity: when validation passes, output_path is published."""
        with tempfile.TemporaryDirectory() as td:
            out_path = Path(td) / "out.md"
            with patch.dict(
                os.environ, {"XLSX_XLSX2MD_POST_VALIDATE": "1"},
            ):
                result = main([
                    str(_FIXTURES / "single_cell.xlsx"),
                    str(out_path),
                ])
            self.assertEqual(result, 0)
            self.assertTrue(out_path.exists())


class TestHelpTextNamesDangerousSchemes(unittest.TestCase):
    """M7: ``--help`` for ``--hyperlink-scheme-allowlist`` must name
    ``javascript:``, ``data:``, ``vbscript:``, ``file:`` explicitly
    (was a vague "NOT recommended").
    """

    def test_help_names_javascript(self) -> None:
        proc = subprocess.run(
            [sys.executable, str(_SHIM), "--help"],
            capture_output=True, text=True,
        )
        self.assertEqual(proc.returncode, 0)
        self.assertIn("javascript:", proc.stdout)

    def test_help_names_data_vbscript_file(self) -> None:
        proc = subprocess.run(
            [sys.executable, str(_SHIM), "--help"],
            capture_output=True, text=True,
        )
        self.assertIn("data:", proc.stdout)
        self.assertIn("vbscript:", proc.stdout)
        self.assertIn("file:", proc.stdout)

    def test_help_says_DANGEROUS(self) -> None:
        """The escalated wording 'DANGEROUS' must appear (vs 'NOT recommended')."""
        proc = subprocess.run(
            [sys.executable, str(_SHIM), "--help"],
            capture_output=True, text=True,
        )
        self.assertIn("DANGEROUS", proc.stdout)


if __name__ == "__main__":  # pragma: no cover
    unittest.main()
