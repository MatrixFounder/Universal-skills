"""Regression cluster for xlsx2md (task 012-08, ARCH §11 row 012-08).

Five targeted regression tests:

1. R6.h  — no-flag output shape pin (single_cell.xlsx baseline locked).
2. R20   — number-format heuristic: #,##0.00 -> 1,234.50
3. R20   — number-format heuristic: 0% -> 42%
4. R3-H1 — sheet named 'History' emitted verbatim as ## History (NOT ## History_).
5. A-A3  — literal U+203A in header is misdetected deterministically across runs.

Tests 2, 3, 4, and 5 use in-test openpyxl workbook construction to avoid
needing pre-built fixtures for these specific scenarios.
"""
from __future__ import annotations

import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

# Make scripts/ root importable.
_SCRIPTS_DIR = Path(__file__).resolve().parents[2]
if str(_SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS_DIR))

_SHIM = _SCRIPTS_DIR / "xlsx2md.py"
_FIXTURES = Path(__file__).resolve().parent / "fixtures"


def _xlsx2md(*args: str) -> subprocess.CompletedProcess[str]:
    """Invoke xlsx2md.py via subprocess; return CompletedProcess."""
    cmd = [sys.executable, str(_SHIM), *args]
    return subprocess.run(cmd, capture_output=True, text=True)


def _build_workbook_to_tempfile(builder) -> str:
    """Call builder(wb, ws) on a fresh single-sheet workbook; return temp path."""
    import openpyxl  # noqa: PLC0415

    wb = openpyxl.Workbook()
    ws = wb.active
    builder(wb, ws)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        path = tmp.name
    wb.save(path)
    return path


class TestRegressions(unittest.TestCase):
    """Regression battery for xlsx-9 (task 012-08 final-gate assertions)."""

    # ------------------------------------------------------------------ R6.h
    def test_no_flag_omitted_shape_pin(self) -> None:
        """R6.h — no-flag output shape EXACTLY matches the locked baseline.

        Baseline: single_cell.xlsx (1×1, A1='hello') → GFM output with
        exactly one H2, one H3, one data row, one separator row.
        The test pins the SHAPE (line count / structure), not pixel bytes,
        to tolerate minor whitespace changes; actual content is asserted.
        """
        proc = _xlsx2md(str(_FIXTURES / "single_cell.xlsx"))
        self.assertEqual(proc.returncode, 0, f"Unexpected exit: {proc.stderr!r}")
        lines = [l for l in proc.stdout.splitlines() if l.strip()]
        # single_cell.xlsx has exactly 1 row (A1='hello'), treated as header.
        # Expected non-blank lines: '## Sheet1', '### Table-1', header, sep
        # (no data rows — the single cell is the header).
        self.assertEqual(len(lines), 4, f"Expected 4 non-blank lines; got {lines!r}")
        self.assertEqual(lines[0], "## Sheet1")
        self.assertEqual(lines[1], "### Table-1")
        self.assertIn("hello", lines[2])   # header row
        self.assertIn("---", lines[3])     # GFM separator

    # ------------------------------------------------------------------ R20
    def test_number_format_heuristic_thousand_separator(self) -> None:
        """R20 — cell with number_format='#,##0.00' and value 1234.5 -> '1,234.50'."""
        def _build(wb, ws):
            ws.title = "Sheet1"
            ws["A1"] = "Amount"
            ws["A2"] = 1234.5
            ws["A2"].number_format = "#,##0.00"

        path = _build_workbook_to_tempfile(_build)
        try:
            proc = _xlsx2md(path)
            self.assertEqual(proc.returncode, 0)
            self.assertIn(
                "1,234.50",
                proc.stdout,
                f"Expected '1,234.50' in output; got:\n{proc.stdout!r}",
            )
        finally:
            Path(path).unlink(missing_ok=True)

    def test_number_format_heuristic_percent(self) -> None:
        """R20 — cell with number_format='0%' and value 0.42 -> '42%'."""
        def _build(wb, ws):
            ws.title = "Sheet1"
            ws["A1"] = "Rate"
            ws["A2"] = 0.42
            ws["A2"].number_format = "0%"

        path = _build_workbook_to_tempfile(_build)
        try:
            proc = _xlsx2md(path)
            self.assertEqual(proc.returncode, 0)
            self.assertIn(
                "42%",
                proc.stdout,
                f"Expected '42%' in output; got:\n{proc.stdout!r}",
            )
        finally:
            Path(path).unlink(missing_ok=True)

    # ------------------------------------------------------------------ R3-H1
    def test_sheet_name_verbatim_no_sanitisation(self) -> None:
        """R3-H1 — sheet named 'History' emitted as '## History' (NOT '## History_').

        xlsx-3 (md_tables2xlsx) sanitises 'History' → 'History_' on write-back,
        but xlsx-9 MUST emit the verbatim sheet name.  This is documented as
        expected asymmetry in references/xlsx-md-shapes.md §6.
        """
        def _build(wb, ws):
            ws.title = "History"
            ws["A1"] = "Event"
            ws["A2"] = "item1"

        path = _build_workbook_to_tempfile(_build)
        try:
            proc = _xlsx2md(path)
            self.assertEqual(proc.returncode, 0)
            self.assertIn(
                "## History",
                proc.stdout,
                f"Expected '## History' (verbatim) in output; got:\n{proc.stdout!r}",
            )
            self.assertNotIn(
                "## History_",
                proc.stdout,
                "Sheet name was incorrectly sanitised to 'History_'",
            )
        finally:
            Path(path).unlink(missing_ok=True)

    # ------------------------------------------------------------------ A-A3
    def test_literal_u203a_in_header_consistent_across_runs(self) -> None:
        """A-A3 — literal U+203A in header cell value is misdetected consistently.

        The ` › ` (U+203A with surrounding spaces) separator is used by xlsx-9
        for multi-row header reconstruction.  A cell whose value literally
        contains ' › ' will be misinterpreted as a multi-row header.  This is
        a documented honest-scope limitation (A-A3).

        The regression asserts DETERMINISM: two independent runs on the same
        workbook produce byte-identical output.  The misinterpretation is
        accepted as long as it is reproducible (no non-determinism from e.g.
        dict ordering in spans computation).
        """
        def _build(wb, ws):
            ws.title = "Sheet1"
            ws["A1"] = "foo › bar"   # literal U+203A surrounded by spaces
            ws["A2"] = "val1"

        path = _build_workbook_to_tempfile(_build)
        try:
            proc1 = _xlsx2md(path)
            proc2 = _xlsx2md(path)
            self.assertEqual(proc1.returncode, 0)
            self.assertEqual(proc2.returncode, 0)
            self.assertEqual(
                proc1.stdout,
                proc2.stdout,
                "Two runs on the same workbook with literal U+203A header "
                "produced different output (non-deterministic misinterpretation).",
            )
            # Confirm both runs produced non-empty output (not a crash).
            self.assertTrue(proc1.stdout.strip(), "Output was unexpectedly empty")
        finally:
            Path(path).unlink(missing_ok=True)


if __name__ == "__main__":
    unittest.main()
