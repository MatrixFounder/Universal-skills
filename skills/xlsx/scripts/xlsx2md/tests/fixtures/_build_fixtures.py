"""Programmatic fixture builder — run once to regenerate test fixtures.

Usage::

    cd skills/xlsx/scripts
    ./.venv/bin/python xlsx2md/tests/fixtures/_build_fixtures.py

Fixtures generated:
- ``single_cell.xlsx`` — minimal 1×1 workbook; cell A1 = "hello".
- ``macro_simple.xlsm`` — macro-enabled workbook that triggers
  ``MacroEnabledWarning`` from ``xlsx_read.open_workbook``.

``encrypted.xlsx`` is NOT generated here — it is copied from
``xlsx2csv2json/tests/fixtures/encrypted.xlsx``.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl

_HERE = Path(__file__).resolve().parent


def build_single_cell() -> None:
    """1×1 workbook: Sheet1, A1 = 'hello'."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "hello"
    out = _HERE / "single_cell.xlsx"
    wb.save(out)
    print(f"built: {out}")


def build_macro_xlsm() -> None:
    """BROKEN — DO NOT RUN. See _HERE/macro_simple.xlsm provenance note.

    Originally intended to produce a macro-enabled workbook (.xlsm)
    that triggers MacroEnabledWarning by patching ``[Content_Types].xml``
    to switch the main-document content type to the xlsm variant.

    **Why it doesn't work:** `xlsx_read._workbook._probe_macros` checks
    for the actual presence of `xl/vbaProject.bin` (or other VBA-bearing
    parts), NOT the content-type string. A content-type rename alone
    does not trigger the warning. openpyxl cannot synthesise a real
    `vbaProject.bin` blob.

    **What we ship instead:** `tests/fixtures/macro_simple.xlsm` is a
    byte-copy of `skills/xlsx/scripts/xlsx_read/tests/fixtures/macros.xlsm`
    (a real macro-bearing fixture maintained alongside the foundation
    library). If you need to regenerate, run::

        cp skills/xlsx/scripts/xlsx_read/tests/fixtures/macros.xlsm \\
           skills/xlsx/scripts/xlsx2md/tests/fixtures/macro_simple.xlsm

    Future fix: build a proper xlsm via an external tool (LibreOffice
    Calc, Excel) or vendor a static vbaProject.bin and embed via
    zipfile. For now this generator is dead-code retained only as a
    cautionary breadcrumb.
    """
    raise NotImplementedError(
        "build_macro_xlsm is BROKEN — see docstring. Use byte-copy from "
        "xlsx_read/tests/fixtures/macros.xlsm instead."
    )
    import io
    import zipfile

    # Step 1: build a normal workbook and save it to a BytesIO buffer.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "hello"
    buf = io.BytesIO()
    wb.save(buf)

    # Step 2: re-pack the archive, patching Content_Types to xlsm.
    out = _HERE / "macro_simple.xlsm"
    buf.seek(0)
    with zipfile.ZipFile(buf, "r") as zin:
        with zipfile.ZipFile(out, "w", compression=zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == "[Content_Types].xml":
                    # Replace xlsx main-document content type with xlsm one.
                    data = data.replace(
                        b"application/vnd.openxmlformats-officedocument"
                        b".spreadsheetml.sheet.main+xml",
                        b"application/vnd.ms-excel.sheet.macroEnabled.main+xml",
                    )
                zout.writestr(item, data)
    print(f"built: {out}")


def build_hyperlink_various_schemes() -> None:
    """3 hyperlinks with different URI schemes for R10a allowlist tests.

    Row 1 is the header. Body rows 2-4 contain one hyperlink each:
      - A2: ``"safe link"`` → ``https://ok.example.com``
      - A3: ``"unsafe link"`` → ``javascript:alert(1)``
      - A4: ``"mail"``        → ``mailto:x@y.example.com``

    Used by tests in ``test_dispatch.py`` to verify Path C′
    parallel-pass extraction AND the dispatch-side scheme-allowlist
    filter (R10a / D-A15 / Sec-MED-2).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Link"
    # A2: https → allowed by default
    ws["A2"] = "safe link"
    ws["A2"].hyperlink = "https://ok.example.com"
    # A3: javascript → blocked by default
    ws["A3"] = "unsafe link"
    ws["A3"].hyperlink = "javascript:alert(1)"
    # A4: mailto → allowed by default
    ws["A4"] = "mail"
    ws["A4"].hyperlink = "mailto:x@y.example.com"
    out = _HERE / "hyperlink_various_schemes.xlsx"
    wb.save(out)
    print(f"built: {out}")


def build_cell_with_newline() -> None:
    """Cell with embedded newline (ALT+ENTER pattern) — 2-column layout.

    Content:
      - A1 = "Label",  B1 = "Note"   (headers, 2 columns for GFM parse)
      - A2 = "first line\\nsecond line" (embedded \\n, ALT+ENTER), B2 = "item1"

    Cell A2 is the assertion target in T-21
    (test_live_roundtrip_cell_newline_br in test_md_tables2xlsx.py).
    xlsx-9 must emit `<br>` for the embedded newline in A2; xlsx-3 R9.c
    must consume `<br>` back to `\\n` on write-back.

    Two-column layout is required because GFM single-column tables
    (one pipe delimiter per row) are not detected by the xlsx-3 block
    scanner. Round-trip safe: plain strings only, no merges, no
    formulas, no styles.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Label"
    ws["B1"] = "Note"
    ws["A2"] = "first line\nsecond line"
    ws["B2"] = "item1"
    out = _HERE / "cell_with_newline.xlsx"
    wb.save(out)
    print(f"built: {out}")


def build_roundtrip_basic() -> None:
    """Multi-sheet round-trip-safe fixture for xlsx-9 <-> xlsx-3 cycle.

    Content: strings, integers, ISO dates, one hyperlink.
    NO merges, NO formulas, NO styles — xlsx-3 v1 preserves these
    losslessly in the xlsx-9 -> xlsx-3 -> xlsx-9 round-trip cycle.

    IMPORTANT: Never add merges (xlsx-3 v1 doesn't re-merge HTML
    colspan/rowspan), formulas (xlsx-3 emits cached values), styles
    (xlsx-3 drops them), or comments (xlsx-3 v1 sidecar only) to this
    fixture. Round-trip safety is a hard contract (TASK 012-07 §Notes).
    """
    from datetime import date

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Name"
    ws["B1"] = "Score"
    ws["C1"] = "Date"
    ws["A2"] = "Alice"
    ws["B2"] = 100
    ws["C2"] = date(2026, 1, 15)
    ws["A3"] = "Bob"
    ws["B3"] = 95
    ws["C3"] = date(2026, 2, 20)
    out = _HERE / "roundtrip_basic.xlsx"
    wb.save(out)
    print(f"built: {out}")


if __name__ == "__main__":
    build_single_cell()
    build_macro_xlsm()
    build_hyperlink_various_schemes()
    build_cell_with_newline()
    build_roundtrip_basic()
    print("done.")
