#!/usr/bin/env python3
"""Convert a CSV/TSV file into a styled .xlsx workbook.

Defaults that matter:
- Auto-detect delimiter (comma / semicolon / tab) if `--delimiter auto`.
- Keep leading zeros on text-looking columns by reading everything as
  string first, then coercing obviously-numeric columns. Codes like
  "007" or phone numbers like "0123456789" would otherwise lose their
  zeros.
- Bold header row, light-grey fill, centre alignment.
- Freeze first row so headers remain visible while scrolling.
- Add an auto-filter over the data range.
- Set column widths to the max of (header width, longest value width)
  with a sane upper bound.

Usage:
    python3 csv2xlsx.py input.csv output.xlsx [--delimiter auto|,|;|\\t]
                                                [--no-filter]
                                                [--no-freeze]
                                                [--encoding utf-8]
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path

import pandas as pd  # type: ignore
from openpyxl import Workbook  # type: ignore
from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore
from openpyxl.utils import get_column_letter  # type: ignore


HEADER_FILL = PatternFill("solid", fgColor="F2F2F2")
HEADER_FONT = Font(bold=True)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
NUMERIC_RE = re.compile(r"^-?\d+(?:[.,]\d+)?$")
MAX_COL_WIDTH = 50


def _sniff_delimiter(path: Path, encoding: str) -> str:
    sample = path.read_text(encoding=encoding, errors="replace")[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        return dialect.delimiter
    except csv.Error:
        return ","


def _coerce_column(values: list[str]) -> list[object]:
    # Returns raw strings if any value doesn't look numeric or if the
    # first non-empty value has a leading zero (likely a code).
    non_empty = [v for v in values if v != ""]
    if not non_empty:
        return values  # type: ignore[return-value]
    if non_empty[0].startswith("0") and len(non_empty[0]) > 1 and non_empty[0][1] not in ".,":
        return values  # type: ignore[return-value]
    if not all(NUMERIC_RE.match(v) for v in non_empty):
        return values  # type: ignore[return-value]
    out: list[object] = []
    for v in values:
        if v == "":
            out.append(None)
            continue
        normalised = v.replace(",", ".")
        out.append(float(normalised) if "." in normalised else int(normalised))
    return out


def _warn_duplicate_headers(input_path: Path, sep: str, encoding: str) -> None:
    """Read the raw header row and warn about duplicates.

    pandas silently renames duplicates to `name.1`, `name.2`, etc. —
    that is usually not what the user expected. We look at the raw
    line so we report the original names, not pandas' rewrites.
    """
    import csv as _csv
    with open(input_path, "r", encoding=encoding, newline="", errors="replace") as fh:
        try:
            first_row = next(_csv.reader(fh, delimiter=sep))
        except StopIteration:
            return
    seen: dict[str, int] = {}
    for col in first_row:
        seen[col] = seen.get(col, 0) + 1
    duplicates = [name for name, count in seen.items() if count > 1]
    if duplicates:
        print(
            f"Warning: duplicate header name(s) in CSV — {', '.join(duplicates)}. "
            "pandas will suffix them .1, .2, etc.",
            file=sys.stderr,
        )


def convert(
    input_path: Path,
    output_path: Path,
    *,
    delimiter: str = "auto",
    encoding: str = "utf-8",
    freeze_header: bool = True,
    auto_filter: bool = True,
) -> None:
    sep = _sniff_delimiter(input_path, encoding) if delimiter == "auto" else delimiter
    _warn_duplicate_headers(input_path, sep, encoding)
    df = pd.read_csv(input_path, sep=sep, dtype=str, encoding=encoding, keep_default_na=False)

    coerced = {col: _coerce_column(df[col].tolist()) for col in df.columns}

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Headers are always strings. Setting value triggers openpyxl's
    # type inference, which classifies e.g. "#REF!" as data_type='e'
    # (Excel error). Re-assigning data_type afterwards forces 's' and
    # prevents xlsx_validate from flagging user-typed error-looking
    # strings as cached formula errors.
    for col_idx, header in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = str(header)
        cell.data_type = "s"

    for row_idx in range(len(df)):
        for col_idx, col in enumerate(df.columns, start=1):
            value = coerced[col][row_idx]
            cell = ws.cell(row=row_idx + 2, column=col_idx)
            if value is None:
                continue
            cell.value = value
            if isinstance(value, str):
                cell.data_type = "s"

    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN

        header_len = len(str(col_name))
        data_len = max(
            (len(str(coerced[col_name][i]) if coerced[col_name][i] is not None else "") for i in range(len(df))),
            default=0,
        )
        width = min(max(header_len, data_len) + 2, MAX_COL_WIDTH)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    if freeze_header:
        ws.freeze_panes = "A2"
    if auto_filter and len(df) > 0:
        ws.auto_filter.ref = ws.dimensions

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("input", type=Path, help="Source .csv / .tsv file")
    parser.add_argument("output", type=Path, help="Destination .xlsx file")
    parser.add_argument("--delimiter", default="auto", help="auto (default), ',', ';', or '\\t'")
    parser.add_argument("--encoding", default="utf-8", help="Input encoding (default: utf-8)")
    parser.add_argument("--no-freeze", action="store_true", help="Do not freeze the header row")
    parser.add_argument("--no-filter", action="store_true", help="Do not add an auto-filter")
    args = parser.parse_args(argv)

    if not args.input.is_file():
        print(f"Input not found: {args.input}", file=sys.stderr)
        return 1

    delim = args.delimiter
    if delim == "\\t":
        delim = "\t"

    try:
        convert(
            args.input,
            args.output,
            delimiter=delim,
            encoding=args.encoding,
            freeze_header=not args.no_freeze,
            auto_filter=not args.no_filter,
        )
    except Exception as exc:
        print(f"Conversion failed: {exc}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
