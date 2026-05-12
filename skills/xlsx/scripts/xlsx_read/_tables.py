"""F4 — 3-tier table detector.

Tier-1: Excel Tables (`<table>` parts in `xl/tables/tableN.xml`).
        Discovered through `ws.tables` (openpyxl exposes them as a
        `TableList` keyed by displayName).
Tier-2: sheet-scope named ranges (`<definedName localSheetId=...>`).
        Workbook-scope names are silently ignored per honest-scope (d).
Tier-3: gap-detection (consecutive empty rows ≥ `gap_rows` AND/OR
        consecutive empty cols ≥ `gap_cols`) over the area not already
        claimed by Tier-1 / Tier-2.

Mode semantics:
- "auto"        : 1 → 2 → 3 fallthrough (default).
- "tables-only" : Tier-1 + Tier-2 only; no gap-detection.
- "whole"       : a single region spanning `ws.dimensions`.

UC-03 A4 (ListObject vs named range overlap): Tier-1 wins — Tier-2
ranges that overlap a Tier-1 region are dropped. Tier-3 is similarly
bounded by previously-claimed bounding boxes.
"""

from __future__ import annotations

from typing import Any

from openpyxl.utils.cell import range_boundaries

from ._types import TableDetectMode, TableRegion


def detect_tables(
    wb: Any,
    sheet_name: str,
    *,
    mode: TableDetectMode = "auto",
    gap_rows: int = 2,
    gap_cols: int = 1,
) -> list[TableRegion]:
    """Detect tables on `sheet_name`. See module docstring for tier rules."""
    if mode not in ("auto", "tables-only", "whole"):
        raise ValueError(f"Unknown mode: {mode!r}")
    ws = wb[sheet_name]

    if mode == "whole":
        return [_whole_sheet_region(ws, sheet_name)]

    tier1 = _listobjects_for_sheet(ws, sheet_name)
    tier2 = _named_ranges_for_sheet(wb, sheet_name, claimed=tier1)
    claimed = tier1 + tier2

    if mode == "tables-only":
        return claimed

    tier3 = _gap_detect(ws, sheet_name, claimed, gap_rows, gap_cols)
    return claimed + tier3


# ---------------------------------------------------------------------------
# Tier-1 — Excel Tables (ListObjects)
# ---------------------------------------------------------------------------


def _listobjects_for_sheet(ws: Any, sheet_name: str) -> list[TableRegion]:
    """Emit one region per `<table>` part bound to `ws`.

    openpyxl ≥ 3.1 surfaces tables via `ws.tables` — a `TableList`
    keyed by `displayName`. **Quirk:** `TableList.items()` yields
    `(name, ref_string)` pairs (NOT `(name, Table)`), while
    `TableList[name]` returns the full `Table` object — verified
    empirically against openpyxl 3.1.5. We iterate keys and look up
    each entry to get the `headerRowCount` attribute.
    """
    tables = getattr(ws, "tables", None)
    if not tables:
        return []
    out: list[TableRegion] = []
    for name in list(tables.keys()):
        table = tables[name]
        ref = getattr(table, "ref", None)
        if not ref:
            continue
        left_col, top_row, right_col, bottom_row = range_boundaries(ref)
        hrc = getattr(table, "headerRowCount", 1)
        # openpyxl uses `None` to mean "default = 1"; normalise.
        if hrc is None:
            hrc = 1
        out.append(
            TableRegion(
                sheet=sheet_name,
                top_row=top_row,
                left_col=left_col,
                bottom_row=bottom_row,
                right_col=right_col,
                source="listobject",
                name=name,
                listobject_header_row_count=hrc,
            )
        )
    return out


# ---------------------------------------------------------------------------
# Tier-2 — sheet-scope named ranges
# ---------------------------------------------------------------------------


def _named_ranges_for_sheet(
    wb: Any, sheet_name: str, claimed: list[TableRegion]
) -> list[TableRegion]:
    """Emit one region per sheet-scope `<definedName>` matching `sheet_name`.

    **openpyxl 3.1+ split:** sheet-scope names live on
    `ws.defined_names` (per-sheet `DefinedNameDict`); workbook-scope
    names live on `wb.defined_names`. The library reads ONLY from
    `ws.defined_names` so workbook-scope names are dropped by
    construction (honest-scope D8, no `localSheetId` filtering
    needed). Ranges that overlap a Tier-1 region are dropped
    (UC-03 A4).
    """
    ws = wb[sheet_name]
    out: list[TableRegion] = []
    sheet_defined = getattr(ws, "defined_names", None)
    if not sheet_defined:
        return out
    for name in list(sheet_defined):
        dn = sheet_defined[name]
        # **L-M3 fix:** a defined name with multiple comma-separated
        # destinations on the same sheet yields multiple regions; we
        # disambiguate with `#1`, `#2` suffixes so downstream consumers
        # that index by `region.name` don't silently collide.
        dest_count = 0
        local_regions: list[TableRegion] = []
        for dest_sheet, dest_range in dn.destinations:
            if dest_sheet != sheet_name:
                # Destination resolves to a *different* sheet — skip.
                continue
            try:
                left_col, top_row, right_col, bottom_row = range_boundaries(dest_range)
            except (TypeError, ValueError):
                continue
            dest_count += 1
            local_regions.append(TableRegion(
                sheet=sheet_name,
                top_row=top_row,
                left_col=left_col,
                bottom_row=bottom_row,
                right_col=right_col,
                source="named_range",
                name=name,
            ))
        # Single-destination → name stays bare; multi-destination →
        # suffix with `#1`, `#2`, ... in iteration order.
        for i, region in enumerate(local_regions):
            final_name = (
                region.name if dest_count == 1
                else f"{region.name}#{i + 1}"
            )
            disambiguated = TableRegion(
                sheet=region.sheet,
                top_row=region.top_row,
                left_col=region.left_col,
                bottom_row=region.bottom_row,
                right_col=region.right_col,
                source=region.source,
                name=final_name,
            )
            if not _has_overlap(disambiguated, claimed):
                out.append(disambiguated)
    return out


# ---------------------------------------------------------------------------
# Tier-3 — gap-detection
# ---------------------------------------------------------------------------


# Defensive cap on the bounding box we will allocate for gap-detect.
# Malformed `<dimension ref="A1:XFD1048576"/>` would otherwise force a
# 137 GB allocation (M8 sibling concern: pathological openpyxl
# `max_row × max_col` from corrupted-or-malicious workbooks). The cap
# is generous for real-world inputs (1M cells well above the 10K × 20
# perf budget) and refuses anything beyond it with a clear error.
_GAP_DETECT_MAX_CELLS: int = 1_000_000


def _gap_detect(
    ws: Any,
    sheet_name: str,
    claimed: list[TableRegion],
    gap_rows: int,
    gap_cols: int,
) -> list[TableRegion]:
    """Split the un-claimed area on ≥ gap_rows empty rows / ≥ gap_cols empty cols.

    Single-pass implementation backed by openpyxl's row-stream
    iterator (`iter_rows`), so the cost scales with **actually
    populated cells**, not with `ws.max_row × ws.max_column`. The
    streaming form is also safe in `read_only=True` mode (the
    historical nested `ws.cell(r, c)` pattern is fatal there —
    each call re-walks the sheet XML).

    1. Bounds: `ws.calculate_dimension()` parsed via openpyxl's
       `range_boundaries`. Refuses inputs with > `_GAP_DETECT_MAX_CELLS`
       cells in the dimension bbox (defense against malformed
       `<dimension>` payloads).
    2. Stream rows via `iter_rows(values_only=True)`. Build the
       occupancy matrix incrementally, masking out cells inside
       any `claimed` region as they pass.
    3. Find row-bands → col-bands → tight bbox per region.
    """
    top, left, bottom, right = _resolve_dim_bbox(ws)
    if top is None:
        return []
    n_rows = bottom - top + 1
    n_cols = right - left + 1
    if n_rows * n_cols > _GAP_DETECT_MAX_CELLS:
        # The bbox is implausibly large for a real spreadsheet — refuse
        # rather than allocate a multi-GB grid. The caller can override
        # by passing `mode="whole"` (one region, no occupancy work) or
        # `mode="tables-only"` (no gap-detect at all).
        raise ValueError(
            f"_gap_detect: dimension bbox {n_rows}×{n_cols}={n_rows * n_cols} "
            f"exceeds {_GAP_DETECT_MAX_CELLS} cells; switch to "
            f"mode='tables-only' or mode='whole' on this sheet"
        )

    claimed_mask = _build_claimed_mask(top, left, bottom, right, claimed)

    # Stream once, mask-and-classify per cell. `values_only=True` keeps
    # the iterator producing Python primitives (no openpyxl Cell
    # objects → no random-access cost in read_only mode).
    occupancy: list[list[bool]] = []
    for r_idx, row in enumerate(
        ws.iter_rows(
            min_row=top, max_row=bottom,
            min_col=left, max_col=right,
            values_only=True,
        )
    ):
        row_occ: list[bool] = []
        for c_idx, v in enumerate(row):
            if claimed_mask[r_idx][c_idx]:
                row_occ.append(False)
            else:
                row_occ.append(v is not None and v != "")
        # Pad if the row is shorter than expected (sparse rows in
        # streaming mode may yield fewer columns).
        if len(row_occ) < n_cols:
            row_occ.extend([False] * (n_cols - len(row_occ)))
        occupancy.append(row_occ)
    # Pad missing rows (iter_rows in read_only mode may skip empties).
    while len(occupancy) < n_rows:
        occupancy.append([False] * n_cols)

    # Step 2: find row-bands.
    row_bands = _split_on_gap(
        [any(r) for r in occupancy], gap_rows, base_index=top
    )
    if not row_bands:
        return []

    out: list[TableRegion] = []
    counter = 1
    for r_start, r_end in row_bands:
        # Slice occupancy rows in this band.
        band_rows = occupancy[r_start - top : r_end - top + 1]
        # Step 3: find col-bands in this row-band — collapse the band
        # vertically: a column is occupied iff *any* row in the band
        # uses it.
        col_occupancy = [any(band_rows[i][c] for i in range(len(band_rows)))
                          for c in range(right - left + 1)]
        col_bands = _split_on_gap(col_occupancy, gap_cols, base_index=left)
        for c_start, c_end in col_bands:
            # Trim each col-band to the actual tight bbox within the
            # row-band so leading/trailing empty edges don't bloat the
            # region.
            tight = _tight_bbox(occupancy, r_start - top, r_end - top,
                                c_start - left, c_end - left)
            if tight is None:
                continue
            tr, tc, br, bc = tight
            out.append(
                TableRegion(
                    sheet=sheet_name,
                    top_row=top + tr,
                    left_col=left + tc,
                    bottom_row=top + br,
                    right_col=left + bc,
                    source="gap_detect",
                    name=f"Table-{counter}",
                )
            )
            counter += 1
    return out


def _resolve_dim_bbox(ws: Any) -> tuple[int, int, int, int] | tuple[None, None, None, None]:
    """Return the (top, left, bottom, right) inclusive bbox of `ws.dimensions`.

    Falls back to (max_row, max_column) when the dimension parse fails.
    Returns four `None`s for a genuinely empty sheet.
    """
    dims = getattr(ws, "dimensions", None) or ""
    if dims:
        try:
            left, top, right, bottom = range_boundaries(dims)
            return top, left, bottom, right
        except (TypeError, ValueError):
            pass
    if ws.max_row and ws.max_column:
        return 1, 1, ws.max_row, ws.max_column
    return None, None, None, None  # type: ignore[return-value]


def _build_claimed_mask(
    top: int, left: int, bottom: int, right: int, claimed: list[TableRegion]
) -> list[list[bool]]:
    h, w = bottom - top + 1, right - left + 1
    mask = [[False] * w for _ in range(h)]
    for region in claimed:
        for r in range(region.top_row, region.bottom_row + 1):
            for c in range(region.left_col, region.right_col + 1):
                if top <= r <= bottom and left <= c <= right:
                    mask[r - top][c - left] = True
    return mask


def _split_on_gap(
    occupied: list[bool], gap: int, base_index: int = 1
) -> list[tuple[int, int]]:
    """Find runs of occupied indices, splitting on `gap`-or-more empties.

    Returns list of (start, end) inclusive in the **absolute** index
    space defined by `base_index` (so e.g. row 1-based or col 1-based
    can be returned without renumbering at the call site).

    A gap of *exactly* `gap` empties splits — e.g. `gap=2` splits on
    `[True, False, False, True]` (2 empties = split) but NOT on
    `[True, False, True]` (1 empty < 2 = same band).
    """
    out: list[tuple[int, int]] = []
    run_start: int | None = None
    empty_run = 0
    for i, occ in enumerate(occupied):
        if occ:
            if run_start is None:
                run_start = i
            empty_run = 0
        else:
            if run_start is not None:
                empty_run += 1
                if empty_run >= gap:
                    out.append((base_index + run_start, base_index + i - empty_run))
                    run_start = None
                    empty_run = 0
    if run_start is not None:
        out.append((base_index + run_start, base_index + len(occupied) - 1 - empty_run))
    return out


def _tight_bbox(
    occupancy: list[list[bool]],
    r_lo: int,
    r_hi: int,
    c_lo: int,
    c_hi: int,
) -> tuple[int, int, int, int] | None:
    """Return the tightest occupied bbox inside the band, or None if empty.

    Returned tuple is `(top, left, bottom, right)` in 0-based grid
    coordinates relative to the occupancy matrix.
    """
    tr, tc, br, bc = None, None, None, None
    for r in range(r_lo, r_hi + 1):
        for c in range(c_lo, c_hi + 1):
            if occupancy[r][c]:
                if tr is None or r < tr:
                    tr = r
                if br is None or r > br:
                    br = r
                if tc is None or c < tc:
                    tc = c
                if bc is None or c > bc:
                    bc = c
    if tr is None:
        return None
    return tr, tc, br, bc  # type: ignore[return-value]


# ---------------------------------------------------------------------------
# Overlap helper
# ---------------------------------------------------------------------------


def _has_overlap(region: TableRegion, claimed: list[TableRegion]) -> bool:
    for c in claimed:
        if (
            region.top_row <= c.bottom_row
            and c.top_row <= region.bottom_row
            and region.left_col <= c.right_col
            and c.left_col <= region.right_col
        ):
            return True
    return False


def _whole_sheet_region(ws: Any, sheet_name: str) -> TableRegion:
    return TableRegion(
        sheet=sheet_name,
        top_row=1,
        left_col=1,
        bottom_row=max(ws.max_row, 1),
        right_col=max(ws.max_column, 1),
        source="gap_detect",
        name="Table-1",
    )
