"""Task 009-06 — F5 `_headers.py` E2E + unit tests."""

from __future__ import annotations

import unittest

import openpyxl

from xlsx_read import TableRegion, open_workbook
from xlsx_read import _headers
from xlsx_read.tests.conftest import FIXTURES_DIR


def _whole_sheet_region(ws, sheet_name: str = "Sheet") -> TableRegion:
    return TableRegion(
        sheet=sheet_name,
        top_row=1,
        left_col=1,
        bottom_row=ws.max_row,
        right_col=ws.max_column,
        source="gap_detect",
    )


class TestDetectHeaderBand(unittest.TestCase):
    """TC-E2E-01..-03: header-band auto-detect across 3 fixtures."""

    def test_single_row(self) -> None:
        with open_workbook(FIXTURES_DIR / "headers_single_row.xlsx") as r:
            ws = r._wb.active
            region = _whole_sheet_region(ws)
            self.assertEqual(_headers.detect_header_band(ws, region, "auto"), 1)

    def test_two_row_merged(self) -> None:
        with open_workbook(FIXTURES_DIR / "headers_two_row_merged.xlsx") as r:
            ws = r._wb.active
            region = _whole_sheet_region(ws)
            self.assertEqual(_headers.detect_header_band(ws, region, "auto"), 2)

    def test_three_row(self) -> None:
        with open_workbook(FIXTURES_DIR / "headers_three_row.xlsx") as r:
            ws = r._wb.active
            region = _whole_sheet_region(ws)
            # Fixture geometry: row 1 has merges B1:C1 + D1:E1 (anchored
            # row 1, max_row=1). Rows 2-3 have no column-spanning merges.
            # Under the TASK 010 §11 contiguous-from-top algorithm, the
            # band runs row 1 (anchored) + row 2 (sub-labels under the
            # banner). Row 3 starts the body. Pinned to 2 — the prior
            # `assertIn(band, (2, 3))` slack was a pre-rewrite tolerance.
            self.assertEqual(_headers.detect_header_band(ws, region, "auto"), 2)


class TestHeaderBandContiguousFromTop(unittest.TestCase):
    """TASK 010 §11 patch: scattered deep-body merges must NOT inflate
    the auto-detected header band. Implementation must match the
    docstring contract: "the first row WITHOUT such a merge ends the
    header band (minimum 1 row)".
    """

    def test_deep_body_merge_does_not_inflate_band(self) -> None:
        """Row 1 banner; rows 2-49 plain; row 50 another banner. Pre-patch
        the function returned band = 50 (deepest anchor + 1). Post-patch
        returns 2 (1 anchored + 1 sub-labels row), because row 2 breaks
        the contiguous-from-top run.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.merge_cells("A1:B1")
        ws["A1"] = "Banner"
        ws["A2"] = "Sub-A"
        ws["B2"] = "Sub-B"
        for r in range(3, 50):
            ws.cell(row=r, column=1).value = r
            ws.cell(row=r, column=2).value = r * 10
        ws.merge_cells("A50:B50")
        ws["A50"] = "Totals"
        region = TableRegion(
            sheet="Sheet", top_row=1, left_col=1,
            bottom_row=50, right_col=2, source="gap_detect",
        )
        self.assertEqual(_headers.detect_header_band(ws, region, "auto"), 2)

    def test_no_merges_returns_one(self) -> None:
        """No column-spanning merges anywhere in the region → default 1."""
        wb = openpyxl.Workbook()
        ws = wb.active
        for c, h in enumerate(["A", "B", "C"], start=1):
            ws.cell(row=1, column=c).value = h
        for r in range(2, 11):
            for c in range(1, 4):
                ws.cell(row=r, column=c).value = r * c
        region = TableRegion(
            sheet="Sheet", top_row=1, left_col=1,
            bottom_row=10, right_col=3, source="gap_detect",
        )
        self.assertEqual(_headers.detect_header_band(ws, region, "auto"), 1)

    def test_contiguous_two_row_banner(self) -> None:
        """Row 1 + Row 2 banners; Row 3 plain → band covers rows 1-2 +
        sub-labels row 3 = 3 total."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.merge_cells("A1:C1")
        ws.merge_cells("A2:C2")
        ws["A1"] = "L0"
        ws["A2"] = "L1"
        ws["A3"] = "h1"
        ws["B3"] = "h2"
        ws["C3"] = "h3"
        for r in range(4, 8):
            ws.cell(row=r, column=1).value = r
        region = TableRegion(
            sheet="Sheet", top_row=1, left_col=1,
            bottom_row=7, right_col=3, source="gap_detect",
        )
        self.assertEqual(_headers.detect_header_band(ws, region, "auto"), 3)


class TestHintPassthrough(unittest.TestCase):
    """TC-E2E-06: integer hint passes through verbatim."""

    def test_int_hint_returns_verbatim(self) -> None:
        with open_workbook(FIXTURES_DIR / "headers_single_row.xlsx") as r:
            ws = r._wb.active
            region = _whole_sheet_region(ws)
            self.assertEqual(_headers.detect_header_band(ws, region, 5), 5)
            self.assertEqual(_headers.detect_header_band(ws, region, 0), 0)


class TestInvalidHints(unittest.TestCase):
    """Defensive: invalid hints raise ValueError."""

    def test_negative_int_rejected(self) -> None:
        wb = openpyxl.Workbook()
        region = _whole_sheet_region(wb.active)
        with self.assertRaises(ValueError):
            _headers.detect_header_band(wb.active, region, -1)

    def test_bogus_str_rejected(self) -> None:
        wb = openpyxl.Workbook()
        region = _whole_sheet_region(wb.active)
        with self.assertRaises(ValueError):
            _headers.detect_header_band(wb.active, region, "bogus")  # type: ignore[arg-type]


class TestFlattenU203A(unittest.TestCase):
    """TC-E2E-02 / TC-UNIT-01: separator is U+203A."""

    def test_two_row_flatten(self) -> None:
        # Simulate post-merge-policy grid: anchor-only fill.
        rows = [
            ["Region", "2026 Plan", None, "Actual"],
            ["", "Q1", "Q2", "Total"],
        ]
        keys, warnings = _headers.flatten_headers(rows, 2)
        # Column 0: "Region" alone (level 1 empty). Column 1: "2026
        # Plan › Q1". Column 2: "2026 Plan › Q2" via sticky-fill-left
        # of "2026 Plan" into col 2. Column 3: "Actual › Total".
        self.assertEqual(keys[0], "Region")
        self.assertEqual(keys[1], "2026 Plan › Q1")
        self.assertEqual(keys[2], "2026 Plan › Q2")
        self.assertEqual(keys[3], "Actual › Total")
        # The U+203A separator MUST be in every multi-level key.
        for k in keys[1:]:
            self.assertIn("›", k)
        self.assertEqual(warnings, [])

    def test_separator_constant_is_u203a(self) -> None:
        self.assertEqual(_headers.HEADER_SEPARATOR, " › ")


class TestStickyFillLeft(unittest.TestCase):
    """TC-UNIT-02: sticky-fill-left propagates anchor labels rightward."""

    def test_horizontal_merge_anchor_propagates(self) -> None:
        rows = [
            ["A", None, None, "B"],
            ["x", "y", "z", "w"],
        ]
        keys, _ = _headers.flatten_headers(rows, 2)
        # Level 0: A › A › A › B (sticky-fill-left).
        # Level 1: x / y / z / w.
        self.assertEqual(keys, ["A › x", "A › y", "A › z", "B › w"])

    def test_consecutive_dedup(self) -> None:
        # When the same label appears at two levels, dedup → single value.
        rows = [
            ["X", "X"],
            ["X", "Y"],
        ]
        keys, _ = _headers.flatten_headers(rows, 2)
        # Column 0: dedup "X" → "X". Column 1: "X › Y" (different values).
        self.assertEqual(keys[0], "X")
        self.assertEqual(keys[1], "X › Y")


class TestMergeScopedStickyFill(unittest.TestCase):
    """TASK 010 §11 patch v2: sticky-fill must NOT propagate beyond the
    column-span of the originating horizontal merge. A title merge
    `A1:F1` covering 6 of 25 columns must produce 6 filled keys + 19
    empty keys, not 25 identical keys.
    """

    def test_title_merge_does_not_overflow_into_uncovered_cols(self) -> None:
        # 25 columns. A1:F1 is the title merge (cols 1-6). G1..Y1 are
        # untouched. After anchor-only policy: row[0] = [title, None×24].
        # Merge-scoped fill should produce [title, title, title, title,
        # title, title, "", "", ..., ""] — 6 fills, 19 empties.
        rows = [["TITLE"] + [None] * 24]
        merges = {(1, 1): (1, 6)}  # min_row=1, min_col=1 → max_row=1, max_col=6
        keys, _ = _headers.flatten_headers(
            rows, 1, merges=merges,
            region_top_row=1, region_left_col=1,
        )
        self.assertEqual(keys[:6], ["TITLE"] * 6)
        self.assertEqual(keys[6:], [""] * 19)

    def test_merge_scoped_fill_within_two_disjoint_banners(self) -> None:
        # Two banners on row 1: A1:B1 ("2026 Plan"), D1:E1 ("Actual").
        # Col C is between them (no merge). Row 2 has sub-labels.
        rows = [
            ["2026 Plan", None, "Mid", "Actual", None],
            ["Q1", "Q2", "X", "Q1", "Q2"],
        ]
        merges = {(1, 1): (1, 2), (1, 4): (1, 5)}
        keys, _ = _headers.flatten_headers(
            rows, 2, merges=merges,
            region_top_row=1, region_left_col=1,
        )
        # Col 0: "2026 Plan / Q1" — under banner.
        # Col 1: "2026 Plan / Q2" — under banner.
        # Col 2: "Mid / X" — not under banner; level-0 anchor cell value.
        # Col 3: "Actual / Q1" — under banner.
        # Col 4: "Actual / Q2" — under banner.
        self.assertEqual(keys[0], "2026 Plan › Q1")
        self.assertEqual(keys[1], "2026 Plan › Q2")
        self.assertEqual(keys[2], "Mid › X")
        self.assertEqual(keys[3], "Actual › Q1")
        self.assertEqual(keys[4], "Actual › Q2")

    def test_legacy_fallback_when_merges_none(self) -> None:
        """Calling flatten_headers without `merges` preserves the
        legacy unconditional sticky-fill (existing synthetic test
        fixtures pre-date this patch and don't model merge ranges).
        """
        rows = [["A", None, None, "B"], ["x", "y", "z", "w"]]
        keys_legacy, _ = _headers.flatten_headers(rows, 2)
        self.assertEqual(keys_legacy, ["A › x", "A › y", "A › z", "B › w"])

    def test_empty_merges_dict_triggers_strict_no_fill(self) -> None:
        """**vdd-multi-2 MED fix:** an explicitly empty `MergeMap`
        (`merges={}`) is a caller assertion of "no merges exist" —
        NOT the same as "merge info unknown" (`merges=None`).
        Strict no-fill semantics apply: `None` header cells stay
        empty, no inheritance from arbitrary left neighbours.

        Before the fix, `merges={}` collapsed to legacy unconditional
        sticky-fill, silently re-introducing the title-spillover bug
        for workbooks whose only merges sit in the body (not the
        header band). The new semantics: caller passed merge info =>
        we honour it, even when empty.
        """
        rows = [["A", None, None, "B"], ["x", "y", "z", "w"]]
        keys, _ = _headers.flatten_headers(rows, 2, merges={})
        # col 0: "A" / "x" → "A › x"
        # col 1: None / "y" → "y"  (no merge → no inheritance)
        # col 2: None / "z" → "z"
        # col 3: "B" / "w" → "B › w"
        self.assertEqual(keys, ["A › x", "y", "z", "B › w"])

    def test_body_only_merges_do_not_trigger_legacy_in_header(self) -> None:
        """**vdd-multi-2 MED fix:** when the workbook has merges but
        none anchor in the header band (e.g. totals-block banners in
        the body), `flatten_headers` must NOT silently fall back to
        legacy unconditional sticky-fill on the header rows. The
        body merges are real merge knowledge — strict mode applies.
        """
        rows = [["Region", None, "Country", None, None]]
        # A merge anchored at row 50 (body, below header band).
        merges = {(50, 1): (50, 3)}
        keys, _ = _headers.flatten_headers(
            rows, 1, merges=merges,
            region_top_row=1, region_left_col=1,
        )
        # Strict mode: None cells are empty, NOT inherited.
        self.assertEqual(keys, ["Region", "", "Country", "", ""])


class TestSyntheticHeaders(unittest.TestCase):
    """TC-UNIT-04: synthetic format unchanged from 009-01."""

    def test_zero_width(self) -> None:
        self.assertEqual(_headers.synthetic_headers(0), [])

    def test_five(self) -> None:
        self.assertEqual(
            _headers.synthetic_headers(5), ["col_1", "col_2", "col_3", "col_4", "col_5"]
        )

    def test_negative_rejected(self) -> None:
        with self.assertRaises(ValueError):
            _headers.synthetic_headers(-1)


class TestAmbiguousBoundary(unittest.TestCase):
    """TC-E2E-04: ambiguous-boundary check fires."""

    def test_straddling_merge_returns_warning(self) -> None:
        with open_workbook(FIXTURES_DIR / "headers_ambiguous.xlsx") as r:
            ws = r._wb.active
            region = _whole_sheet_region(ws)
            warning = _headers._ambiguous_boundary_check(
                list(ws.merged_cells.ranges), region, header_rows=1
            )
        self.assertIsNotNone(warning)
        self.assertIn("Ambiguous header boundary", warning)
        self.assertIn("B1:B2", warning)

    def test_non_straddling_merge_clean(self) -> None:
        # Two-row-merged fixture: the merge is INSIDE the header band
        # (rows 1-2), not straddling row1/row2 vs row3. Detect_header_band
        # returns 2, so the cut is between row 2 and row 3. No straddle.
        with open_workbook(FIXTURES_DIR / "headers_two_row_merged.xlsx") as r:
            ws = r._wb.active
            region = _whole_sheet_region(ws)
            warning = _headers._ambiguous_boundary_check(
                list(ws.merged_cells.ranges), region, header_rows=2
            )
        self.assertIsNone(warning)


class TestSyntheticUsedForZeroHeaders(unittest.TestCase):
    """TC-E2E-05: explicit hint=0 → synthetic emit via caller."""

    def test_band_zero_then_synthetic(self) -> None:
        with open_workbook(FIXTURES_DIR / "headers_single_row.xlsx") as r:
            ws = r._wb.active
            region = _whole_sheet_region(ws)
            band = _headers.detect_header_band(ws, region, hint=0)
            self.assertEqual(band, 0)
            # Caller will then call synthetic_headers(width=4) — verify
            # the standalone helper.
            width = region.right_col - region.left_col + 1
            keys = _headers.synthetic_headers(width)
            self.assertEqual(keys, [f"col_{i+1}" for i in range(width)])


if __name__ == "__main__":  # pragma: no cover
    unittest.main()
