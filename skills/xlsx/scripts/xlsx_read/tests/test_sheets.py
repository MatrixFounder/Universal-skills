"""Task 009-03 — F2 `_sheets.py` E2E + unit tests."""

from __future__ import annotations

import unittest

import openpyxl

from xlsx_read import SheetInfo, SheetNotFound, open_workbook
from xlsx_read import _sheets
from xlsx_read.tests.conftest import FIXTURES_DIR


class TestEnumerateThreeMixed(unittest.TestCase):
    """TC-E2E-01: enumerate visible+hidden+special-char names verbatim."""

    def test_three_sheetinfos_in_document_order(self) -> None:
        with open_workbook(FIXTURES_DIR / "three_sheets_mixed.xlsx") as reader:
            infos = reader.sheets()
        self.assertEqual(len(infos), 3)
        names = [i.name for i in infos]
        self.assertEqual(names, ["Visible1", "Hidden1", "Bob's Sheet"])

    def test_state_field_correct(self) -> None:
        with open_workbook(FIXTURES_DIR / "three_sheets_mixed.xlsx") as reader:
            states = {i.name: i.state for i in reader.sheets()}
        self.assertEqual(states["Visible1"], "visible")
        self.assertEqual(states["Hidden1"], "hidden")
        self.assertEqual(states["Bob's Sheet"], "visible")

    def test_indexes_are_dense_zero_based(self) -> None:
        with open_workbook(FIXTURES_DIR / "three_sheets_mixed.xlsx") as reader:
            indexes = [i.index for i in reader.sheets()]
        self.assertEqual(indexes, [0, 1, 2])

    def test_sheetinfos_are_immutable(self) -> None:
        with open_workbook(FIXTURES_DIR / "three_sheets_mixed.xlsx") as reader:
            info = reader.sheets()[0]
        with self.assertRaises(Exception):  # FrozenInstanceError
            info.name = "X"  # type: ignore[misc]


class TestResolveSheet(unittest.TestCase):
    """TC-E2E-02..-04: resolver NAME / all / missing."""

    def _open(self):
        return open_workbook(FIXTURES_DIR / "three_sheets_mixed.xlsx")

    def test_resolver_name_returns_one(self) -> None:
        with self._open() as r:
            self.assertEqual(_sheets.resolve_sheet(r._wb, "Hidden1"), "Hidden1")

    def test_resolver_special_char_name(self) -> None:
        with self._open() as r:
            self.assertEqual(_sheets.resolve_sheet(r._wb, "Bob's Sheet"), "Bob's Sheet")

    def test_resolver_all_returns_full_order(self) -> None:
        with self._open() as r:
            self.assertEqual(
                _sheets.resolve_sheet(r._wb, "all"),
                ["Visible1", "Hidden1", "Bob's Sheet"],
            )

    def test_resolver_missing_raises(self) -> None:
        with self._open() as r:
            with self.assertRaises(SheetNotFound) as ctx:
                _sheets.resolve_sheet(r._wb, "Nonexistent")
            self.assertIn("Nonexistent", str(ctx.exception))

    def test_resolver_case_sensitive(self) -> None:
        # "hidden1" (lowercase) is NOT a match for "Hidden1" — Excel
        # stores sheet names case-preserving; library matches verbatim.
        with self._open() as r:
            with self.assertRaises(SheetNotFound):
                _sheets.resolve_sheet(r._wb, "hidden1")


class TestStateMapper(unittest.TestCase):
    """TC-UNIT-01: `_state_from_openpyxl` covers each valid state + rejects invalid."""

    def test_visible(self) -> None:
        wb = openpyxl.Workbook()
        wb.active.sheet_state = "visible"
        self.assertEqual(_sheets._state_from_openpyxl(wb.active), "visible")

    def test_hidden(self) -> None:
        wb = openpyxl.Workbook()
        wb.active.sheet_state = "hidden"
        self.assertEqual(_sheets._state_from_openpyxl(wb.active), "hidden")

    def test_very_hidden(self) -> None:
        wb = openpyxl.Workbook()
        wb.active.sheet_state = "veryHidden"
        self.assertEqual(_sheets._state_from_openpyxl(wb.active), "veryHidden")

    def test_invalid_state_raises(self) -> None:
        class FakeWS:
            sheet_state = "bogus"

        with self.assertRaises(RuntimeError):
            _sheets._state_from_openpyxl(FakeWS())


class TestEmptyWorkbook(unittest.TestCase):
    """TC-UNIT-02: workbook with one default sheet → one SheetInfo, index=0."""

    def test_empty_returns_one_sheet(self) -> None:
        with open_workbook(FIXTURES_DIR / "empty.xlsx") as reader:
            infos = reader.sheets()
        self.assertEqual(len(infos), 1)
        self.assertEqual(infos[0].index, 0)
        self.assertEqual(infos[0].state, "visible")


class TestSpecialCharVerbatim(unittest.TestCase):
    """TC-UNIT-03: special-character sheet names preserved byte-for-byte."""

    def test_apostrophe_preserved(self) -> None:
        with open_workbook(FIXTURES_DIR / "three_sheets_mixed.xlsx") as reader:
            names = [i.name for i in reader.sheets()]
        self.assertIn("Bob's Sheet", names)


class TestSheetInfoReturnType(unittest.TestCase):
    """Regression: returned objects are `SheetInfo`, NOT openpyxl Worksheets (R1)."""

    def test_no_openpyxl_leak(self) -> None:
        with open_workbook(FIXTURES_DIR / "three_sheets_mixed.xlsx") as reader:
            for info in reader.sheets():
                self.assertIsInstance(info, SheetInfo)
                # The dataclass carries primitive attributes only.
                self.assertIsInstance(info.name, str)
                self.assertIsInstance(info.index, int)
                self.assertIsInstance(info.state, str)


if __name__ == "__main__":  # pragma: no cover
    unittest.main()
