"""Task 009-04 — F3 `_merges.py` E2E + unit tests."""

from __future__ import annotations

import copy
import unittest

import openpyxl

from types import SimpleNamespace

from xlsx_read import OverlappingMerges, TooManyMerges, open_workbook
from xlsx_read import _merges
from xlsx_read.tests.conftest import FIXTURES_DIR


# ---------------------------------------------------------------------------
# End-to-end — parse_merges across 3 fixture shapes
# ---------------------------------------------------------------------------


class TestParseMerges(unittest.TestCase):
    """TC-E2E-01..-03: parse the 3 merge shapes."""

    def _ws(self, name):
        wb = openpyxl.load_workbook(filename=str(FIXTURES_DIR / name))
        return wb.active

    def test_parse_row_merge(self) -> None:
        m = _merges.parse_merges(self._ws("merges_row.xlsx"))
        self.assertEqual(m, {(1, 1): (1, 3)})

    def test_parse_col_merge(self) -> None:
        m = _merges.parse_merges(self._ws("merges_col.xlsx"))
        self.assertEqual(m, {(1, 1): (3, 1)})

    def test_parse_rect_merge(self) -> None:
        m = _merges.parse_merges(self._ws("merges_rect.xlsx"))
        self.assertEqual(m, {(2, 2): (4, 4)})

    def test_parse_empty_sheet(self) -> None:
        wb = openpyxl.load_workbook(filename=str(FIXTURES_DIR / "empty.xlsx"))
        self.assertEqual(_merges.parse_merges(wb.active), {})


# ---------------------------------------------------------------------------
# End-to-end — three policies on three shapes (9-cell behavioural matrix)
# ---------------------------------------------------------------------------


class TestApplyPolicyAnchorOnly(unittest.TestCase):
    """TC-E2E-04: anchor-only across 3 fixtures."""

    def test_row_anchor_only(self) -> None:
        rows = [["ROW", "B1raw", "C1raw"]]
        out = _merges.apply_merge_policy(rows, {(1, 1): (1, 3)}, "anchor-only")
        self.assertEqual(out, [["ROW", None, None]])

    def test_col_anchor_only(self) -> None:
        rows = [["COL"], ["A2raw"], ["A3raw"]]
        out = _merges.apply_merge_policy(rows, {(1, 1): (3, 1)}, "anchor-only")
        self.assertEqual(out, [["COL"], [None], [None]])

    def test_rect_anchor_only(self) -> None:
        rows = [
            ["r1c1", "r1c2", "r1c3", "r1c4"],
            ["r2c1", "RECT", "r2c3", "r2c4"],
            ["r3c1", "r3c2", "r3c3", "r3c4"],
            ["r4c1", "r4c2", "r4c3", "r4c4"],
        ]
        out = _merges.apply_merge_policy(rows, {(2, 2): (4, 4)}, "anchor-only")
        self.assertEqual(out[0], ["r1c1", "r1c2", "r1c3", "r1c4"])
        self.assertEqual(out[1], ["r2c1", "RECT", None, None])
        self.assertEqual(out[2], ["r3c1", None, None, None])
        self.assertEqual(out[3], ["r4c1", None, None, None])


class TestApplyPolicyFill(unittest.TestCase):
    """TC-E2E-05: fill across 3 fixtures."""

    def test_row_fill(self) -> None:
        rows = [["ROW", "B1raw", "C1raw"]]
        out = _merges.apply_merge_policy(rows, {(1, 1): (1, 3)}, "fill")
        self.assertEqual(out, [["ROW", "ROW", "ROW"]])

    def test_col_fill(self) -> None:
        rows = [["COL"], ["A2raw"], ["A3raw"]]
        out = _merges.apply_merge_policy(rows, {(1, 1): (3, 1)}, "fill")
        self.assertEqual(out, [["COL"], ["COL"], ["COL"]])

    def test_rect_fill(self) -> None:
        rows = [
            ["r1c1", "r1c2", "r1c3", "r1c4"],
            ["r2c1", "RECT", "r2c3", "r2c4"],
            ["r3c1", "r3c2", "r3c3", "r3c4"],
            ["r4c1", "r4c2", "r4c3", "r4c4"],
        ]
        out = _merges.apply_merge_policy(rows, {(2, 2): (4, 4)}, "fill")
        self.assertEqual(out[1], ["r2c1", "RECT", "RECT", "RECT"])
        self.assertEqual(out[2], ["r3c1", "RECT", "RECT", "RECT"])
        self.assertEqual(out[3], ["r4c1", "RECT", "RECT", "RECT"])


class TestApplyPolicyBlank(unittest.TestCase):
    """TC-E2E-06: blank ≡ anchor-only in v1 (semantic alias)."""

    def test_blank_equals_anchor_only(self) -> None:
        rows = [["X", "y", "z"], ["a", "b", "c"]]
        merges = {(1, 1): (2, 3)}
        ao = _merges.apply_merge_policy(rows, merges, "anchor-only")
        bl = _merges.apply_merge_policy(rows, merges, "blank")
        self.assertEqual(ao, bl)


# ---------------------------------------------------------------------------
# E2E — overlap detector fail-loud
# ---------------------------------------------------------------------------


class TestOverlapDetector(unittest.TestCase):
    """TC-E2E-07: `_overlapping_merges_check` raises on the fixture."""

    def test_overlap_fixture_raises(self) -> None:
        wb = openpyxl.load_workbook(filename=str(FIXTURES_DIR / "overlapping_merges.xlsx"))
        with self.assertRaises(OverlappingMerges) as ctx:
            _merges._overlapping_merges_check(wb.active.merged_cells.ranges)
        msg = str(ctx.exception)
        # The two injected ranges must both appear in the message.
        self.assertIn("A1:B2", msg)
        self.assertIn("B2:C3", msg)


# ---------------------------------------------------------------------------
# Unit tests
# ---------------------------------------------------------------------------


class TestPolicyPurity(unittest.TestCase):
    """TC-UNIT-01: `apply_merge_policy` never mutates input."""

    def test_input_unchanged_anchor_only(self) -> None:
        original = [["a", "b"], ["c", "d"]]
        snapshot = copy.deepcopy(original)
        _merges.apply_merge_policy(original, {(1, 1): (2, 2)}, "anchor-only")
        self.assertEqual(original, snapshot)

    def test_input_unchanged_fill(self) -> None:
        original = [["a", "b"], ["c", "d"]]
        snapshot = copy.deepcopy(original)
        _merges.apply_merge_policy(original, {(1, 1): (2, 2)}, "fill")
        self.assertEqual(original, snapshot)


class TestEmptyMergesPassthrough(unittest.TestCase):
    """TC-UNIT-02: empty merges → grid shape preserved."""

    def test_passthrough_anchor_only(self) -> None:
        rows = [[1, 2, 3], [4, 5, 6]]
        out = _merges.apply_merge_policy(rows, {}, "anchor-only")
        self.assertEqual(out, rows)
        self.assertIsNot(out, rows)  # but a *new* list (purity)


class TestOverlapDetectorNoFalsePositive(unittest.TestCase):
    """TC-UNIT-03: adjacent non-overlapping ranges are clean."""

    def test_adjacent_vertical(self) -> None:
        wb = openpyxl.Workbook()
        wb.active.merge_cells("A1:A3")
        wb.active.merge_cells("A4:A6")
        # No raise expected.
        self.assertIsNone(_merges._overlapping_merges_check(wb.active.merged_cells.ranges))

    def test_adjacent_horizontal(self) -> None:
        wb = openpyxl.Workbook()
        wb.active.merge_cells("A1:C1")
        wb.active.merge_cells("D1:F1")
        self.assertIsNone(_merges._overlapping_merges_check(wb.active.merged_cells.ranges))


class TestOverlapDetectorCornerOverlap(unittest.TestCase):
    """TC-UNIT-04: single-cell corner overlap raises.

    Note: openpyxl's `Workbook.merge_cells()` itself REJECTS overlapping
    merges with `MergedCellRangeException` at build time. The fixture
    `overlapping_merges.xlsx` was assembled by post-save XML patching
    precisely to test this case. So this unit test exercises the
    detector against an in-memory raw-range list rather than going
    through `merge_cells()`.
    """

    def test_single_cell_corner_raises(self) -> None:
        # The detector duck-types on `.min_row / .min_col / .max_row /
        # .max_col`; we don't need a real openpyxl `MergedCellRange`
        # here. (Going through `Workbook.merge_cells()` is impossible
        # anyway — openpyxl rejects overlaps at build time.)
        class _R:
            def __init__(self, s, mr, mc, Mr, Mc):
                self._s = s
                self.min_row, self.min_col = mr, mc
                self.max_row, self.max_col = Mr, Mc

            def __str__(self):
                return self._s

        r1 = _R("A1:B2", 1, 1, 2, 2)
        r2 = _R("B2:C3", 2, 2, 3, 3)
        with self.assertRaises(OverlappingMerges):
            _merges._overlapping_merges_check([r1, r2])


# ---------------------------------------------------------------------------
# Defensive — invalid policy
# ---------------------------------------------------------------------------


class TestInvalidPolicy(unittest.TestCase):
    """`apply_merge_policy("nonsense")` raises ValueError."""

    def test_invalid_policy_raises(self) -> None:
        with self.assertRaises(ValueError):
            _merges.apply_merge_policy([[1]], {}, "nonsense")  # type: ignore[arg-type]


class TestOverlapMemoSoundnessAcrossRegions(unittest.TestCase):
    """NEW-S-M1 regression (security iter-2): the overlap-check memo
    must NOT silently skip a second region whose overlapping merges
    sit outside the first-checked region's bbox.

    Fixture has clean region A (A1:C3) and overlapping merges (E10:F11
    ∩ F11:G12) inside region B. Iter-2 introduced a bug where reading
    region A first marked the sheet as checked, allowing region B's
    read to silently corrupt the merge-policy output. Iter-3 fixes by
    running the overlap check on the WHOLE sheet on first read.
    """

    def test_region_B_overlap_detected_after_region_A_clean_read(self) -> None:
        from xlsx_read import OverlappingMerges, TableRegion, open_workbook
        from xlsx_read.tests.conftest import FIXTURES_DIR
        path = FIXTURES_DIR / "overlap_in_region_B_only.xlsx"
        with open_workbook(path) as r:
            ws = r._wb.active
            region_a = TableRegion(
                sheet=ws.title, top_row=1, left_col=1,
                bottom_row=3, right_col=3, source="gap_detect",
            )
            region_b = TableRegion(
                sheet=ws.title, top_row=10, left_col=5,
                bottom_row=20, right_col=8, source="gap_detect",
            )
            # FIRST read on region A must raise — even though A
            # itself contains no overlapping merges, the sheet does.
            # Soundness requires the check be sheet-wide.
            with self.assertRaises(OverlappingMerges):
                r.read_table(region_a, header_rows=0)


# ---------------------------------------------------------------------------
# xlsx-8a-02 (R2, Sec-MED-3) — `_MAX_MERGES` cap regression
# ---------------------------------------------------------------------------


def _mock_ws_with_n_merges(n: int, title: str = "MockSheet") -> Any:
    """Build a `SimpleNamespace` worksheet whose `merged_cells.ranges`
    yields `n` mock-range objects (each is a 1×1 single-cell merge at
    `(row=1, col=k)` for `k in range(1, n+1)`).

    Used by `_MAX_MERGES` boundary tests — avoids the 5-10s cost of
    building a real openpyxl Workbook with that many merges, and
    pinpoints the cap-guard behaviour deterministically.
    """
    ranges = [
        SimpleNamespace(min_row=1, max_row=1, min_col=k, max_col=k)
        for k in range(1, n + 1)
    ]
    return SimpleNamespace(
        title=title,
        merged_cells=SimpleNamespace(ranges=ranges),
    )


class TestParseMergesCap(unittest.TestCase):
    """xlsx-8a-02 (R2): bounded merge-count in `parse_merges(ws)`."""

    def test_R2_parse_merges_at_100000_passes(self) -> None:
        """Cap allows exactly 100_000 merge entries; no raise."""
        ws = _mock_ws_with_n_merges(100_000)
        out = _merges.parse_merges(ws)
        self.assertEqual(len(out), 100_000)

    def test_R2_parse_merges_at_100001_raises(self) -> None:
        """The 100_001st merge insertion raises `TooManyMerges`."""
        ws = _mock_ws_with_n_merges(100_001)
        with self.assertRaises(TooManyMerges) as ctx:
            _merges.parse_merges(ws)
        # Basename-safe message (sheet name is workbook-controlled
        # but `_validate_sheet_path_components` is downstream of the
        # library — here we only check the message carries the
        # title for diagnostic purposes, no absolute paths).
        self.assertIn("MockSheet", str(ctx.exception))
        self.assertIn("100000", str(ctx.exception))


if __name__ == "__main__":  # pragma: no cover
    unittest.main()
