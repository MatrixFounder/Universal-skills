"""Task 009-07 — F6 `_values.py` E2E + unit tests."""

from __future__ import annotations

import unittest
from datetime import date, datetime

import openpyxl

from xlsx_read import open_workbook
from xlsx_read import _values
from xlsx_read.tests.conftest import FIXTURES_DIR


class TestNumberFormatHeuristic(unittest.TestCase):
    """TC-E2E-01..-06 (number-format) + TC-UNIT-01/-02 (regex coverage)."""

    def _wb(self):
        return openpyxl.load_workbook(filename=str(FIXTURES_DIR / "values_numformat.xlsx"))

    def test_decimal_with_thousands(self) -> None:
        wb = self._wb()
        v, w = _values.extract_cell(wb.active["A1"])
        self.assertEqual(v, "1,234.50")
        self.assertIsNone(w)

    def test_percent_1dp(self) -> None:
        wb = self._wb()
        v, w = _values.extract_cell(wb.active["A2"])
        self.assertEqual(v, "5.0%")

    def test_percent_2dp(self) -> None:
        wb = self._wb()
        v, _ = _values.extract_cell(wb.active["A6"])
        self.assertEqual(v, "12.30%")

    def test_date_iso(self) -> None:
        wb = self._wb()
        v, _ = _values.extract_cell(wb.active["A3"], datetime_format="ISO")
        # ISO of datetime → "2026-03-05T00:00:00" (string).
        self.assertIsInstance(v, str)
        self.assertTrue(v.startswith("2026-03-05"))

    def test_date_excel_serial(self) -> None:
        wb = self._wb()
        v, _ = _values.extract_cell(wb.active["A3"], datetime_format="excel-serial")
        self.assertIsInstance(v, float)
        # 2026-03-05 from epoch 1899-12-30 = 46086 days (true Python
        # day-count). Honest-scope note: Excel's own serial is +1
        # (=46087) due to the 1900-02-29 leap-year bug; the v1 library
        # surfaces TRUE calendar days. Excel-bug-compat is out of
        # scope for v1 (document; future v2 flag if needed).
        self.assertAlmostEqual(v, 46086.0, places=2)

    def test_date_raw(self) -> None:
        wb = self._wb()
        v, _ = _values.extract_cell(wb.active["A3"], datetime_format="raw")
        self.assertIsInstance(v, (datetime, date))
        self.assertEqual((v.year, v.month, v.day), (2026, 3, 5))

    def test_leading_zero_text(self) -> None:
        wb = self._wb()
        v, _ = _values.extract_cell(wb.active["A4"])
        self.assertEqual(v, "00042")

    def test_general_passthrough(self) -> None:
        wb = self._wb()
        v, _ = _values.extract_cell(wb.active["A5"])
        self.assertEqual(v, 42)

    def test_decimal_no_decimals(self) -> None:
        wb = self._wb()
        v, _ = _values.extract_cell(wb.active["A7"])
        # `#,##0` → "1,000" (no decimals).
        self.assertEqual(v, "1,000")


class TestFormulaPath(unittest.TestCase):
    """TC-E2E-07..-09: formula cached / emitted / stale-cache.

    Honest-scope (xlsx-10.A v1): openpyxl exposes EITHER cached values
    OR formula strings, not both. The `keep_formulas` flag at
    `open_workbook` time picks the side.
    """

    def test_cached_default(self) -> None:
        # Default open (keep_formulas=False, openpyxl data_only=True):
        # cell.value is the cached <v>7</v> value injected at fixture
        # build time. include_formulas is irrelevant in this mode.
        with open_workbook(FIXTURES_DIR / "values_formula_cached.xlsx") as r:
            cell = r._wb.active["C1"]
            v, w = _values.extract_cell(cell)
        self.assertEqual(v, 7)
        self.assertIsNone(w)

    def test_formula_emitted_when_opted_in(self) -> None:
        # keep_formulas=True open → openpyxl data_only=False →
        # cell.value is the formula string; include_formulas=True
        # passes it through verbatim.
        with open_workbook(
            FIXTURES_DIR / "values_formula_cached.xlsx", keep_formulas=True
        ) as r:
            cell = r._wb.active["C1"]
            v, w = _values.extract_cell(cell, include_formulas=True)
        self.assertEqual(v, "=A1+B1")
        self.assertIsNone(w)

    def test_stale_cache_warns_when_cached_value_requested_in_keep_mode(self) -> None:
        # keep_formulas=True + include_formulas=False: caller asked
        # for a cached value but the loader doesn't have it. Warning
        # surfaces with the documented "reopen with keep_formulas=
        # False" hint.
        with open_workbook(
            FIXTURES_DIR / "values_formula_cached.xlsx", keep_formulas=True
        ) as r:
            cell = r._wb.active["C1"]
            v, w = _values.extract_cell(cell, include_formulas=False)
        self.assertIsNone(v)
        self.assertIsNotNone(w)
        self.assertIn("stale cache", w)


class TestHyperlink(unittest.TestCase):
    """TC-E2E-10..-11: hyperlink off / on."""

    def test_off_returns_display(self) -> None:
        wb = openpyxl.load_workbook(filename=str(FIXTURES_DIR / "values_hyperlink.xlsx"))
        v, _ = _values.extract_cell(wb.active["A1"], include_hyperlinks=False)
        self.assertEqual(v, "Click here")

    def test_on_returns_target(self) -> None:
        wb = openpyxl.load_workbook(filename=str(FIXTURES_DIR / "values_hyperlink.xlsx"))
        v, _ = _values.extract_cell(wb.active["A1"], include_hyperlinks=True)
        self.assertEqual(v, "https://example.com")


class TestRichText(unittest.TestCase):
    """TC-E2E-12: rich-text spans flattened to plain text."""

    def test_richtext_concat(self) -> None:
        wb = openpyxl.load_workbook(
            filename=str(FIXTURES_DIR / "values_richtext.xlsx"),
            rich_text=True,
        )
        v, _ = _values.extract_cell(wb.active["A1"])
        # Bold span + plain span → "Bold then plain".
        self.assertEqual(v, "Bold then plain")


class TestExcelEpochRoundtrip(unittest.TestCase):
    """TC-UNIT-03: Excel epoch anchor (1899-12-30, not 1900-01-01)."""

    def test_epoch_anchor(self) -> None:
        self.assertEqual(_values._EXCEL_EPOCH, datetime(1899, 12, 30))

    def test_jan1_1900_is_serial_2(self) -> None:
        v = _values._apply_datetime_format(datetime(1900, 1, 1), "excel-serial")
        self.assertEqual(v, 2.0)

    def test_epoch_is_serial_0(self) -> None:
        v = _values._apply_datetime_format(datetime(1899, 12, 30), "excel-serial")
        self.assertEqual(v, 0.0)


class TestFlattenRichTextPassthrough(unittest.TestCase):
    """TC-UNIT-05: non-rich-text inputs return verbatim."""

    def test_none_passthrough(self) -> None:
        self.assertIsNone(_values._flatten_rich_text(None))

    def test_string_passthrough(self) -> None:
        self.assertEqual(_values._flatten_rich_text("plain"), "plain")

    def test_int_passthrough(self) -> None:
        self.assertEqual(_values._flatten_rich_text(42), 42)


class TestStaleCacheGuards(unittest.TestCase):
    """TC-UNIT-06: stale-cache only fires on data_type=='f' cells."""

    def test_non_formula_with_none_value(self) -> None:
        class FakeCell:
            data_type = "n"
            value = None
            number_format = "General"
            coordinate = "A1"
            hyperlink = None

        v, w = _values.extract_cell(FakeCell())
        self.assertIsNone(v)
        self.assertIsNone(w)


class TestInvalidDateFormat(unittest.TestCase):
    def test_bogus_datetime_format_raises(self) -> None:
        with self.assertRaises(ValueError):
            _values._apply_datetime_format(datetime(2026, 1, 1), "bogus")  # type: ignore[arg-type]


if __name__ == "__main__":  # pragma: no cover
    unittest.main()
