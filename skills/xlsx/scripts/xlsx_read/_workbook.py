"""F1 — workbook open + cross-cutting pre-flight.

`open_workbook(path)` runs the cross-cutting probes (encryption,
macros) **before** invoking openpyxl, picks a `read_only` mode based
on file size, opens the workbook, and returns a `WorkbookReader`.

Cross-cutting contracts honoured:
- **cross-3** (encryption): `_probe_encryption` raises
  `EncryptedWorkbookError` if the OPC archive carries an
  `EncryptedPackage` part. Caller maps to exit 3.
- **cross-4** (macros): `_probe_macros` returns True iff
  `xl/vbaProject.bin` is present; `open_workbook` then emits a
  `MacroEnabledWarning` via `warnings.warn`. The library NEVER
  raises on macros (UC-01 A3).

Encryption detection re-implements the same heuristic as
`office_passwd.py` (CFB-magic-bytes prefix + OPC `EncryptedPackage`
stream), **without** importing the cross-skill replicated module —
CLAUDE.md §2 forbids reading from `office_passwd.py` here (and the
heuristic is only a few lines). The detection covers Office 2010+
agile encryption and the legacy CFB compound-file form.
"""

from __future__ import annotations

import warnings
import zipfile
from pathlib import Path
from typing import TYPE_CHECKING

import openpyxl

from ._exceptions import EncryptedWorkbookError, MacroEnabledWarning
from ._types import WorkbookReader

if TYPE_CHECKING:
    pass

# CFB (Compound File Binary) header magic — every legacy MS-Office
# encrypted container starts with this 8-byte sequence. Detecting it
# pre-empts openpyxl, which would otherwise raise an obscure
# `BadZipFile` because CFB containers are not ZIPs.
_CFB_MAGIC: bytes = b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"

# OPC parts present in any agile-encrypted .xlsx (Office 2010+). The
# DataSpaceMap part is the most reliable signal; EncryptedPackage is
# the actual ciphertext payload. Either is sufficient.
_OPC_ENCRYPTED_NAMES: frozenset[str] = frozenset({
    "EncryptedPackage",
    "\x06DataSpaces/DataSpaceMap",
})

# OPC part that flags macros. Present in `.xlsm` and `.xltm`; absent
# from `.xlsx`. Probed purely by name — we never read or execute it.
_OPC_MACRO_PART: str = "xl/vbaProject.bin"

# Default size threshold above which `read_only=True` is preferred —
# openpyxl streams rows one at a time in that mode, capping memory.
# 10 MiB matches the xlsx-7 precedent (verified in
# `xlsx_check_rules/scope_resolver.py`).
_DEFAULT_READ_ONLY_THRESHOLD: int = 10 * 1024 * 1024


def _probe_encryption(path: Path) -> None:
    """Raise `EncryptedWorkbookError` iff `path` looks encrypted.

    Two complementary checks:
    1. Legacy CFB compound-file: first 8 bytes match `_CFB_MAGIC`.
       openpyxl cannot read those at all; bail before it tries.
    2. Agile-encrypted OPC: ZIP-archive contains the `EncryptedPackage`
       or `DataSpaceMap` stream. openpyxl would raise a generic
       `InvalidFileException` here; we replace that with the typed
       error so callers can route exit 3 without string-matching.

    **S-M2 fix:** error messages use `path.name` (basename only) — the
    full resolved path is leakable in shared logs / CI surfaces. The
    caller has the full path in scope already; we don't need to echo
    it back. The path is still bound to `__cause__` via Python's
    exception chaining if openpyxl raises downstream.
    """
    # CFB magic — short read, then bail.
    with path.open("rb") as fh:
        head = fh.read(len(_CFB_MAGIC))
    if head == _CFB_MAGIC:
        raise EncryptedWorkbookError(
            f"Workbook is encrypted (CFB compound file): {path.name}"
        )

    # OPC agile encryption — peek at the archive namelist.
    try:
        with zipfile.ZipFile(path, "r") as zf:
            namelist = set(zf.namelist())
    except zipfile.BadZipFile:
        # Not a ZIP at all; not our concern — propagate to caller via
        # the regular `openpyxl.load_workbook` failure path.
        return
    if namelist & _OPC_ENCRYPTED_NAMES:
        raise EncryptedWorkbookError(
            f"Workbook is encrypted (OPC agile): {path.name}"
        )


def _probe_macros(path: Path) -> bool:
    """Return True iff the workbook carries `xl/vbaProject.bin`.

    Pure name probe — never reads the part contents, never executes
    anything. False on non-ZIP inputs (a BadZipFile is openpyxl's
    problem, not ours).
    """
    try:
        with zipfile.ZipFile(path, "r") as zf:
            return _OPC_MACRO_PART in zf.namelist()
    except zipfile.BadZipFile:
        return False


def _decide_read_only(
    path: Path,
    override: bool | None,
    threshold: int,
) -> bool:
    """Pick `read_only` mode: caller override wins, else size > threshold."""
    if override is not None:
        return override
    return path.stat().st_size > threshold


def open_workbook(
    path: Path,
    *,
    read_only_mode: bool | None = None,
    size_threshold_bytes: int = _DEFAULT_READ_ONLY_THRESHOLD,
    keep_formulas: bool = False,
) -> WorkbookReader:
    """Open an `.xlsx` / `.xlsm` workbook for read.

    Pipeline:
    1. Resolve `path` strictly (raises `FileNotFoundError` on missing).
    2. `_probe_encryption` — fail-fast on encrypted inputs (cross-3).
    3. `_probe_macros` — emit `MacroEnabledWarning` (cross-4); never
       raise.
    4. `_decide_read_only` — pick streaming mode for large files.
    5. `openpyxl.load_workbook(..., data_only=<inverse-of-keep_formulas>,
       keep_links=False)`. **Honest-scope (xlsx-10.A v1):** openpyxl
       cannot expose **both** formula strings AND cached values from a
       single load. The library picks a side at open time:
         - default `keep_formulas=False` → `data_only=True` → cached
           values are surfaced (matches the common case where xlsx-8 /
           xlsx-9 emit serialised data).
         - `keep_formulas=True` → `data_only=False` → formula strings
           are surfaced; `cell.value` returns `"=A1+B1"` etc. Callers
           that need both must open the workbook twice (once with
           each mode) — the cost is documented and intentional.
    6. Wrap in `WorkbookReader` with the resolved flags.

    Raises:
        FileNotFoundError: when `path` does not exist.
        EncryptedWorkbookError: from `_probe_encryption`.
        zipfile.BadZipFile / openpyxl.utils.exceptions.InvalidFileException:
            propagated unchanged for corrupted-but-not-encrypted files.
    """
    resolved = Path(path).resolve(strict=True)
    _probe_encryption(resolved)
    # **P-H3 fix:** macros are only legal in `.xlsm`/`.xltm` containers.
    # Skip the zip-open probe on `.xlsx` (the common case) — drops ~10–
    # 50 ms per open for files where the answer is structurally
    # impossible. The `.xlsm` extension itself is the loudest signal.
    if resolved.suffix.lower() in (".xlsm", ".xltm") and _probe_macros(resolved):
        warnings.warn(
            # **S-M2 fix:** echo basename only — full resolved path is
            # a side-channel in shared logs (multi-tenant CI, etc.).
            f"Workbook carries xl/vbaProject.bin — macros are present "
            f"and IGNORED: {resolved.name}",
            category=MacroEnabledWarning,
            stacklevel=2,
        )
    read_only = _decide_read_only(resolved, read_only_mode, size_threshold_bytes)
    wb = openpyxl.load_workbook(
        filename=str(resolved),
        read_only=read_only,
        data_only=not keep_formulas,
        keep_links=False,
    )
    return WorkbookReader(
        path=resolved,
        _wb=wb,
        _read_only=read_only,
        _keep_formulas=keep_formulas,
    )
