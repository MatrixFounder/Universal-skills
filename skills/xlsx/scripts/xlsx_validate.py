#!/usr/bin/env python3
"""Scan an .xlsx for formula errors without recalculating.

Pairs with `xlsx_recalc.py`: first recalc (LibreOffice), then validate
(this script, fast, no external tools). If you validate without a
recalc pass, a freshly-openpyxl-written file will typically report
zero errors because formulas are stored as strings with no cached
value.

Usage:
    python3 xlsx_validate.py input.xlsx [--json] [--fail-empty]

Exit codes:
    0 — no errors found
    1 — errors found (any of #REF!, #DIV/0!, #VALUE!, #NAME?, #N/A,
        #NUM!, #NULL!)
    2 — input missing
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

from openpyxl import load_workbook  # type: ignore

from office._encryption import EncryptedFileError, assert_not_encrypted


ERROR_TOKENS = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NUM!", "#NULL!"}


def scan(input_path: Path) -> tuple[dict[str, list[str]], int]:
    """Scan only cells whose Excel data_type is 'e' (error).

    This distinguishes a cached formula-error value from a string cell
    that happens to contain the literal text "#REF!" — which openpyxl
    would expose with the same .value string but with data_type 's'.
    """
    wb = load_workbook(str(input_path), data_only=True, read_only=True)
    hits: dict[str, list[str]] = {}
    non_empty_cells = 0
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    non_empty_cells += 1
                    if cell.data_type == "e" and isinstance(cell.value, str) and cell.value in ERROR_TOKENS:
                        hits.setdefault(cell.value, []).append(f"{sheet_name}!{cell.coordinate}")
    finally:
        wb.close()
    return hits, non_empty_cells


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("input", type=Path, help="Source .xlsx")
    parser.add_argument("--json", action="store_true", help="Emit JSON report")
    parser.add_argument(
        "--fail-empty",
        action="store_true",
        help="Exit 1 if every cell is None — likely means formulas were never recalculated",
    )
    args = parser.parse_args(argv)

    if not args.input.is_file():
        print(f"Input not found: {args.input}", file=sys.stderr)
        return 2

    try:
        assert_not_encrypted(args.input)
    except EncryptedFileError as exc:
        print(str(exc), file=sys.stderr)
        return 3

    hits, non_empty = scan(args.input)
    if args.json:
        print(json.dumps({"ok": not hits, "errors": hits, "non_empty_cells": non_empty},
                         ensure_ascii=False, indent=2))
    else:
        if not hits:
            print(f"OK — {non_empty} non-empty cells, no formula errors.")
        else:
            total = sum(len(v) for v in hits.values())
            print(f"{total} formula-error cells across {len(hits)} error types:")
            for code, cells in hits.items():
                print(f"  {code}: {len(cells)} — {', '.join(cells[:10])}"
                      + ("..." if len(cells) > 10 else ""))

    if hits:
        return 1
    if args.fail_empty and non_empty == 0:
        print("Workbook has no cached values — run xlsx_recalc.py first.", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
