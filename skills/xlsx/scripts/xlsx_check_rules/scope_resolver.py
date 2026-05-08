"""F6 — Scope resolver: 10 SPEC §4 scope forms → concrete cell sequences.

Owns sheet-qualifier parsing (plain / quoted / apostrophe-escaped per
ECMA-376), header lookup with case-sensitive whitespace-strip,
Excel-Tables auto-detect via openpyxl's `ws.tables` API, merged-cell
anchor resolution, and the `--visible-only` filter.

Security note: Excel-Tables metadata is consumed via openpyxl's
parser. A separately-hardened lxml read of `xl/tables/tableN.xml`
(mirroring xlsx-6's VML parser) was considered and deferred —
`office/unpack` enforces the zip / path-traversal boundary, and a v2
hardening pass can layer defusedxml on top if a real vector emerges.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from typing import Any, Iterator

from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import coordinate_from_string, range_boundaries

from .ast_nodes import (
    CellRef, ColRef, MultiColRef, NamedRef, RangeRef, RowRef,
    SheetRef, TableRef,
)
from .cell_types import ClassifiedCell, classify
from .exceptions import (
    AmbiguousHeader, CorruptInput, HeaderNotFound,
    MergedHeaderUnsupported, RulesParseError,
)

__all__ = [
    "ScopeResult",
    "parse_sheet_qualifier",
    "resolve_sheet",
    "resolve_header",
    "iter_cells",
    "resolve_named",
    "resolve_scope",
]

# Excel-forbidden chars in unquoted sheet names (ECMA-376 §18.3.1.81).
_FORBIDDEN_SHEET_CHARS = set(r"\/?*[]:")


# === Public types =========================================================

@dataclass
class ScopeResult:
    sheet_name: str
    cells: list[ClassifiedCell]
    column_letter: str | None = None    # populated for col:HEADER / col:LETTER
    is_table_resolved: bool = False     # True iff Excel-Tables fallback fired
    info_findings: list[dict[str, Any]] = field(default_factory=list)  # merged-cell-resolution etc.


# === Sheet qualifier ======================================================

def parse_sheet_qualifier(text: str) -> tuple[str | None, str]:
    """Split sheet qualifier per ECMA-376 §18.3.1.81: plain, quoted, or
    apostrophe-escaped. Returns ``(sheet | None, ref)``."""
    s = text.strip()
    if not s:
        raise RulesParseError("empty scope text", subtype="BadGrammar")

    # Quoted form: starts with apostrophe; '' is the escape.
    if s.startswith("'"):
        i = 1
        end = -1
        while i < len(s):
            if s[i] == "'":
                if i + 1 < len(s) and s[i + 1] == "'":
                    i += 2  # escaped apostrophe; advance past pair
                    continue
                end = i
                break
            i += 1
        if end == -1:
            raise RulesParseError(f"unterminated quoted sheet name: {text!r}",
                                    subtype="BadGrammar")
        if end + 1 >= len(s) or s[end + 1] != "!":
            raise RulesParseError(f"missing '!' after quoted sheet name: {text!r}",
                                    subtype="BadGrammar")
        sheet = s[1:end].replace("''", "'")
        return sheet, s[end + 2:]

    # Plain form must contain '!' to be qualified; otherwise unqualified.
    if "!" not in s:
        return None, s

    sheet, _, ref = s.partition("!")
    if not sheet:
        raise RulesParseError(f"empty sheet name: {text!r}", subtype="BadGrammar")
    bad = sorted({ch for ch in sheet if ch in _FORBIDDEN_SHEET_CHARS})
    if bad:
        raise RulesParseError(
            f"sheet name contains prohibited char(s) {bad}: {text!r}",
            subtype="BadGrammar",
        )
    if sheet.startswith("'") or sheet.endswith("'"):
        raise RulesParseError(
            f"unquoted sheet name has leading/trailing apostrophe: {text!r}",
            subtype="BadGrammar",
        )
    return sheet, ref


# === Sheet resolution =====================================================

def resolve_sheet(qualifier: str | None, workbook: Any) -> Any:
    """SPEC §4.1 — None → first non-hidden sheet in workbook-XML order."""
    if qualifier is None:
        for name in workbook.sheetnames:
            ws = workbook[name]
            if getattr(ws, "sheet_state", "visible") not in ("hidden", "veryHidden"):
                return ws
        raise RulesParseError(
            "no visible sheet in workbook (all hidden / veryHidden); "
            "qualify the scope with an explicit sheet name",
            subtype="BadGrammar",
        )
    if qualifier not in workbook.sheetnames:
        raise RulesParseError(
            f"sheet not found: {qualifier!r}",
            subtype="BadGrammar", available=list(workbook.sheetnames),
        )
    return workbook[qualifier]


# === Header lookup ========================================================

def _header_row_has_merged_cells(ws: Any, header_row: int) -> bool:
    """SPEC §4.2 — header row with `<mergeCell>` ranges is rejected."""
    for merge in ws.merged_cells.ranges:
        # range_boundaries returns (min_col, min_row, max_col, max_row)
        _, min_row, _, max_row = merge.min_col, merge.min_row, merge.max_col, merge.max_row
        if min_row <= header_row <= max_row:
            return True
    return False


def _excel_tables_for_sheet(ws: Any) -> dict[str, Any]:
    """`{table_name: Table}` via openpyxl's `ws.tables`."""
    return dict(getattr(ws, "tables", {}))


def _table_column_lookup(name: str, ws: Any) -> str | None:
    """Return the column letter for `name` from any Excel Table on the sheet,
    else None. Case-sensitive, whitespace-stripped (SPEC §4.3)."""
    for table in _excel_tables_for_sheet(ws).values():
        for idx, tc in enumerate(getattr(table, "tableColumns", None) or []):
            if (getattr(tc, "name", "") or "").strip() == name:
                min_col, _, _, _ = range_boundaries(table.ref)
                return get_column_letter(min_col + idx)
    return None


def resolve_header(name: str, ws: Any, defaults: dict[str, Any] | None,
                   allow_table_fallback: bool = True) -> tuple[str, bool]:
    """SPEC §4.2 + §4.3 — return `(column_letter, is_table_resolved)`.

    Resolution order (per SPEC §4.3 "Table's header takes precedence"):
      1. If `allow_table_fallback`, consult Excel Tables first.
      2. Else fall through to the cell-grid `header_row` lookup.

    Raises:
        AmbiguousHeader: two columns share the header in the cell grid.
        HeaderNotFound:  no column has the header (after Table consult).
        MergedHeaderUnsupported: header row contains `<mergeCell>` ranges.
    """
    header_row = int((defaults or {}).get("header_row", 1))
    if header_row > 0 and _header_row_has_merged_cells(ws, header_row):
        raise MergedHeaderUnsupported(
            f"header row {header_row} on sheet {ws.title!r} contains merged cells",
            sheet=ws.title, header_row=header_row,
        )

    # SPEC §4.3 — Table headers take precedence over cell-grid header_row.
    if allow_table_fallback:
        table_col = _table_column_lookup(name, ws)
        if table_col is not None:
            return table_col, True

    # Cell-grid lookup (case-sensitive, whitespace-stripped).
    matches: list[str] = []
    available: list[str] = []
    if header_row > 0:
        for cell in ws[header_row]:
            v = cell.value
            if v is None:
                continue
            text = str(v).strip()
            if text:
                available.append(text)
            if text == name:
                matches.append(cell.column_letter)

    if len(matches) > 1:
        raise AmbiguousHeader(
            f"header {name!r} appears in columns {matches} on sheet {ws.title!r}",
            sheet=ws.title, columns=matches,
        )
    if len(matches) == 1:
        return matches[0], False

    raise HeaderNotFound(
        f"header {name!r} not found on sheet {ws.title!r}",
        sheet=ws.title, header=name,
        available=available[:50],  # truncate to first 50 per SPEC §4.2
    )


# === Cell iteration =======================================================

def iter_cells(scope_result: ScopeResult, opts: dict[str, Any] | None = None
                ) -> Iterator[ClassifiedCell]:
    """Apply ``--visible-only`` (opts['visible_only']) filter."""
    visible_only = bool((opts or {}).get("visible_only", False))
    for c in scope_result.cells:
        if visible_only and c.is_hidden:
            continue
        yield c


# === Named-range resolution ===============================================

def _split_named_areas(value: str) -> list[str]:
    # Bare-comma split is safe: $-prefixed refs and apostrophe-escaped
    # sheet names never contain commas.
    return [s.strip() for s in value.split(",") if s.strip()]


def resolve_named(name: str, workbook: Any) -> str:
    """Resolve a `definedName` to a single area ref. Multi-area names
    rejected (R13.i / SPEC §4 honest scope)."""
    dn = workbook.defined_names.get(name)
    if dn is None:
        raise RulesParseError(
            f"defined name not found: {name!r}",
            subtype="BadGrammar",
        )
    value = getattr(dn, "value", None) or getattr(dn, "attr_text", "") or ""
    areas = _split_named_areas(value)
    if len(areas) > 1:
        raise RulesParseError(
            f"multi-area defined name not supported: {name!r} → {areas}",
            subtype="MultiAreaName", name=name, areas=areas,
        )
    if not areas:
        raise RulesParseError(
            f"empty defined name: {name!r}",
            subtype="BadGrammar",
        )
    return areas[0]


# === Per-form cell collectors =============================================

def _classify_cell_at(ws: Any, row: int, col_letter: str, opts: dict[str, Any] | None,
                       merge_lookup: dict[tuple[int, str], tuple[Any, str]],
                       info_sink: list[dict[str, Any]]) -> ClassifiedCell:
    # merge_lookup[(row, col)] = (anchor|None, range_str). anchor=None
    # means *this* cell IS the anchor; otherwise redirect through it.
    cell = ws[f"{col_letter}{row}"]
    anchor_cell, merge_range = merge_lookup.get((row, col_letter), (None, None))

    classify_opts = dict(opts or {})
    classify_opts["merge_range"] = merge_range
    classify_opts["is_anchor_of_merge"] = bool(merge_range and anchor_cell is None)

    if anchor_cell is not None:
        # Non-anchor inside a merge — redirect to the anchor's value.
        info_sink.append({
            "rule_id": "merged-cell-resolution",
            "sheet": ws.title, "row": row, "column": col_letter,
            "merge_range": merge_range, "anchor": anchor_cell.coordinate,
        })
        return classify(anchor_cell, opts=classify_opts)
    return classify(cell, opts=classify_opts)


def _build_merge_lookup(ws: Any) -> dict[tuple[int, str], tuple[Any, str]]:
    lookup: dict[tuple[int, str], tuple[Any, str]] = {}
    for merge in ws.merged_cells.ranges:
        range_str = str(merge)
        anchor = ws.cell(row=merge.min_row, column=merge.min_col)
        for r in range(merge.min_row, merge.max_row + 1):
            for c in range(merge.min_col, merge.max_col + 1):
                col_letter = get_column_letter(c)
                if r == merge.min_row and c == merge.min_col:
                    lookup[(r, col_letter)] = (None, range_str)  # this IS the anchor
                else:
                    lookup[(r, col_letter)] = (anchor, range_str)
    return lookup


def _data_range(ws: Any, header_row: int) -> tuple[int, int]:
    return max(header_row + 1, 1), (ws.max_row or 1)


def _collect_column(ws: Any, col_letter: str, defaults: dict[str, Any] | None,
                     opts: dict[str, Any] | None,
                     merge_lookup: dict[tuple[int, str], tuple[Any, str]],
                     info_sink: list[dict[str, Any]]) -> list[ClassifiedCell]:
    header_row = int((defaults or {}).get("header_row", 1))
    first, last = _data_range(ws, header_row)
    out: list[ClassifiedCell] = []
    for r in range(first, last + 1):
        out.append(_classify_cell_at(ws, r, col_letter, opts, merge_lookup, info_sink))
    return out


# === Top-level dispatch ===================================================

def resolve_scope(scope_node: Any, workbook: Any, defaults: dict[str, Any] | None = None,
                   opts: dict[str, Any] | None = None) -> ScopeResult:
    """Top-level dispatch over the 10 scope-node types (SPEC §4)."""
    defaults = defaults or {}
    opts = opts or {}
    allow_tables = not bool(opts.get("no_table_autodetect", False))
    info_sink: list[dict[str, Any]] = []

    if isinstance(scope_node, CellRef):
        ws = resolve_sheet(scope_node.sheet, workbook)
        col_letter, row = coordinate_from_string(scope_node.ref)
        merge_lookup = _build_merge_lookup(ws)
        c = _classify_cell_at(ws, row, col_letter, opts, merge_lookup, info_sink)
        return ScopeResult(ws.title, [c], info_findings=info_sink)

    if isinstance(scope_node, RangeRef):
        ws = resolve_sheet(scope_node.sheet, workbook)
        min_col, min_row, max_col, max_row = range_boundaries(
            f"{scope_node.start}:{scope_node.end}"
        )
        merge_lookup = _build_merge_lookup(ws)
        cells: list[ClassifiedCell] = []
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                cells.append(_classify_cell_at(
                    ws, r, get_column_letter(c), opts, merge_lookup, info_sink,
                ))
        return ScopeResult(ws.title, cells, info_findings=info_sink)

    if isinstance(scope_node, ColRef):
        ws = resolve_sheet(scope_node.sheet, workbook)
        if scope_node.is_letter:
            col_letter = scope_node.name_or_letter
            is_table = False
        else:
            col_letter, is_table = resolve_header(
                scope_node.name_or_letter, ws, defaults,
                allow_table_fallback=allow_tables,
            )
        merge_lookup = _build_merge_lookup(ws)
        cells = _collect_column(ws, col_letter, defaults, opts, merge_lookup, info_sink)
        return ScopeResult(ws.title, cells, column_letter=col_letter,
                            is_table_resolved=is_table, info_findings=info_sink)

    if isinstance(scope_node, MultiColRef):
        ws = resolve_sheet(scope_node.sheet, workbook)
        merge_lookup = _build_merge_lookup(ws)
        cells = []
        any_table_resolved = False
        for child in scope_node.children:
            if child.is_letter:
                letter = child.name_or_letter
            else:
                letter, used_tbl = resolve_header(
                    child.name_or_letter, ws, defaults,
                    allow_table_fallback=allow_tables,
                )
                any_table_resolved = any_table_resolved or used_tbl
            cells.extend(_collect_column(ws, letter, defaults, opts, merge_lookup, info_sink))
        return ScopeResult(ws.title, cells, is_table_resolved=any_table_resolved,
                            info_findings=info_sink)

    if isinstance(scope_node, RowRef):
        ws = resolve_sheet(scope_node.sheet, workbook)
        header_row = int((defaults or {}).get("header_row", 1))
        merge_lookup = _build_merge_lookup(ws)
        cells = []
        if scope_node.n != header_row:  # exclude header row per SPEC §4
            for c in range(1, (ws.max_column or 1) + 1):
                cells.append(_classify_cell_at(
                    ws, scope_node.n, get_column_letter(c), opts, merge_lookup, info_sink,
                ))
        return ScopeResult(ws.title, cells, info_findings=info_sink)

    if isinstance(scope_node, SheetRef):
        ws = workbook[scope_node.name] if scope_node.name in workbook.sheetnames \
             else resolve_sheet(scope_node.name, workbook)
        header_row = int((defaults or {}).get("header_row", 1))
        first, last = _data_range(ws, header_row)
        merge_lookup = _build_merge_lookup(ws)
        cells = []
        for r in range(first, last + 1):
            for c in range(1, (ws.max_column or 1) + 1):
                cells.append(_classify_cell_at(
                    ws, r, get_column_letter(c), opts, merge_lookup, info_sink,
                ))
        return ScopeResult(ws.title, cells, info_findings=info_sink)

    if isinstance(scope_node, NamedRef):
        area = resolve_named(scope_node.name, workbook)
        # Strip $-anchors and resolve as a RangeRef recursively.
        sheet, ref = parse_sheet_qualifier(area)
        cleaned = ref.replace("$", "")
        if ":" in cleaned:
            start, end = cleaned.split(":", 1)
            return resolve_scope(RangeRef(sheet, start, end), workbook, defaults, opts)
        return resolve_scope(CellRef(sheet, cleaned), workbook, defaults, opts)

    if isinstance(scope_node, TableRef):
        return _resolve_table_ref(scope_node, workbook, defaults, opts)

    raise RulesParseError(
        f"unknown scope node type: {type(scope_node).__name__}",
        subtype="BadGrammar",
    )


def _resolve_table_ref(scope_node: TableRef, workbook: Any,
                        defaults: dict[str, Any] | None,
                        opts: dict[str, Any] | None) -> ScopeResult:
    """`table:T1` (data area) or `table:T1[Col]` (single Table column)."""
    info_sink: list[dict[str, Any]] = []
    for ws_name in workbook.sheetnames:
        ws = workbook[ws_name]
        tables = _excel_tables_for_sheet(ws)
        for table in tables.values():
            tname = getattr(table, "displayName", None) or getattr(table, "name", None)
            if tname != scope_node.name:
                continue
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            merge_lookup = _build_merge_lookup(ws)
            # Header row is row 1 of the Table's range; data starts at row 2.
            data_first = min_row + 1
            data_last = max_row
            if scope_node.column is None:
                cells: list[ClassifiedCell] = []
                for r in range(data_first, data_last + 1):
                    for c in range(min_col, max_col + 1):
                        cells.append(_classify_cell_at(
                            ws, r, get_column_letter(c), opts, merge_lookup, info_sink,
                        ))
                return ScopeResult(ws.title, cells, is_table_resolved=True,
                                    info_findings=info_sink)
            # Single-column form: find the named TableColumn.
            col_names = [tc.name for tc in (table.tableColumns or [])]
            if scope_node.column not in col_names:
                raise HeaderNotFound(
                    f"Table {scope_node.name!r} has no column {scope_node.column!r}",
                    table=scope_node.name, header=scope_node.column,
                    available=col_names[:50],
                )
            offset = col_names.index(scope_node.column)
            col_letter = get_column_letter(min_col + offset)
            cells = []
            for r in range(data_first, data_last + 1):
                cells.append(_classify_cell_at(
                    ws, r, col_letter, opts, merge_lookup, info_sink,
                ))
            return ScopeResult(ws.title, cells, column_letter=col_letter,
                                is_table_resolved=True, info_findings=info_sink)
    raise RulesParseError(
        f"Excel Table not found: {scope_node.name!r}",
        subtype="BadGrammar", name=scope_node.name,
    )
