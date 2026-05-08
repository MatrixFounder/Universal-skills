"""F10 — Workbook output writer (remarks column).

Two write paths:

  - Full-fidelity (`write_remarks`): `openpyxl.load_workbook` + per-cell
    write + `save()`. Preserves comments / drawings / charts / defined
    names on cells NOT modified by xlsx-7 (R8.g round-trip lock).

  - Streaming (`write_remarks_streaming`): M-1 architect-locked dual-
    stream design. Source opened via `load_workbook(read_only=True)`
    iter; destination via `WriteOnlyWorkbook`. Per source row: build
    `[cell_or_remark for col in range(1, max_col+1)]` and `append`.
    The remark-column letter need NOT be rightmost — substitution
    works at any column index ≤ max_col, and for letters past max_col
    the row is extended with empty cells up to the remark index.

Same-path guard (cross-7 H1): `Path.resolve()` follows symlinks. The
TOCTOU window between resolve and open is honest scope (mirrors
xlsx-6).
"""
from __future__ import annotations

import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter

from .exceptions import SelfOverwriteRefused

__all__ = [
    "write_remarks",
    "write_remarks_streaming",
    "allocate_remark_column",
    "apply_remark_mode",
    "apply_pattern_fill",
    "assert_distinct_paths",
]

# PatternFill colours per severity (R8.d). aRGB hex (8 chars: AARRGGBB).
_FILL_ERROR = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE",
                          fill_type="solid")  # red
_FILL_WARNING = PatternFill(start_color="FFFFEB9C", end_color="FFFFEB9C",
                            fill_type="solid")  # yellow
_FILL_INFO = PatternFill(start_color="FFB7DEE8", end_color="FFB7DEE8",
                         fill_type="solid")  # blue


def assert_distinct_paths(input_path: Path, output_path: Path) -> None:
    """cross-7 H1 same-path guard. `Path.resolve()` follows symlinks."""
    if input_path.resolve() == output_path.resolve():
        raise SelfOverwriteRefused(
            f"--output resolves to the same path as input: {input_path.resolve()}",
            subtype="SelfOverwriteRefused",
            input=str(input_path), output=str(output_path),
        )


# === Column allocation =====================================================

def allocate_remark_column(sheet: Any, mode: str | None,
                            explicit: str | None,
                            existing_max_col: int) -> tuple[str, str]:
    """SPEC §8.1 — auto / LETTER / HEADER. Returns `(column_letter, label)`.

    `mode` is `'replace' | 'append' | 'new'` (default `new` from the
    CLI argparse layer). `explicit` is the `--remark-column` argument
    value (`'auto' | LETTER | HEADER`).

    Logic:
      - `explicit == 'auto'` → first free column letter past max_col;
        label = "Remarks".
      - `explicit` is a column letter (≤ 3 upper alpha) → use directly;
        label = sheet's row-1 cell at that letter (if any) else "Remarks".
      - else `explicit` is a header — find an exact match in row 1; if
        found, use that column. If not found, allocate a fresh column
        named `explicit` to the right.
      - For mode `new` AND the chosen letter has existing data in row 1
        (a "Remarks" header or any other), append `_2` (then `_3`, …)
        until a free letter is found.
    """
    if explicit == "auto" or explicit is None:
        new_idx = (existing_max_col or 0) + 1
        return get_column_letter(new_idx), "Remarks"

    # Letter form (e.g. "Z").
    if explicit.isalpha() and explicit.isupper() and 1 <= len(explicit) <= 3:
        # Compute next-free if mode='new' and the column is occupied.
        if mode == "new":
            return _next_free_column(sheet, explicit, existing_max_col)
        return explicit, _row1_value_or(sheet, explicit, "Remarks")

    # Header form: try to find a row-1 match.
    for cell in (sheet[1] if sheet.max_row else []):
        v = (cell.value or "")
        if str(v).strip() == explicit:
            if mode == "new":
                return _next_free_column(sheet, cell.column_letter, existing_max_col)
            return cell.column_letter, explicit
    # Fallback: append a fresh column with the header name.
    new_idx = (existing_max_col or 0) + 1
    return get_column_letter(new_idx), explicit


def _row1_value_or(sheet: Any, letter: str, default: str) -> str:
    try:
        v = sheet[f"{letter}1"].value
    except Exception:  # noqa: BLE001 — defensive on out-of-bounds
        return default
    return str(v).strip() if v is not None else default


def _next_free_column(sheet: Any, base_letter: str,
                       existing_max_col: int) -> tuple[str, str]:
    """For mode='new': if `base_letter` has any row-1 content, walk
    rightward until a fully-empty column is found, then label it
    `<original-header>_2` / `_3` etc. so user data is preserved."""
    base_idx = column_index_from_string(base_letter)
    base_label = _row1_value_or(sheet, base_letter, "Remarks")
    if not _column_has_data(sheet, base_idx):
        return base_letter, base_label
    # Find first free column to the right; label gets _N suffix.
    n = 2
    candidate_idx = base_idx
    while True:
        candidate_idx += 1
        if not _column_has_data(sheet, candidate_idx):
            break
        n += 1
        if candidate_idx > (existing_max_col or 0) + 8:  # defensive cap
            break
    return get_column_letter(candidate_idx), f"{base_label}_{n}"


def _column_has_data(sheet: Any, col_idx: int) -> bool:
    """True if any cell in `col_idx` has a non-empty value."""
    if col_idx > (sheet.max_column or 0):
        return False
    letter = get_column_letter(col_idx)
    for row in range(1, (sheet.max_row or 0) + 1):
        if sheet[f"{letter}{row}"].value not in (None, ""):
            return True
    return False


# === Cell-value mode (replace / append / new) =============================

def apply_remark_mode(existing_value: Any, new_message: str, mode: str) -> str:
    """`replace` overwrites; `append` newline-concats; `new` writes the
    message verbatim (the `_2` suffix is handled at column-allocation
    time so by the time we get here the column is fresh)."""
    if mode == "replace" or existing_value is None or existing_value == "":
        return new_message
    if mode == "append":
        return f"{existing_value}\n{new_message}"
    if mode == "new":
        return new_message
    raise ValueError(f"unknown remark-column-mode: {mode!r}")


def apply_pattern_fill(cell: Any, severity: str) -> None:
    """Set the cell's fill colour per severity (R8.d)."""
    fills = {"error": _FILL_ERROR, "warning": _FILL_WARNING, "info": _FILL_INFO}
    cell.fill = fills.get(severity, _FILL_INFO)


def _format_messages(findings: list[Any]) -> tuple[str, str]:
    """Combine all findings on a cell into a single remark string +
    pick the worst severity (`error > warning > info`)."""
    sev_order = {"error": 3, "warning": 2, "info": 1}
    worst = max(findings, key=lambda f: sev_order.get(f.severity, 0))
    msg = "; ".join(f.message for f in findings)
    return msg, worst.severity


# === Full-fidelity write path =============================================

def write_remarks(input_path: Path, output_path: Path,
                   findings_per_cell: dict[tuple[str, int, str], list[Any]],
                   opts: Any) -> None:
    """Full-fidelity copy + remark column. Comments / drawings / charts
    / defined names on cells NOT touched by xlsx-7 are preserved."""
    assert_distinct_paths(input_path, output_path)
    wb = load_workbook(input_path)

    mode = getattr(opts, "remark_column_mode", "new") or "new"
    explicit = getattr(opts, "remark_column", None)

    for sheet in wb.worksheets:
        sheet_findings = {(r, c): fs
                          for (s, r, c), fs in findings_per_cell.items()
                          if s == sheet.title}
        if not sheet_findings:
            continue  # no findings on this sheet; leave it untouched
        existing_max_col = sheet.max_column or 0
        col_letter, label = allocate_remark_column(
            sheet, mode, explicit, existing_max_col,
        )
        # Write the header label in row 1 if the column is freshly added.
        if column_index_from_string(col_letter) > existing_max_col \
                or _row1_value_or(sheet, col_letter, "") == "":
            sheet[f"{col_letter}1"] = label
        for (row, _col), fs in sheet_findings.items():
            cell_ref = f"{col_letter}{row}"
            existing = sheet[cell_ref].value
            new_msg, severity = _format_messages(fs)
            sheet[cell_ref] = apply_remark_mode(existing, new_msg, mode)
            apply_pattern_fill(sheet[cell_ref], severity)
    wb.save(output_path)


# === Streaming write path (M-1 architect-locked dual-stream) ==============

def write_remarks_streaming(input_path: Path, output_path: Path,
                              findings_per_cell: dict[tuple[str, int, str], list[Any]],
                              opts: Any) -> None:
    """M-1 architect-locked dual-stream: read source `read_only=True`,
    write dest via `WriteOnlyWorkbook`. Single-pass per workbook.

    The remark-column letter need NOT be rightmost — substitution at
    the chosen column index works for any letter ≤ source max_col, and
    for letters past max_col the row is extended with empty cells up
    to the remark index.

    DEP-4 (auto) and DEP-5 (append) are rejected at arg-parse, so by
    the time this function runs we always have an explicit letter and
    a non-`append` mode.
    """
    assert_distinct_paths(input_path, output_path)

    sys.stderr.write(
        "NOTE: --streaming-output mode trades cell-level styles for "
        "single-pass writes; PatternFill on remark column is dropped "
        "(SPEC §11.2 honest scope).\n"
    )

    explicit = getattr(opts, "remark_column", None)
    if explicit is None or explicit == "auto":
        # CLI DEP-4 should have rejected this; guard defensively anyway.
        raise ValueError("--streaming-output requires explicit --remark-column LETTER")
    remark_letter = explicit if (explicit.isalpha() and explicit.isupper()) else "Z"
    remark_col_idx = column_index_from_string(remark_letter)

    # Pre-aggregate findings by (sheet, row) — the remark column carries
    # the union of all findings on that row, regardless of which cell
    # actually fired the rule. This matches the full-fidelity write
    # path (per-row remark) and is the documented xlsx-7 contract.
    by_sheet_row: dict[tuple[str, int], list[Any]] = {}
    for key, fs in findings_per_cell.items():
        by_sheet_row.setdefault(key[:2], []).extend(fs)

    src = load_workbook(input_path, read_only=True, data_only=False)
    dst = Workbook(write_only=True)
    for src_sheet in src.worksheets:
        dst_sheet = dst.create_sheet(title=src_sheet.title)
        src_max_col = src_sheet.max_column or 1
        max_idx = max(src_max_col, remark_col_idx)
        for src_row_idx, src_row in enumerate(
                src_sheet.iter_rows(values_only=False), start=1):
            row_vec: list[Any] = []
            row_findings = by_sheet_row.get((src_sheet.title, src_row_idx), [])
            for col_idx in range(1, max_idx + 1):
                if col_idx == remark_col_idx and row_findings:
                    msg, _sev = _format_messages(row_findings)
                    row_vec.append(msg)
                elif col_idx == remark_col_idx and src_row_idx == 1 \
                        and remark_col_idx > src_max_col:
                    # Fresh remark column past source max_col → write header.
                    row_vec.append("Remarks")
                elif col_idx <= src_max_col and (col_idx - 1) < len(src_row):
                    row_vec.append(src_row[col_idx - 1].value)
                else:
                    row_vec.append(None)
            dst_sheet.append(row_vec)
    src.close()
    dst.save(str(output_path))
