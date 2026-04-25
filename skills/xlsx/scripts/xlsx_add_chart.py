#!/usr/bin/env python3
"""Add a bar / line / pie chart to an existing .xlsx workbook.

The most common follow-up to `csv2xlsx.py` is "and put a chart on top
of it" — this script does that without making the user open Excel /
LibreOffice. It uses openpyxl's chart API (XML, no rasterisation), so
the chart stays editable in the consuming application and re-renders
correctly on theme changes.

Usage:
    xlsx_add_chart.py INPUT.xlsx --type bar|line|pie \\
        --data RANGE [--categories RANGE] [--title TEXT] \\
        [--sheet NAME] [--anchor CELL] [--output OUT.xlsx]

Range syntax: standard Excel notation, e.g. `B2:B11` for a single
series or `B1:D11` for a header-row + 3-column block. When --data
spans multiple columns the first row is treated as series titles.

--categories selects the X-axis labels (typically a column to the
left of --data). If omitted, the chart uses 1..N as default labels.

--anchor is the top-left cell of the chart on the sheet (default:
two columns to the right of the data block). --sheet defaults to
the active sheet. --output defaults to overwriting INPUT.xlsx.

Limitations: one chart per invocation; no secondary axes; no
trendlines. For elaborate dashboards drop down to openpyxl directly.
"""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

from openpyxl import load_workbook  # type: ignore
from openpyxl.chart import BarChart, LineChart, PieChart, Reference  # type: ignore
from openpyxl.utils import column_index_from_string, get_column_letter  # type: ignore

from _errors import add_json_errors_argument, report_error
from office._encryption import EncryptedFileError, assert_not_encrypted
from office._macros import warn_if_macros_will_be_dropped


CHART_TYPES = {"bar": BarChart, "line": LineChart, "pie": PieChart}

RANGE_RE = re.compile(r"^([A-Z]+)(\d+):([A-Z]+)(\d+)$")
CELL_RE = re.compile(r"^([A-Z]+)(\d+)$")


def _parse_range(text: str) -> tuple[int, int, int, int]:
    """Parse 'B2:D11' → (min_col, min_row, max_col, max_row), all 1-based."""
    m = RANGE_RE.match(text.upper().replace("$", ""))
    if not m:
        raise ValueError(
            f"Range '{text}' is not in 'A1:B10' notation. "
            "Use uppercase letters and 1-based row numbers."
        )
    return (
        column_index_from_string(m.group(1)),
        int(m.group(2)),
        column_index_from_string(m.group(3)),
        int(m.group(4)),
    )


def _parse_cell(text: str) -> tuple[int, int]:
    m = CELL_RE.match(text.upper().replace("$", ""))
    if not m:
        raise ValueError(f"Cell '{text}' is not in 'A1' notation.")
    return column_index_from_string(m.group(1)), int(m.group(2))


def _first_cell_is_text(ws, col: int, row: int) -> bool:
    """Used by the titles-from-data auto-detect. A non-numeric value
    in the data range's top-left cell almost certainly means the user
    laid out the data with a header row and expects that header to be
    the series title — not a data point."""
    cell = ws.cell(row=row, column=col)
    if cell.value is None:
        return False
    if isinstance(cell.value, (int, float)):
        return False
    return True


def add_chart(
    input_path: Path,
    output_path: Path,
    *,
    chart_type: str,
    data_range: str,
    categories_range: str | None,
    title: str | None,
    sheet_name: str | None,
    anchor: str | None,
    titles_from_data: bool | None = None,
) -> dict:
    if chart_type not in CHART_TYPES:
        raise ValueError(f"--type must be one of {', '.join(CHART_TYPES)}")

    wb = load_workbook(str(input_path))
    ws = wb[sheet_name] if sheet_name else wb.active

    data_min_col, data_min_row, data_max_col, data_max_row = _parse_range(data_range)

    # Resolve titles_from_data: explicit user flag wins; otherwise
    # auto-detect from the top-left cell. Multi-column data with a
    # text header → titles. Single-column data with a text header →
    # also titles (otherwise the header lands as the first bar).
    if titles_from_data is None:
        titles_from_data = _first_cell_is_text(ws, data_min_col, data_min_row)

    chart = CHART_TYPES[chart_type]()
    if title:
        chart.title = title

    # `from_rows=False` (default) treats each COLUMN as a series.
    data_ref = Reference(
        ws,
        min_col=data_min_col, max_col=data_max_col,
        min_row=data_min_row, max_row=data_max_row,
    )
    chart.add_data(data_ref, titles_from_data=titles_from_data)

    if categories_range:
        cat_min_col, cat_min_row, cat_max_col, cat_max_row = _parse_range(categories_range)
        cats = Reference(
            ws,
            min_col=cat_min_col, max_col=cat_max_col,
            min_row=cat_min_row, max_row=cat_max_row,
        )
        chart.set_categories(cats)

    if anchor:
        # Validate the anchor up front via the existing helper; bad
        # input gives a clear ValueError instead of a confusing
        # openpyxl exception three frames deep.
        _parse_cell(anchor)
        anchor_cell_str = anchor.upper()
    else:
        # Default placement: 2 columns to the right of the data block,
        # at the same top row. If that overflows the sheet (col > XFD),
        # fall back to the row below the data.
        target_col = data_max_col + 2
        if target_col <= 16384:  # XFD = 16384
            anchor_cell_str = f"{get_column_letter(target_col)}{data_min_row}"
        else:
            anchor_cell_str = f"A{data_max_row + 2}"

    # openpyxl wants the anchor as a cell reference string, e.g. "F2".
    ws.add_chart(chart, anchor_cell_str)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))

    return {
        "chart_type": chart_type,
        "sheet": ws.title,
        "data": data_range,
        "categories": categories_range,
        "anchor": anchor_cell_str,
        "output": str(output_path),
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("input", type=Path, help="Source .xlsx")
    parser.add_argument("--type", required=True, choices=list(CHART_TYPES),
                        help="Chart kind: bar, line, or pie.")
    parser.add_argument("--data", required=True,
                        help="Cell range with the values, e.g. 'B2:B11' or 'B1:D11'.")
    parser.add_argument("--categories", default=None,
                        help="Cell range with the X-axis labels (typically the leftmost column).")
    parser.add_argument("--title", default=None, help="Chart title text.")
    parser.add_argument("--sheet", default=None,
                        help="Worksheet name (default: active sheet).")
    parser.add_argument("--anchor", default=None,
                        help="Top-left cell where the chart is placed (default: 2 cols right of --data).")
    titles_group = parser.add_mutually_exclusive_group()
    titles_group.add_argument("--titles-from-data", dest="titles_from_data", action="store_true",
                              default=None,
                              help="Treat the first row of --data as series titles (default: auto-detect from cell type).")
    titles_group.add_argument("--no-titles-from-data", dest="titles_from_data", action="store_false",
                              help="Treat every cell in --data as a value (no header row).")
    parser.add_argument("--output", type=Path, default=None,
                        help="Destination .xlsx (default: overwrite INPUT).")
    add_json_errors_argument(parser)
    args = parser.parse_args(argv)
    je = args.json_errors

    if not args.input.is_file():
        parser.error(f"input not found: {args.input}")
    try:
        assert_not_encrypted(args.input)
    except EncryptedFileError as exc:
        return report_error(
            str(exc), code=3, error_type="EncryptedFileError",
            details={"path": str(args.input)}, json_mode=je,
        )
    output = args.output or args.input
    warn_if_macros_will_be_dropped(args.input, output, sys.stderr)

    try:
        report = add_chart(
            args.input, output,
            chart_type=args.type,
            data_range=args.data,
            categories_range=args.categories,
            title=args.title,
            sheet_name=args.sheet,
            anchor=args.anchor,
            titles_from_data=args.titles_from_data,
        )
    except (ValueError, KeyError) as exc:
        return report_error(
            f"xlsx_add_chart: {exc}",
            code=1, error_type=type(exc).__name__, json_mode=je,
        )

    import json
    print(json.dumps(report, indent=2, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    sys.exit(main())
